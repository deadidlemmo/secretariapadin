import os
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils.dates import extract_te_date_from_text
from utils.excel import set_merged_cell_value
from utils.text import (
    build_colmap,
    is_missing_text,
    is_missing_value,
    norm_header_compact,
    pick_col,
    safe_str,
)


RX_EJA_TRANSFER = re.compile(
    r"(?i)(?<![A-Z0-9])(TE|MC|MCC)\s*[-:\s\u2013\u2014]*\s*(\d{1,2})\s*/\s*(\d{1,2})(?:\s*/\s*(\d{2,4}))?"
)


def label_set(ws, addr: str, label: str, value: str):
    """
    Mantem prefixo do template quando existir, por exemplo:
    'Unidade Escolar: ...'.
    """
    current = ws[addr].value
    value_text = safe_str(value)

    if isinstance(current, str) and ":" in current:
        left = current.split(":", 1)[0].strip()
        if norm_header_compact(left) == norm_header_compact(label):
            set_merged_cell_value(ws, addr, f"{left}: {value_text}")
            return

    set_merged_cell_value(ws, addr, value_text)


def push_missing_info_alert(alert_items, seen_keys, *, turma, nome, ra, tipo, data_str, campo, detalhe):
    """
    Guarda inconsistencias sem duplicar o mesmo caso.
    """
    key = (
        safe_str(turma),
        safe_str(nome),
        safe_str(ra),
        safe_str(tipo),
        safe_str(data_str),
        safe_str(campo),
        safe_str(detalhe),
    )
    if key in seen_keys:
        return
    seen_keys.add(key)

    alert_items.append(
        {
            "turma": safe_str(turma) or "-",
            "nome": safe_str(nome) or "-",
            "ra": safe_str(ra) or "-",
            "tipo": safe_str(tipo) or "-",
            "data": safe_str(data_str) or "-",
            "campo": safe_str(campo) or "-",
            "detalhe": safe_str(detalhe) or "-",
        }
    )


def replace_workbook_sheet(wb, title: str):
    if title in wb.sheetnames:
        wb.remove(wb[title])
    return wb.create_sheet(title)


def add_transfer_alerts_sheet(wb, alert_items):
    if not alert_items:
        return
    ws_alert = replace_workbook_sheet(wb, "ALERTAS")
    ws_alert.append(["Turma", "Nome", "RA", "Tipo", "Data", "Campo", "Detalhe"])
    for item in alert_items:
        ws_alert.append(
            [
                item.get("turma", "-"),
                item.get("nome", "-"),
                item.get("ra", "-"),
                item.get("tipo", "-"),
                item.get("data", "-"),
                item.get("campo", "-"),
                item.get("detalhe", "-"),
            ]
        )
    for col in range(1, 8):
        ws_alert.column_dimensions[get_column_letter(col)].width = 24


def serie_key_from_value(serie_val: str):
    """Extrai 2o/3o/4o/5o de valores como '4oF', '5o D', etc."""
    text = "" if serie_val is None else str(serie_val).strip()
    match = re.search("(?<!\\d)([2345])\\s*[\\u00ba\\u00b0o]", text, flags=re.IGNORECASE)
    if not match:
        return None
    number = match.group(1)
    if number in {"2", "3", "4", "5"}:
        return f"{number}\u00ba"
    return None


def normalize_tipo_te(value) -> str:
    """
    Normaliza o TIPO TE para bater no mapping do modelo.
    Fora disso, retorna o valor original.
    """
    if is_missing_value(value):
        return "Sem Informação"

    raw = str(value).strip()
    normalized = norm_header_compact(raw)

    if "DENTRO" in normalized or "REDEMUNICIPAL" in normalized or "MUNICIPAL" in normalized:
        return "Dentro da Rede"
    if "ESTAD" in normalized:
        return "Rede Estadual"
    if "LITORAL" in normalized or "BAIXADA" in normalized:
        return "Litoral"
    if "MUDANCA" in normalized and "MUNICIP" in normalized:
        return "Mudança de Municipio"
    if "SAOPAULO" in normalized:
        return "São Paulo"
    if "ABCD" in normalized:
        return "ABCD"
    if "INTERIOR" in normalized:
        return "Interior"
    if "OUTROSESTADOS" in normalized or ("OUTROS" in normalized and "ESTAD" in normalized):
        return "Outros Estados"
    if "PARTICULAR" in normalized:
        return "Particular"
    if "PAIS" in normalized:
        return "País"
    if "SEMINFORMA" in normalized:
        return "Sem Informação"

    return raw


class TransferenciasError(ValueError):
    pass


@dataclass
class TransferenciasBuildResult:
    output: BytesIO
    filename: str
    missing_info_alerts: list
    debug: list


def _format_date_br(value) -> str:
    if pd.isna(value):
        return ""
    try:
        parsed = pd.to_datetime(value, errors="coerce")
    except Exception:
        return ""
    if pd.isna(parsed):
        return ""
    return parsed.strftime("%d/%m/%Y")


def _local_te_or_dash(value) -> str:
    text = safe_str(value)
    return "-" if is_missing_text(text) else text


def _append_fundamental_records(
    df_fundamental,
    *,
    period_start,
    period_end,
    transfer_records,
    missing_info_alerts,
    missing_info_seen,
    debug,
) -> int:
    colmap = build_colmap(df_fundamental)

    col_serie = pick_col(colmap, "SERIE", "SÉRIE")
    col_nome = pick_col(colmap, "NOME")
    col_dn = pick_col(colmap, "DATA NASC.", "DATA NASC", "DATANASC")
    col_ra = pick_col(colmap, "RA")
    col_obs = pick_col(colmap, "OBS", "OBSERVACAO", "OBSERVAÇÃO")
    col_local_te = pick_col(colmap, "LOCAL TE", "LOCALTE")

    if not col_nome or not col_dn or not col_ra or not col_serie or not col_obs:
        raise TransferenciasError(
            "A aba 'LISTA CORRIDA' não contém cabeçalhos essenciais "
            "(SÉRIE, NOME, DATA NASC., RA, OBS)."
        )

    debug.append("[quadro_transferencias] Aba lida: LISTA CORRIDA (Fundamental)")
    debug.append(
        f"[quadro_transferencias] Colunas detectadas: "
        f"SÉRIE='{col_serie}', NOME='{col_nome}', DATA NASC.='{col_dn}', "
        f"RA='{col_ra}', OBS='{col_obs}', LOCAL TE='{col_local_te}'"
    )

    if not col_local_te:
        push_missing_info_alert(
            missing_info_alerts,
            missing_info_seen,
            turma="(Estrutura do arquivo)",
            nome="-",
            ra="-",
            tipo="TE",
            data_str="-",
            campo="LOCAL TE",
            detalhe="A coluna 'LOCAL TE' não foi encontrada na Lista Piloto Fundamental.",
        )

    invalid_te_dates = 0
    use_cols = [col_serie, col_nome, col_dn, col_ra, col_obs]
    if col_local_te and col_local_te not in use_cols:
        use_cols.append(col_local_te)

    df_sub = df_fundamental[use_cols].copy()

    for row in df_sub.itertuples(index=False, name=None):
        row_dict = dict(zip(df_sub.columns, row))

        te_dt, te_match_txt, _ = extract_te_date_from_text(
            row_dict.get(col_obs), period_start, period_end
        )

        if te_match_txt and not te_dt:
            invalid_te_dates += 1
            continue

        if not te_dt or not (period_start <= te_dt <= period_end):
            continue

        nome = safe_str(row_dict.get(col_nome))
        ra = safe_str(row_dict.get(col_ra))
        nivel_classe = safe_str(row_dict.get(col_serie))
        local_te_raw = safe_str(row_dict.get(col_local_te)) if col_local_te else ""
        local_te = _local_te_or_dash(local_te_raw)

        if is_missing_text(local_te_raw):
            push_missing_info_alert(
                missing_info_alerts,
                missing_info_seen,
                turma=nivel_classe,
                nome=nome,
                ra=ra,
                tipo="TE",
                data_str=te_dt.strftime("%d/%m/%Y"),
                campo="LOCAL TE",
                detalhe="Registro encontrado no período, mas o campo LOCAL TE está vazio ou inválido.",
            )

        transfer_records.append(
            {
                "nome": nome,
                "dn": _format_date_br(row_dict.get(col_dn)),
                "ra": ra,
                "situacao": "Parcial",
                "breda": "Não",
                "nivel_classe": nivel_classe,
                "tipo": "TE",
                "observacao": local_te,
                "remanejamento": "-",
                "data": te_dt.strftime("%d/%m/%Y"),
            }
        )

    debug.append(f"[quadro_transferencias] TE válidos no período: {len(transfer_records)}")
    debug.append(f"[quadro_transferencias] Datas TE inválidas descartadas: {invalid_te_dates}")
    return invalid_te_dates


def _append_eja_records(
    df_eja,
    *,
    period_start,
    period_end,
    transfer_records,
    missing_info_alerts,
    missing_info_seen,
) -> None:
    colmap_eja = build_colmap(df_eja)
    eja_col_nome = pick_col(colmap_eja, "NOME")
    eja_col_dn = pick_col(colmap_eja, "DATA NASC.", "DATA NASC")
    eja_col_ra = pick_col(colmap_eja, "RA")
    eja_col_serie = pick_col(colmap_eja, "SÉRIE", "SERIE")
    eja_col_obs = pick_col(colmap_eja, "OBS")
    eja_col_local_te = pick_col(colmap_eja, "LOCAL TE", "LOCALTE")

    if not eja_col_local_te:
        push_missing_info_alert(
            missing_info_alerts,
            missing_info_seen,
            turma="(Estrutura do arquivo EJA)",
            nome="-",
            ra="-",
            tipo="TE/MC/MCC",
            data_str="-",
            campo="LOCAL TE",
            detalhe="A coluna 'LOCAL TE' não foi encontrada na Lista Piloto EJA.",
        )

    if not (eja_col_nome and eja_col_ra and eja_col_serie and eja_col_obs):
        return

    df_eja_sub = df_eja[
        [c for c in [eja_col_serie, eja_col_nome, eja_col_dn, eja_col_ra, eja_col_obs, eja_col_local_te] if c]
    ].copy()

    for row in df_eja_sub.itertuples(index=False, name=None):
        row_dict = dict(zip(df_eja_sub.columns, row))
        match = RX_EJA_TRANSFER.search(safe_str(row_dict.get(eja_col_obs)))
        if not match:
            continue

        tipo_str = match.group(1).upper()
        day = int(match.group(2))
        month = int(match.group(3))
        year_raw = match.group(4)

        if year_raw:
            year = int(year_raw)
            if year < 100:
                year += 2000
        else:
            year = period_start.year

        try:
            dt = datetime(year, month, day)
        except Exception:
            continue

        if not (period_start <= dt <= period_end):
            continue

        nome = safe_str(row_dict.get(eja_col_nome))
        ra = safe_str(row_dict.get(eja_col_ra))
        nivel_classe = safe_str(row_dict.get(eja_col_serie))
        local_te_raw = safe_str(row_dict.get(eja_col_local_te)) if eja_col_local_te else ""
        local_te = _local_te_or_dash(local_te_raw)

        if is_missing_text(local_te_raw):
            push_missing_info_alert(
                missing_info_alerts,
                missing_info_seen,
                turma=nivel_classe,
                nome=nome,
                ra=ra,
                tipo=tipo_str,
                data_str=dt.strftime("%d/%m/%Y"),
                campo="LOCAL TE",
                detalhe="Registro encontrado no período, mas o campo LOCAL TE está vazio ou inválido.",
            )

        transfer_records.append(
            {
                "nome": nome,
                "dn": _format_date_br(row_dict.get(eja_col_dn)) if eja_col_dn else "",
                "ra": ra,
                "situacao": "Parcial",
                "breda": "Não",
                "nivel_classe": nivel_classe,
                "tipo": tipo_str,
                "observacao": local_te,
                "remanejamento": "-",
                "data": dt.strftime("%d/%m/%Y"),
            }
        )


def collect_transfer_records(
    fundamental_path,
    period_start,
    period_end,
    *,
    enable_eja=False,
    eja_path=None,
):
    transfer_records = []
    missing_info_alerts = []
    missing_info_seen = set()
    debug = []

    try:
        df_fundamental = pd.read_excel(fundamental_path, sheet_name="LISTA CORRIDA")
    except Exception as exc:
        raise TransferenciasError(f"Erro ao ler a Lista Piloto Fundamental: {exc}") from exc

    _append_fundamental_records(
        df_fundamental,
        period_start=period_start,
        period_end=period_end,
        transfer_records=transfer_records,
        missing_info_alerts=missing_info_alerts,
        missing_info_seen=missing_info_seen,
        debug=debug,
    )

    if enable_eja and eja_path and os.path.exists(eja_path):
        try:
            df_eja = pd.read_excel(eja_path, sheet_name="LISTA CORRIDA")
        except Exception as exc:
            raise TransferenciasError(f"Erro ao ler a Lista Piloto EJA: {exc}") from exc

        _append_eja_records(
            df_eja,
            period_start=period_start,
            period_end=period_end,
            transfer_records=transfer_records,
            missing_info_alerts=missing_info_alerts,
            missing_info_seen=missing_info_seen,
        )

    if not transfer_records:
        raise TransferenciasError("Nenhum registro de TE/MC/MCC encontrado no período especificado.")

    debug.append(f"[quadro_transferencias] Linhas preenchidas no modelo: {len(transfer_records)} (início A12)")
    debug.append(f"[quadro_transferencias] Alertas de falta de informação: {len(missing_info_alerts)}")
    return transfer_records, missing_info_alerts, debug


def build_transferencias_workbook(
    *,
    model_path,
    transfer_records,
    missing_info_alerts,
    responsavel,
    diretor_nome,
    data_quadro_dt,
):
    if not os.path.exists(model_path):
        raise TransferenciasError("Modelo de Quadro Informativo (Transferências) não encontrado.")

    try:
        with open(model_path, "rb") as model_file:
            wb = load_workbook(model_file, data_only=False)
    except Exception as exc:
        raise TransferenciasError(f"Erro ao ler o modelo: {exc}") from exc

    ws = wb.active

    label_set(ws, "A7", "Unidade Escolar", "E.M José Padin Mouta")
    label_set(ws, "A8", "Diretor(a)", diretor_nome)

    set_merged_cell_value(ws, "B9", responsavel)
    set_merged_cell_value(ws, "J9", data_quadro_dt.strftime("%d/%m/%Y"))

    current_row = 12
    for record in transfer_records:
        set_merged_cell_value(ws, f"A{current_row}", record["nome"])
        set_merged_cell_value(ws, f"B{current_row}", record["dn"])
        set_merged_cell_value(ws, f"C{current_row}", record["ra"])
        set_merged_cell_value(ws, f"D{current_row}", record["situacao"])
        set_merged_cell_value(ws, f"E{current_row}", record["breda"])
        set_merged_cell_value(ws, f"F{current_row}", record["nivel_classe"])
        set_merged_cell_value(ws, f"G{current_row}", record["tipo"])
        set_merged_cell_value(ws, f"H{current_row}", record["observacao"])
        set_merged_cell_value(ws, f"I{current_row}", record.get("remanejamento", "-"))
        set_merged_cell_value(ws, f"J{current_row}", record["data"])
        current_row += 1

    add_transfer_alerts_sheet(wb, missing_info_alerts)
    return wb


def build_transferencias_file(
    *,
    fundamental_path,
    model_path,
    period_start,
    period_end,
    responsavel,
    diretor_nome,
    data_quadro_dt,
    enable_eja=False,
    eja_path=None,
):
    transfer_records, missing_info_alerts, debug = collect_transfer_records(
        fundamental_path,
        period_start,
        period_end,
        enable_eja=enable_eja,
        eja_path=eja_path,
    )

    wb = build_transferencias_workbook(
        model_path=model_path,
        transfer_records=transfer_records,
        missing_info_alerts=missing_info_alerts,
        responsavel=responsavel,
        diretor_nome=diretor_nome,
        data_quadro_dt=data_quadro_dt,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Quadro_de_Transferencias_{period_start.strftime('%d%m')}_{period_end.strftime('%d%m')}.xlsx"
    return TransferenciasBuildResult(
        output=output,
        filename=filename,
        missing_info_alerts=missing_info_alerts,
        debug=debug,
    )
