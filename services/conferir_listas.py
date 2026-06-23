import hashlib
import io
import json
import os
import re
import unicodedata
from collections import defaultdict
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from services.confere_escolas import default_confere_school_config


LISTA_STATUS_LABELS = {
    "MA": "ativo",
    "TE": "transferido",
    "REM": "remanejado",
}

SED_TRANSFER_STATUSES = {"BXTR", "TRANSF", "TRANS", "TRAN"}
SED_STATUS_LABELS = {
    "ATIVO": "ativo",
    "REMA": "remanejado",
    **{status: "transferido" for status in SED_TRANSFER_STATUSES},
}
NAME_PARTICLES = {"DA", "DE", "DO", "DAS", "DOS", "E"}

RESULT_COLUMNS = [
    "status_conferencia",
    "categoria",
    "turma_lista",
    "turma_sed",
    "nome_lista",
    "nome_sed",
    "ra_lista",
    "ra_sed",
    "data_nascimento_lista",
    "data_nascimento_sed",
    "situacao_lista",
    "situacao_sed",
    "pdf_origem",
    "observacao",
]

PRINT_COLUMNS = [
    ("turma", "Turma"),
    ("tipo", "Tipo"),
    ("nome_lista", "Aluno na Lista Piloto"),
    ("nome_sed", "Aluno no SED"),
    ("ra_lista", "RA Lista"),
    ("ra_sed", "RA SED"),
    ("data_nascimento_lista", "Nasc. Lista"),
    ("data_nascimento_sed", "Nasc. SED"),
    ("situacao_lista", "Sit. Lista"),
    ("situacao_sed", "Sit. SED"),
    ("observacao", "O que verificar"),
    ("manual", "Correcao / observacao manual"),
]

CATEGORY_SHEETS = [
    ("inconsistencia_situacao", "Inconsistencias"),
    ("nao_encontrado_sed", "Sem SED"),
    ("nao_encontrado_lista", "Sem Lista"),
    ("divergencia_cadastral", "Cadastro"),
    ("ok", "OK"),
]

CATEGORY_LABELS = {
    "ok": "OK",
    "inconsistencia_situacao": "Inconsistencia",
    "nao_encontrado_sed": "Nao encontrado no SED",
    "nao_encontrado_lista": "Nao encontrado na Lista",
    "divergencia_cadastral": "Divergencia cadastral",
}

CATEGORY_COLORS = {
    "ok": "E6F4EA",
    "inconsistencia_situacao": "FCE8E6",
    "nao_encontrado_sed": "FCE8E6",
    "nao_encontrado_lista": "FCE8E6",
    "divergencia_cadastral": "FFF2CC",
}

PENDING_CATEGORIES = {
    "inconsistencia_situacao",
    "nao_encontrado_sed",
    "nao_encontrado_lista",
    "divergencia_cadastral",
}

EXCEL_LISTA_STATUS_LABELS = {
    "MA": "Matricula Ativa",
    "REM": "Remanejado",
    "TE": "Transferido",
}

EXCEL_SED_STATUS_LABELS = {
    "ATIVO": "Ativo",
    "REMA": "Remanejado",
    **{status: "Transferido" for status in SED_TRANSFER_STATUSES},
}

VIEW_LISTA_STATUS_LABELS = {
    "MA": "Ativo",
    "REM": "Remanejado",
    "TE": "Transferido",
}

VIEW_SED_STATUS_LABELS = {
    "ATIVO": "Ativo",
    "REMA": "Remanejado",
    **{status: "Transferido" for status in SED_TRANSFER_STATUSES},
}


def normalize_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


def normalize_name_for_match(value):
    return re.sub(r"\s+", "", normalize_text(value))


def normalize_name_without_particles(value):
    tokens = [token for token in normalize_text(value).split() if token not in NAME_PARTICLES]
    return "".join(tokens)


def names_probably_same(first, second):
    first_match = normalize_name_for_match(first)
    second_match = normalize_name_for_match(second)
    if first_match and second_match and first_match == second_match:
        return True
    first_without_particles = normalize_name_without_particles(first)
    second_without_particles = normalize_name_without_particles(second)
    return bool(
        first_without_particles
        and second_without_particles
        and first_without_particles == second_without_particles
    )


def clean_display(value):
    text = "" if value is None else str(value)
    text = text.replace("\n", " ")
    text = re.sub(r"(\d{2}/\d{2}/)\s+(\d{4})", r"\1\2", text)
    text = re.sub(r"\s+", " ", text).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def normalize_date(value):
    text = clean_display(value)
    if not text:
        return ""
    iso_match = re.search(r"\b(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if iso_match:
        year, month, day = iso_match.groups()
        return f"{int(day):02d}/{int(month):02d}/{int(year):04d}"
    match = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b", text)
    if match:
        day, month, year = match.groups()
        if len(year) == 2:
            year = "20" + year
        return f"{int(day):02d}/{int(month):02d}/{int(year):04d}"
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return text
    return parsed.strftime("%d/%m/%Y")


def normalize_ra(value):
    text = "" if value is None else str(value).upper()
    if text.endswith(".0"):
        text = text[:-2]
    text = re.sub(r"[^0-9X]", "", text)
    return text


def ra_keys(ra, digito=None):
    raw_ra = "" if ra is None else str(ra).upper()
    base = normalize_ra(ra)
    dig = normalize_ra(digito)
    keys = set()
    for value in [base, base + dig if base and dig else ""]:
        if value:
            keys.add(value)
            keys.add(value.lstrip("0") or "0")
    if base and "-" in raw_ra and len(base) > 1:
        base_before_suffix = normalize_ra(raw_ra.split("-", 1)[0])
        if base_before_suffix:
            keys.add(base_before_suffix)
            keys.add(base_before_suffix.lstrip("0") or "0")
        base_without_digit = base[:-1]
        keys.add(base_without_digit)
        keys.add(base_without_digit.lstrip("0") or "0")
    return {key for key in keys if key}


def normalize_turma_lista(value):
    text = normalize_text(value)
    match = re.search(r"\b(\d{1,2})\s*O?\s*([A-Z])\b", text)
    if not match:
        match = re.search(r"\b(\d{1,2})O?([A-Z])\b", text)
    if not match:
        return ""
    return f"{match.group(1)}{match.group(2)}"


def normalize_turma_sed(value):
    text = normalize_text(value)
    match = re.search(r"\b(\d{1,2})\s*ANO\s*([A-Z])(?:I)?\b", text)
    if not match:
        match = re.search(r"\b(\d{1,2})\D+([A-Z])I\b", text)
    if not match:
        return ""
    return f"{match.group(1)}{match.group(2)}"


def format_turma_key(value):
    text = normalize_text(value)
    match = re.fullmatch(r"(\d{1,2})([A-Z])", text)
    if not match:
        return clean_display(value) or "-"
    return f"{match.group(1)}º{match.group(2)}"


def _turma_curta(record):
    return format_turma_key(record.get("turma_key")) if record.get("turma_key") else (record.get("turma") or "-")


def _view_turma_lista_label(value):
    turma_key = normalize_turma_lista(value)
    return format_turma_key(turma_key) if turma_key else (clean_display(value) or "-")


def _view_turma_sed_label(value):
    turma_key = normalize_turma_sed(value)
    return format_turma_key(turma_key) if turma_key else (clean_display(value) or "-")


def _view_lista_status_label(status):
    if not clean_display(status):
        return "-"
    normalized = normalizar_status_lista_piloto(status)
    return VIEW_LISTA_STATUS_LABELS.get(normalized, clean_display(status))


def _view_sed_status_label(status):
    if not clean_display(status):
        return "-"
    normalized = normalizar_status_sed(status)
    return VIEW_SED_STATUS_LABELS.get(normalized, clean_display(status))


def _mesmo_ra(lista_record, sed_record):
    lista_keys = set(lista_record.get("ra_keys", []))
    sed_keys = set(sed_record.get("ra_keys", []))
    return bool(lista_keys and sed_keys and not lista_keys.isdisjoint(sed_keys))


def normalizar_status_lista_piloto(status):
    text = normalize_text(status)
    if text.startswith("REM"):
        return "REM"
    if text.startswith("TE"):
        return "TE"
    if text.startswith("MA"):
        return "MA"
    return text


def normalizar_status_sed(status):
    text = normalize_text(status)
    for candidate in ["TRANSF", "TRANS", "TRAN", "BXTR", "REMA", "ATIVO"]:
        if candidate in text.split() or candidate == text:
            return candidate
    return text


def status_compativel(status_lista, status_sed):
    lista_norm = normalizar_status_lista_piloto(status_lista)
    sed_norm = normalizar_status_sed(status_sed)
    if lista_norm == "MA":
        return sed_norm == "ATIVO"
    if lista_norm == "TE":
        return sed_norm in SED_TRANSFER_STATUSES
    if lista_norm == "REM":
        return sed_norm == "REMA"
    return False


def status_observacao(status_lista, status_sed):
    lista_norm = normalizar_status_lista_piloto(status_lista)
    sed_norm = normalizar_status_sed(status_sed)
    sed_label = SED_STATUS_LABELS.get(sed_norm, sed_norm or "sem status")
    if lista_norm == "MA":
        return f"Aluno ativo na Lista Piloto, mas no SED aparece como {sed_label}."
    if lista_norm == "TE":
        return f"Aluno transferido na Lista Piloto, mas no SED aparece como {sed_label}."
    if lista_norm == "REM":
        return f"Aluno remanejado na Lista Piloto, mas no SED aparece como {sed_label}."
    return f"Status da Lista Piloto nao mapeado: {status_lista or '-'}."


def is_sed_transfer_status(status):
    return normalizar_status_sed(status) in SED_TRANSFER_STATUSES


def is_sed_inactive_status(status):
    return is_sed_transfer_status(status) or normalizar_status_sed(status) == "REMA"


def fallback_key(record):
    return (
        record.get("nome_match_norm") or record.get("nome_norm", ""),
        record.get("data_nascimento_norm", ""),
        record.get("turma_key", ""),
    )


def _pick_by_position(row, index):
    if index >= len(row):
        return ""
    return clean_display(row[index])


def _find_column_index(df, candidates, fallback=None):
    normalized_columns = [(index, normalize_text(column)) for index, column in enumerate(df.columns)]
    normalized_candidates = [normalize_text(candidate) for candidate in candidates]

    for candidate in normalized_candidates:
        for index, column in normalized_columns:
            if column == candidate:
                return index

    for candidate in normalized_candidates:
        for index, column in normalized_columns:
            if column.startswith(candidate) or candidate in column:
                return index

    return fallback


def _lista_piloto_column_indexes(df, school_config=None):
    school_config = school_config or default_confere_school_config()
    if school_config.column_mode == "letters":
        columns = {}
        for field_name, column_letter in school_config.columns.items():
            try:
                columns[field_name] = column_index_from_string(str(column_letter).strip()) - 1
            except Exception as exc:
                raise ValueError(
                    f"Coluna invalida na configuracao da escola {school_config.nome}: "
                    f"{field_name}={column_letter!r}."
                ) from exc
        return columns

    configured_columns = {}
    for field_name, candidates in school_config.columns.items():
        if isinstance(candidates, str):
            candidates = [candidates]
        configured_columns[field_name] = _find_column_index(df, candidates)

    missing_required = [
        field_name
        for field_name in ["turma", "nome", "data_nascimento", "ra", "situacao"]
        if configured_columns.get(field_name) is None
    ]
    if missing_required:
        missing = ", ".join(missing_required)
        raise ValueError(
            f"Campo obrigatorio nao localizado na planilha para {school_config.nome}: {missing}. "
            "Confira se a escola selecionada corresponde ao modelo da Lista Piloto enviada."
        )

    if configured_columns.get("observacoes") is None:
        configured_columns["observacoes"] = -1

    return configured_columns


def _read_lista_dataframe(file_obj, school_config):
    sheet_name = 0 if school_config.sheet_name is None else school_config.sheet_name
    if school_config.column_mode == "letters":
        return pd.read_excel(file_obj, sheet_name=sheet_name, header=None, dtype=str)

    header_row = max(int(school_config.header_row or 1), 1) - 1
    return pd.read_excel(file_obj, sheet_name=sheet_name, header=header_row, dtype=str)


def _normalizar_status_lista_por_escola(value, school_config):
    text = normalize_text(clean_display(value))
    compact_text = text.replace(" ", "")
    for key in (text, compact_text):
        mapped = school_config.status_map.get(key)
        if mapped:
            return mapped
    return normalizar_status_lista_piloto(value)


def _safe_row_value(row, index):
    if index is None or index < 0 or index >= len(row):
        return ""
    return row.iloc[index]


def read_lista_piloto(file_obj, school_config=None):
    school_config = school_config or default_confere_school_config()
    try:
        df = _read_lista_dataframe(file_obj, school_config)
    except Exception as exc:
        sheet_label = school_config.sheet_name if school_config.sheet_name is not None else "primeira aba"
        raise ValueError(
            f"Erro ao ler a Lista Piloto para {school_config.nome} ({sheet_label}): {exc}"
        ) from exc

    if df.empty:
        raise ValueError("A Lista Piloto enviada esta vazia.")

    columns = _lista_piloto_column_indexes(df, school_config)
    data_start_row = int(school_config.data_start_row or 0)
    header_offset = 0 if school_config.column_mode == "letters" else max(int(school_config.header_row or 1), 1)
    records = []
    for idx, row in df.iterrows():
        row_number = int(idx) + 1 if school_config.column_mode == "letters" else int(idx) + header_offset + 1
        if data_start_row and row_number < data_start_row:
            continue

        turma = clean_display(_safe_row_value(row, columns["turma"]))
        nome = clean_display(_safe_row_value(row, columns["nome"]))
        data_nascimento = normalize_date(_safe_row_value(row, columns["data_nascimento"]))
        ra_original = clean_display(_safe_row_value(row, columns["ra"]))
        situacao = _normalizar_status_lista_por_escola(_safe_row_value(row, columns["situacao"]), school_config)
        observacoes = clean_display(_safe_row_value(row, columns.get("observacoes", -1)))

        if not nome or nome in {"0", "#REF#"}:
            continue
        if normalize_text(nome) in {"NOME", "NOME DO ALUNO", "ALUNO"}:
            continue
        if normalize_text(ra_original) in {"RA", "R A"}:
            continue
        if not turma and not ra_original and not situacao:
            continue

        records.append(
            {
                "source": "lista",
                "row_number": row_number,
                "school_id": school_config.id,
                "school_name": school_config.nome,
                "turma": turma,
                "turma_key": normalize_turma_lista(turma),
                "nome": nome,
                "nome_norm": normalize_text(nome),
                "nome_match_norm": normalize_name_for_match(nome),
                "data_nascimento": data_nascimento,
                "data_nascimento_norm": data_nascimento,
                "ra": ra_original,
                "ra_keys": sorted(ra_keys(ra_original)),
                "situacao": situacao,
                "observacoes": observacoes,
            }
        )
    return records


def _extract_turma_from_page(page):
    text = page.extract_text() or ""
    match = re.search(r"Turma:\s*(.+)", text)
    if not match:
        return ""
    return clean_display(match.group(1).splitlines()[0])


def _extract_status_from_sed_cells(*cells):
    combined = normalize_text(" ".join(clean_display(cell) for cell in cells))
    for candidate in ["TRANSF", "TRANS", "TRAN", "BXTR", "REMA", "ATIVO"]:
        if re.search(rf"\b{candidate}\b", combined):
            return candidate
    return combined


def _is_sed_header_row(nome, ra_original, data_nascimento, situacao):
    nome_norm = normalize_text(nome)
    ra_norm = normalize_text(ra_original)
    situacao_norm = normalize_text(situacao)
    data_norm = normalize_text(data_nascimento)
    return (
        nome_norm in {"NOME", "NOME DO ALUNO", "ALUNO"}
        or ra_norm == "RA"
        or data_norm == "DATA DE NASCIMENTO"
        or "DATA MOVIMENTACAO" in situacao_norm
    )


def _is_blank_or_year_residual(value):
    text = normalize_text(value)
    return not text or bool(re.fullmatch(r"\d{4}", text))


def _is_sed_name_continuation_row(
    nome,
    ra_original,
    digito,
    uf_ra,
    data_nascimento,
    data_movimentacao,
    situacao,
):
    nome_norm = normalize_text(nome)
    if not nome_norm:
        return False
    if nome_norm in {"NOME", "NOME DO ALUNO", "ALUNO"}:
        return False
    return (
        not normalize_ra(ra_original)
        and not normalize_ra(digito)
        and not normalize_text(uf_ra)
        and _is_blank_or_year_residual(data_nascimento)
        and _is_blank_or_year_residual(data_movimentacao)
        and _is_blank_or_year_residual(situacao)
    )


def _append_sed_name_continuation(record, continuation):
    continuation_text = clean_display(continuation)
    if not continuation_text:
        return
    current_name = clean_display(record.get("nome"))
    continuation_key = normalize_name_for_match(continuation_text)
    current_key = normalize_name_for_match(current_name)
    if continuation_key and continuation_key in current_key:
        return
    combined_name = clean_display(f"{current_name} {continuation_text}")
    record["nome"] = combined_name
    record["nome_norm"] = normalize_text(combined_name)
    record["nome_match_norm"] = normalize_name_for_match(combined_name)


def extract_sed_pdf_records(pdf_bytes, filename):
    records = []
    current_turma = ""
    pending_page_tail_record = None
    pending_page_tail_number = None
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_number, page in enumerate(pdf.pages, 1):
            page_turma = _extract_turma_from_page(page)
            if page_turma:
                current_turma = page_turma
            turma = page_turma or current_turma
            page_last_record = None
            continuation_target = (
                pending_page_tail_record
                if pending_page_tail_record is not None and pending_page_tail_number == page_number - 1
                else None
            )
            merged_continuation = False
            tables = page.extract_tables(
                table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 3,
                    "join_tolerance": 3,
                    "intersection_tolerance": 5,
                    "text_x_tolerance": 2,
                    "text_y_tolerance": 3,
                }
            )
            for table in tables:
                if not table or len(table) < 2:
                    continue
                for raw_row in table[1:]:
                    row = [clean_display(cell) for cell in raw_row]
                    if len(row) < 9:
                        continue
                    nome = _pick_by_position(row, 2)
                    ra_original = _pick_by_position(row, 3)
                    digito = _pick_by_position(row, 4)
                    uf_ra = _pick_by_position(row, 5)
                    data_nascimento_original = _pick_by_position(row, 6)
                    data_movimentacao_original = _pick_by_position(row, 7)
                    data_nascimento = normalize_date(data_nascimento_original)
                    data_movimentacao = normalize_date(data_movimentacao_original)
                    situacao = _extract_status_from_sed_cells(_pick_by_position(row, 8), data_movimentacao_original)
                    if _is_sed_header_row(nome, ra_original, data_nascimento_original, situacao):
                        continue
                    if _is_sed_name_continuation_row(
                        nome,
                        ra_original,
                        digito,
                        uf_ra,
                        data_nascimento_original,
                        data_movimentacao_original,
                        situacao,
                    ):
                        if continuation_target is not None:
                            _append_sed_name_continuation(continuation_target, nome)
                            merged_continuation = True
                        continue
                    if not nome or not ra_original:
                        continue
                    ra_display = ra_original
                    record_ra_keys = sorted(ra_keys(ra_original, digito))
                    if not record_ra_keys:
                        continue

                    record = {
                        "source": "sed",
                        "page_number": page_number,
                        "turma": turma,
                        "turma_key": normalize_turma_sed(turma),
                        "serie": _pick_by_position(row, 0),
                        "nr": _pick_by_position(row, 1),
                        "nome": nome,
                        "nome_norm": normalize_text(nome),
                        "nome_match_norm": normalize_name_for_match(nome),
                        "ra": ra_display,
                        "digito_ra": digito,
                        "ra_keys": record_ra_keys,
                        "uf_ra": uf_ra,
                        "data_nascimento": data_nascimento,
                        "data_nascimento_norm": data_nascimento,
                        "data_movimentacao": data_movimentacao,
                        "situacao": normalizar_status_sed(situacao),
                        "condicoes": _pick_by_position(row, 9),
                        "transtornos": _pick_by_position(row, 10),
                        "pdf_origem": filename,
                    }
                    records.append(record)
                    page_last_record = record
            if page_last_record is not None:
                pending_page_tail_record = page_last_record
                pending_page_tail_number = page_number
            elif merged_continuation:
                pending_page_tail_record = continuation_target
                pending_page_tail_number = page_number
            else:
                pending_page_tail_record = None
                pending_page_tail_number = None
    if not records:
        raise ValueError("Nenhum aluno foi extraido do PDF.")
    return records


def extract_sed_pdfs(file_items):
    records = []
    errors = []
    successful_files = []
    duplicate_files = []
    file_scopes = []
    seen_hashes = set()

    for file_item in file_items:
        filename = file_item["filename"]
        pdf_bytes = file_item["content"]
        digest = hashlib.sha256(pdf_bytes).hexdigest()
        if digest in seen_hashes:
            duplicate_files.append(filename)
            continue
        seen_hashes.add(digest)
        try:
            extracted = extract_sed_pdf_records(pdf_bytes, filename)
        except Exception as exc:
            errors.append(
                {
                    "arquivo": filename,
                    "erro": f"Nao foi possivel ler o arquivo {filename}. Verifique se o PDF foi gerado corretamente pelo SED. Detalhe: {exc}",
                }
            )
            continue
        records.extend(extracted)
        successful_files.append(filename)
        turma_keys = sorted(
            {record.get("turma_key", "") for record in extracted if record.get("turma_key")}
        )
        file_scopes.append(
            {
                "arquivo": filename,
                "turmas": [format_turma_key(key) for key in turma_keys],
                "turma_keys": turma_keys,
                "alunos": len(extracted),
            }
        )

    return {
        "records": records,
        "errors": errors,
        "successful_files": successful_files,
        "duplicate_files": duplicate_files,
        "file_scopes": file_scopes,
        "total_files": len(file_items),
    }


def preview_sed_pdf_scope(file_items):
    pdf_info = extract_sed_pdfs(file_items)
    turma_keys = sorted(
        {record.get("turma_key", "") for record in pdf_info["records"] if record.get("turma_key")}
    )
    return {
        "turmas": [format_turma_key(key) for key in turma_keys],
        "turma_keys": turma_keys,
        "pdfs_validos": len(pdf_info["successful_files"]),
        "pdfs_ignorados": len(pdf_info["errors"]) + len(pdf_info["duplicate_files"]),
        "pdfs_selecionados": pdf_info["total_files"],
        "errors": pdf_info["errors"],
        "duplicate_files": pdf_info["duplicate_files"],
        "file_scopes": pdf_info["file_scopes"],
    }


def _index_by_ra(records):
    index = defaultdict(list)
    for idx, record in enumerate(records):
        for key in record.get("ra_keys", []):
            index[key].append(idx)
    return index


def _index_by_fallback(records):
    index = defaultdict(list)
    for idx, record in enumerate(records):
        key = fallback_key(record)
        if all(key):
            index[key].append(idx)
    return index


def _index_lista_by_name_birth(records):
    index = defaultdict(list)
    for idx, record in enumerate(records):
        key = (
            record.get("nome_match_norm") or record.get("nome_norm", ""),
            record.get("data_nascimento_norm", ""),
        )
        if all(key):
            index[key].append(idx)
    return index


def _unique_indices(indices):
    unique = []
    seen = set()
    for idx in indices:
        if idx in seen:
            continue
        seen.add(idx)
        unique.append(idx)
    return unique


def _match_score(lista_record, sed_record):
    lista_nome_match = lista_record.get("nome_match_norm") or lista_record.get("nome_norm")
    sed_nome_match = sed_record.get("nome_match_norm") or sed_record.get("nome_norm")
    return (
        bool(lista_record.get("turma_key") and lista_record.get("turma_key") == sed_record.get("turma_key")),
        status_compativel(lista_record.get("situacao"), sed_record.get("situacao")),
        bool(names_probably_same(lista_record.get("nome"), sed_record.get("nome"))),
        bool(lista_nome_match and lista_nome_match == sed_nome_match),
        bool(
            lista_record.get("data_nascimento_norm")
            and lista_record.get("data_nascimento_norm") == sed_record.get("data_nascimento_norm")
        ),
        normalizar_status_sed(sed_record.get("situacao")) == "ATIVO",
    )


def _pick_best_sed_match(lista_record, candidate_indices, sed_records, used_sed):
    available_indices = [idx for idx in _unique_indices(candidate_indices) if idx not in used_sed]
    if not available_indices:
        return None
    return max(available_indices, key=lambda idx: _match_score(lista_record, sed_records[idx]))


def _find_any_sed_candidate(lista_record, sed_records, sed_by_ra, sed_by_fallback):
    candidate_indices = []
    for key in lista_record.get("ra_keys", []):
        candidate_indices.extend(sed_by_ra.get(key, []))
    if candidate_indices:
        return max(_unique_indices(candidate_indices), key=lambda idx: _match_score(lista_record, sed_records[idx]))

    key = fallback_key(lista_record)
    if all(key):
        candidate_indices = sed_by_fallback.get(key, [])
        if candidate_indices:
            return max(_unique_indices(candidate_indices), key=lambda idx: _match_score(lista_record, sed_records[idx]))
    return None


def _find_probable_identity_sed_candidate(lista_record, sed_records, used_sed):
    candidates = []
    lista_birth = lista_record.get("data_nascimento_norm", "")
    lista_turma = lista_record.get("turma_key", "")
    if not lista_birth or not lista_turma:
        return None

    for idx, sed_record in enumerate(sed_records):
        if idx in used_sed:
            continue
        if sed_record.get("data_nascimento_norm") != lista_birth:
            continue
        if sed_record.get("turma_key") != lista_turma:
            continue
        if not names_probably_same(lista_record.get("nome"), sed_record.get("nome")):
            continue
        candidates.append(idx)

    if not candidates:
        return None
    return max(candidates, key=lambda idx: _match_score(lista_record, sed_records[idx]))


def _lista_match_priority(lista_record, sed_records, sed_by_ra, sed_by_fallback):
    candidate_idx = _find_any_sed_candidate(lista_record, sed_records, sed_by_ra, sed_by_fallback)
    if candidate_idx is None:
        return (False, False, False, False, False, False)
    return _match_score(lista_record, sed_records[candidate_idx])


def _find_match(lista_record, sed_records, sed_by_ra, sed_by_fallback, used_sed):
    candidate_indices = []
    for key in lista_record.get("ra_keys", []):
        candidate_indices.extend(sed_by_ra.get(key, []))
    best_idx = _pick_best_sed_match(lista_record, candidate_indices, sed_records, used_sed)
    if best_idx is not None:
        return best_idx, "ra"

    key = fallback_key(lista_record)
    if all(key):
        best_idx = _pick_best_sed_match(lista_record, sed_by_fallback.get(key, []), sed_records, used_sed)
        if best_idx is not None:
            return best_idx, "fallback"
    best_idx = _find_probable_identity_sed_candidate(lista_record, sed_records, used_sed)
    if best_idx is not None:
        return best_idx, "probable_identity"
    return None, None


def _result_row(category, label, lista=None, sed=None, observacao="", campos_divergentes=None):
    lista = lista or {}
    sed = sed or {}
    turma_lista = lista.get("turma", "")
    turma_sed = sed.get("turma", "")
    situacao_lista = lista.get("situacao", "")
    situacao_sed = sed.get("situacao", "")
    return {
        "status_conferencia": label,
        "categoria": category,
        "turma_lista": turma_lista,
        "turma_sed": turma_sed,
        "turma_lista_display": _view_turma_lista_label(turma_lista),
        "turma_sed_display": _view_turma_sed_label(turma_sed),
        "nome_lista": lista.get("nome", ""),
        "nome_sed": sed.get("nome", ""),
        "ra_lista": lista.get("ra", ""),
        "ra_sed": sed.get("ra", ""),
        "data_nascimento_lista": lista.get("data_nascimento", ""),
        "data_nascimento_sed": sed.get("data_nascimento", ""),
        "situacao_lista": situacao_lista,
        "situacao_sed": situacao_sed,
        "situacao_lista_display": _view_lista_status_label(situacao_lista),
        "situacao_sed_display": _view_sed_status_label(situacao_sed),
        "pdf_origem": sed.get("pdf_origem", ""),
        "observacao": observacao,
        "campos_divergentes": sorted(set(campos_divergentes or [])),
    }


def _data_divergence_details(lista, sed, match_method):
    divergences = []
    lista_nome_match = lista.get("nome_match_norm") or lista.get("nome_norm")
    sed_nome_match = sed.get("nome_match_norm") or sed.get("nome_norm")
    if (
        lista_nome_match
        and sed_nome_match
        and lista_nome_match != sed_nome_match
        and not names_probably_same(lista.get("nome"), sed.get("nome"))
    ):
        divergences.append(("nome", "Nome divergente entre as bases."))
    if (
        lista.get("data_nascimento_norm")
        and sed.get("data_nascimento_norm")
        and lista["data_nascimento_norm"] != sed["data_nascimento_norm"]
    ):
        divergences.append(("data_nascimento", "Data de nascimento divergente entre as bases."))
    if lista.get("turma_key") and sed.get("turma_key") and lista["turma_key"] != sed["turma_key"]:
        divergences.append(("turma", "Turma divergente entre as bases."))
    if match_method in {"fallback", "probable_identity"}:
        lista_keys = set(lista.get("ra_keys", []))
        sed_keys = set(sed.get("ra_keys", []))
        if lista_keys and sed_keys and lista_keys.isdisjoint(sed_keys):
            divergences.append(("ra", "RA divergente entre as bases."))
    return divergences


def _data_divergences(lista, sed, match_method):
    return [message for _field, message in _data_divergence_details(lista, sed, match_method)]


def _data_divergence_fields(lista, sed, match_method):
    return [field for field, _message in _data_divergence_details(lista, sed, match_method)]


def _record_divergence_fields(lista_record, sed_record):
    fields = []
    lista_nome_match = lista_record.get("nome_match_norm") or lista_record.get("nome_norm")
    sed_nome_match = sed_record.get("nome_match_norm") or sed_record.get("nome_norm")
    if (
        lista_nome_match
        and sed_nome_match
        and lista_nome_match != sed_nome_match
        and not names_probably_same(lista_record.get("nome"), sed_record.get("nome"))
    ):
        fields.append("nome")
    if (
        lista_record.get("data_nascimento_norm")
        and sed_record.get("data_nascimento_norm")
        and lista_record["data_nascimento_norm"] != sed_record["data_nascimento_norm"]
    ):
        fields.append("data_nascimento")
    if (
        lista_record.get("turma_key")
        and sed_record.get("turma_key")
        and lista_record["turma_key"] != sed_record["turma_key"]
    ):
        fields.append("turma")
    lista_keys = set(lista_record.get("ra_keys", []))
    sed_keys = set(sed_record.get("ra_keys", []))
    if lista_keys and sed_keys and lista_keys.isdisjoint(sed_keys):
        fields.append("ra")
    return fields


def _scope_lista_by_sed_turmas(lista_records, sed_records):
    sed_turma_keys = sorted({record.get("turma_key", "") for record in sed_records if record.get("turma_key")})
    if not sed_turma_keys:
        return lista_records, []

    scoped_records = [record for record in lista_records if record.get("turma_key") in sed_turma_keys]
    return scoped_records, sed_turma_keys


def _dedupe_sed_records(records):
    deduped = []
    seen = set()
    for record in records:
        key = (
            record.get("pdf_origem", ""),
            record.get("turma_key", ""),
            record.get("nome_match_norm") or record.get("nome_norm", ""),
            tuple(record.get("ra_keys", [])),
            record.get("data_nascimento_norm", ""),
            record.get("situacao", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(record)
    return deduped


def _sed_duplicate_observation(records):
    details = []
    seen = set()
    for record in records:
        detail = f"{record.get('turma') or '-'} / {record.get('situacao') or '-'} / {record.get('pdf_origem') or '-'}"
        if detail in seen:
            continue
        seen.add(detail)
        details.append(detail)
    return "Aluno encontrado em mais de uma turma/PDF do SED: " + "; ".join(details) + "."


def _lista_record_already_has_sed_match(lista_record, sed_records, sed_by_ra, sed_by_fallback, used_sed):
    candidate_idx = _find_any_sed_candidate(lista_record, sed_records, sed_by_ra, sed_by_fallback)
    if candidate_idx is None or candidate_idx not in used_sed:
        return None
    return sed_records[candidate_idx]


def _identidade_divergente(lista_record, sed_record):
    lista_nome_match = lista_record.get("nome_match_norm") or lista_record.get("nome_norm")
    sed_nome_match = sed_record.get("nome_match_norm") or sed_record.get("nome_norm")
    nome_divergente = bool(lista_nome_match and sed_nome_match and lista_nome_match != sed_nome_match)
    nascimento_divergente = bool(
        lista_record.get("data_nascimento_norm")
        and sed_record.get("data_nascimento_norm")
        and lista_record["data_nascimento_norm"] != sed_record["data_nascimento_norm"]
    )
    return nome_divergente or nascimento_divergente


def _lista_duplicate_observation(lista_record, sed_record):
    if _mesmo_ra(lista_record, sed_record):
        if _identidade_divergente(lista_record, sed_record):
            return f"Mesmo RA com dados diferentes: Lista {_turma_curta(lista_record)}; SED {_turma_curta(sed_record)}. Confira RA, nome e nascimento."
        return f"Aluno duplicado em turmas diferentes: Lista {_turma_curta(lista_record)}; SED {_turma_curta(sed_record)}. Verifique duplicidade na Lista Piloto."
    return f"Turma divergente: Lista {_turma_curta(lista_record)}; SED {_turma_curta(sed_record)}. Verifique duplicidade na Lista Piloto."


def _lista_has_duplicate_record(lista_record, lista_by_ra, lista_by_name_birth):
    candidate_indices = []
    for key in lista_record.get("ra_keys", []):
        candidate_indices.extend(lista_by_ra.get(key, []))
    if not candidate_indices:
        key = (
            lista_record.get("nome_match_norm") or lista_record.get("nome_norm", ""),
            lista_record.get("data_nascimento_norm", ""),
        )
        if all(key):
            candidate_indices.extend(lista_by_name_birth.get(key, []))
    return len(set(candidate_indices)) > 1


def _find_lista_candidate_for_sed(sed_record, lista_records, lista_by_ra, lista_by_name_birth):
    candidate_indices = []
    for key in sed_record.get("ra_keys", []):
        candidate_indices.extend(lista_by_ra.get(key, []))
    if not candidate_indices:
        key = (
            sed_record.get("nome_match_norm") or sed_record.get("nome_norm", ""),
            sed_record.get("data_nascimento_norm", ""),
        )
        if all(key):
            candidate_indices.extend(lista_by_name_birth.get(key, []))
    if not candidate_indices:
        return None

    return max(
        _unique_indices(candidate_indices),
        key=lambda idx: (
            bool(lista_records[idx].get("turma_key") and lista_records[idx].get("turma_key") == sed_record.get("turma_key")),
            status_compativel(lista_records[idx].get("situacao"), sed_record.get("situacao")),
            normalizar_status_lista_piloto(lista_records[idx].get("situacao")) == "MA",
        ),
    )


def _sed_turma_mismatch_observation(lista_record, sed_record):
    if _mesmo_ra(lista_record, sed_record):
        if _identidade_divergente(lista_record, sed_record):
            return f"Mesmo RA com dados diferentes: Lista {_turma_curta(lista_record)}; SED {_turma_curta(sed_record)}. Confira RA, nome e nascimento."
        return f"Aluno em turma diferente: Lista {_turma_curta(lista_record)}; SED {_turma_curta(sed_record)}. Confira a turma correta."
    return f"Turma divergente: Lista {_turma_curta(lista_record)}; SED {_turma_curta(sed_record)}. Verifique a turma correta."


def _is_sed_duplicate_group_relevant(records):
    statuses = [normalizar_status_sed(record.get("situacao")) for record in records]
    active_count = sum(1 for status in statuses if status == "ATIVO")
    unknown_count = sum(1 for status in statuses if status not in {"ATIVO", "REMA"} and status not in SED_TRANSFER_STATUSES)
    return active_count > 1 or unknown_count > 0


def compare_lista_piloto_sed(lista_records, sed_records, pdf_info):
    sed_records = _dedupe_sed_records(sed_records)
    lista_scope_records, sed_turma_keys = _scope_lista_by_sed_turmas(lista_records, sed_records)
    rows = []
    used_sed = set()
    sed_by_ra = _index_by_ra(sed_records)
    sed_by_fallback = _index_by_fallback(sed_records)
    lista_by_ra = _index_by_ra(lista_records)
    lista_by_name_birth = _index_lista_by_name_birth(lista_records)
    lista_scope_records = sorted(
        lista_scope_records,
        key=lambda record: _lista_match_priority(record, sed_records, sed_by_ra, sed_by_fallback),
        reverse=True,
    )

    for lista in lista_scope_records:
        sed_idx, match_method = _find_match(lista, sed_records, sed_by_ra, sed_by_fallback, used_sed)
        if sed_idx is None:
            matched_sed = _lista_record_already_has_sed_match(lista, sed_records, sed_by_ra, sed_by_fallback, used_sed)
            if matched_sed:
                rows.append(
                    _result_row(
                        "divergencia_cadastral",
                        "Divergencia cadastral",
                        lista=lista,
                        sed=matched_sed,
                        observacao=_lista_duplicate_observation(lista, matched_sed),
                        campos_divergentes=_record_divergence_fields(lista, matched_sed),
                    )
                )
                continue
            rows.append(
                _result_row(
                    "nao_encontrado_sed",
                    "Nao encontrado no SED",
                    lista=lista,
                    observacao="Aluno existe na Lista Piloto, mas nao foi localizado nos PDFs do SED.",
                )
            )
            continue

        sed = sed_records[sed_idx]
        used_sed.add(sed_idx)
        data_divergences = _data_divergences(lista, sed, match_method)
        data_divergence_fields = _data_divergence_fields(lista, sed, match_method)
        compatible = status_compativel(lista.get("situacao"), sed.get("situacao"))

        if not compatible:
            rows.append(
                _result_row(
                    "inconsistencia_situacao",
                    "Inconsistencia de situacao",
                    lista=lista,
                    sed=sed,
                    observacao=" ".join([status_observacao(lista.get("situacao"), sed.get("situacao"))] + data_divergences),
                    campos_divergentes=["situacao", *data_divergence_fields],
                )
            )
        elif data_divergences:
            rows.append(
                _result_row(
                    "divergencia_cadastral",
                    "Divergencia cadastral",
                    lista=lista,
                    sed=sed,
                    observacao=" ".join(data_divergences),
                    campos_divergentes=data_divergence_fields,
                )
            )
        else:
            rows.append(_result_row("ok", "OK", lista=lista, sed=sed, observacao="Dados compativeis."))

    scoped_lista_ids = {id(record) for record in lista_scope_records}
    for lista in lista_records:
        if id(lista) in scoped_lista_ids:
            continue
        sed_idx = _find_any_sed_candidate(lista, sed_records, sed_by_ra, sed_by_fallback)
        if sed_idx is None:
            continue
        sed = sed_records[sed_idx]
        if _lista_has_duplicate_record(lista, lista_by_ra, lista_by_name_birth):
            observacao = _lista_duplicate_observation(lista, sed)
        else:
            observacao = _sed_turma_mismatch_observation(lista, sed)
        rows.append(
            _result_row(
                "divergencia_cadastral",
                "Divergencia cadastral",
                lista=lista,
                sed=sed,
                observacao=observacao,
                campos_divergentes=_record_divergence_fields(lista, sed),
            )
        )
        used_sed.add(sed_idx)

    for idx, sed in enumerate(sed_records):
        if idx in used_sed:
            continue
        if is_sed_inactive_status(sed.get("situacao")):
            continue
        lista_idx = _find_lista_candidate_for_sed(sed, lista_records, lista_by_ra, lista_by_name_birth)
        if lista_idx is not None:
            lista = lista_records[lista_idx]
            rows.append(
                _result_row(
                    "divergencia_cadastral",
                    "Divergencia cadastral",
                    lista=lista,
                    sed=sed,
                    observacao=_sed_turma_mismatch_observation(lista, sed),
                    campos_divergentes=_record_divergence_fields(lista, sed),
                )
            )
            continue
        rows.append(
            _result_row(
                "nao_encontrado_lista",
                "Nao encontrado na Lista Piloto",
                sed=sed,
                observacao="Aluno existe no SED, mas nao foi localizado na Lista Piloto.",
            )
        )

    duplicate_groups = []
    seen_duplicates = set()
    for key, indices in sed_by_ra.items():
        unique_indices = tuple(sorted(set(indices)))
        if len(unique_indices) <= 1 or unique_indices in seen_duplicates:
            continue
        seen_duplicates.add(unique_indices)
        duplicate_records = [sed_records[idx] for idx in unique_indices]
        if not _is_sed_duplicate_group_relevant(duplicate_records):
            continue
        duplicate_groups.append(duplicate_records)
        first = sed_records[unique_indices[0]]
        rows.append(
            _result_row(
                "divergencia_cadastral",
                "Divergencia cadastral",
                sed=first,
                observacao=_sed_duplicate_observation(duplicate_records),
            )
        )

    counts = {category: sum(1 for row in rows if row["categoria"] == category) for category in [
        "ok",
        "inconsistencia_situacao",
        "nao_encontrado_sed",
        "nao_encontrado_lista",
        "divergencia_cadastral",
    ]}
    total_sed_inativos_sem_lista_ignorados = sum(
        1
        for idx, sed in enumerate(sed_records)
        if idx not in used_sed and is_sed_inactive_status(sed.get("situacao"))
    )
    summary = {
        "total_lista_piloto": len(lista_scope_records),
        "total_lista_piloto_geral": len(lista_records),
        "total_sed": len(sed_records),
        "turmas_conferidas": ", ".join(sed_turma_keys) if sed_turma_keys else "Todas",
        "total_pdfs_processados": pdf_info.get("total_files", 0),
        "total_pdfs_lidos": len(pdf_info.get("successful_files", [])),
        "total_pdfs_com_erro": len(pdf_info.get("errors", [])),
        "total_pdfs_duplicados_ignorados": len(pdf_info.get("duplicate_files", [])),
        "total_ok": counts["ok"],
        "total_inconsistencias_situacao": counts["inconsistencia_situacao"],
        "total_nao_encontrados_sed": counts["nao_encontrado_sed"],
        "total_nao_encontrados_lista": counts["nao_encontrado_lista"],
        "total_divergencias_cadastrais": counts["divergencia_cadastral"],
        "total_duplicidades_sed": len(duplicate_groups),
        "total_sed_inativos_sem_lista_ignorados": total_sed_inativos_sem_lista_ignorados,
    }
    return {
        "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "summary": summary,
        "rows": rows,
        "pdf_errors": pdf_info.get("errors", []),
        "successful_files": pdf_info.get("successful_files", []),
        "duplicate_files": pdf_info.get("duplicate_files", []),
    }


def run_conferencia(lista_file, sed_file_items, school_config=None):
    school_config = school_config or default_confere_school_config()
    lista_records = read_lista_piloto(lista_file, school_config=school_config)
    pdf_info = extract_sed_pdfs(sed_file_items)
    result = compare_lista_piloto_sed(lista_records, pdf_info["records"], pdf_info)
    result["school"] = {"id": school_config.id, "nome": school_config.nome}
    return result


def save_result(result, folder, result_id):
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, f"{result_id}.json")
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(result, handle, ensure_ascii=False, indent=2)
    return path


def load_result(folder, result_id):
    path = os.path.join(folder, f"{result_id}.json")
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def prepare_result_for_view(result):
    """Add presentation-only labels without changing the raw comparison values."""
    if not result:
        return result
    for row in result.get("rows", []):
        row["turma_lista_display"] = _view_turma_lista_label(row.get("turma_lista"))
        row["turma_sed_display"] = _view_turma_sed_label(row.get("turma_sed"))
        row["situacao_lista_display"] = _view_lista_status_label(row.get("situacao_lista"))
        row["situacao_sed_display"] = _view_sed_status_label(row.get("situacao_sed"))
    return result


def _result_row_turma_label(row):
    turma_key = normalize_turma_lista(row.get("turma_lista")) or normalize_turma_sed(row.get("turma_sed"))
    return format_turma_key(turma_key) if turma_key else "-"


def _result_sort_key(row):
    category_order = {
        "inconsistencia_situacao": 0,
        "nao_encontrado_sed": 1,
        "nao_encontrado_lista": 2,
        "divergencia_cadastral": 3,
        "ok": 4,
    }
    turma_key = normalize_turma_lista(row.get("turma_lista")) or normalize_turma_sed(row.get("turma_sed"))
    turma_match = re.fullmatch(r"(\d{1,2})([A-Z])", turma_key or "")
    turma_sort = (int(turma_match.group(1)), turma_match.group(2)) if turma_match else (999, turma_key or "")
    name = row.get("nome_lista") or row.get("nome_sed") or ""
    return turma_sort + (category_order.get(row.get("categoria"), 99), normalize_text(name))


def _excel_lista_status_label(status):
    if not clean_display(status):
        return "-"
    normalized = normalizar_status_lista_piloto(status)
    return EXCEL_LISTA_STATUS_LABELS.get(normalized, clean_display(status))


def _excel_sed_status_label(status):
    if not clean_display(status):
        return "-"
    normalized = normalizar_status_sed(status)
    return EXCEL_SED_STATUS_LABELS.get(normalized, clean_display(status))


def _print_row_values(row):
    values = {
        **row,
        "turma": _result_row_turma_label(row),
        "tipo": CATEGORY_LABELS.get(row.get("categoria"), row.get("status_conferencia", "")),
        "situacao_lista": _excel_lista_status_label(row.get("situacao_lista")),
        "situacao_sed": _excel_sed_status_label(row.get("situacao_sed")),
        "manual": "",
    }
    return [values.get(key, "") for key, _label in PRINT_COLUMNS]


def _configure_print_sheet(ws, fit_height=False):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1 if fit_height else 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.print_title_rows = "1:1"
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions


def _style_tabular_sheet(ws, header_fill, header_font, thin, category_column=None):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if category_column:
        for row_number in range(2, ws.max_row + 1):
            category = ws.cell(row_number, category_column).value
            color = CATEGORY_COLORS.get(category)
            if color:
                for cell in ws[row_number]:
                    cell.fill = PatternFill("solid", fgColor=color)


def _add_print_rows(ws, rows):
    ws.append([label for _key, label in PRINT_COLUMNS])
    for row in rows:
        ws.append(_print_row_values(row))


def _set_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _add_summary_sheet(wb, result, title_fill, white_font, header_font, thin):
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False
    summary = result.get("summary", {})
    total_pendencias = (
        summary.get("total_inconsistencias_situacao", 0)
        + summary.get("total_nao_encontrados_sed", 0)
        + summary.get("total_nao_encontrados_lista", 0)
        + summary.get("total_divergencias_cadastrais", 0)
    )

    ws.append(["Conferencia Lista Piloto x SED"])
    ws.merge_cells("A1:D1")
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append(["Gerado em", result.get("generated_at", ""), "Turmas conferidas", summary.get("turmas_conferidas", "")])
    ws.append([])
    ws.append(["Indicador", "Total", "Uso na conferencia", "Aba relacionada"])

    summary_rows = [
        ("Pendencias para revisar", total_pendencias, "Itens que precisam de conferencia manual", "Para imprimir"),
        ("OK", summary.get("total_ok", 0), "Registros compativeis entre Lista Piloto e SED", "OK"),
        ("Inconsistencias de situacao", summary.get("total_inconsistencias_situacao", 0), "Situacao escolar diferente entre as bases", "Inconsistencias"),
        ("Nao encontrados no SED", summary.get("total_nao_encontrados_sed", 0), "Aluno da Lista Piloto nao localizado nos PDFs enviados", "Sem SED"),
        ("Nao encontrados na Lista", summary.get("total_nao_encontrados_lista", 0), "Aluno do SED nao localizado na Lista Piloto", "Sem Lista"),
        ("Divergencias cadastrais", summary.get("total_divergencias_cadastrais", 0), "Nome, RA, nascimento ou turma para conferir", "Cadastro"),
        ("PDFs lidos", f"{summary.get('total_pdfs_lidos', 0)}/{summary.get('total_pdfs_processados', 0)}", "Arquivos do SED usados na conferencia", "Erros PDFs"),
    ]
    for item in summary_rows:
        ws.append(list(item))

    ws.append([])
    ws.append(["Legenda"])
    ws.append(["A aba Para imprimir contem apenas pendencias e uma coluna em branco para anotacao manual."])
    ws.append(["A aba Base completa preserva a exportacao detalhada de auditoria."])

    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor="DDEFEA")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws["A14"].font = Font(bold=True, color="1F2D33")
    ws["A15"].alignment = Alignment(wrap_text=True)
    ws["A16"].alignment = Alignment(wrap_text=True)
    _set_widths(ws, [32, 18, 58, 22])
    _configure_print_sheet(ws, fit_height=True)
    return ws


def build_excel_report(result):
    wb = Workbook()
    title_fill = PatternFill("solid", fgColor="1F6F5F")
    header_fill = PatternFill("solid", fgColor="DDEFEA")
    header_font = Font(bold=True, color="1F2D33")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="C9D8E6")

    rows = sorted(result.get("rows", []), key=_result_sort_key)
    pending_rows = [row for row in rows if row.get("categoria") in PENDING_CATEGORIES]

    _add_summary_sheet(wb, result, title_fill, white_font, header_font, thin)

    ws_print = wb.create_sheet("Para imprimir")
    _add_print_rows(ws_print, pending_rows)
    _style_tabular_sheet(ws_print, header_fill, header_font, thin)
    _set_widths(ws_print, [10, 20, 34, 34, 16, 16, 13, 13, 18, 18, 58, 38])
    for row_number in range(2, ws_print.max_row + 1):
        ws_print.row_dimensions[row_number].height = 54
    _configure_print_sheet(ws_print)

    for category, sheet_name in CATEGORY_SHEETS:
        ws = wb.create_sheet(sheet_name)
        category_rows = [row for row in rows if row.get("categoria") == category]
        _add_print_rows(ws, category_rows)
        _style_tabular_sheet(ws, header_fill, header_font, thin)
        _set_widths(ws, [10, 20, 34, 34, 16, 16, 13, 13, 18, 18, 58, 30])
        _configure_print_sheet(ws)

    ws_errors = wb.create_sheet("Erros PDFs")
    ws_errors.append(["Arquivo", "Ocorrencia"])
    for error in result.get("pdf_errors", []):
        ws_errors.append([error.get("arquivo", ""), error.get("erro", "")])
    for duplicate in result.get("duplicate_files", []):
        ws_errors.append([duplicate, "PDF duplicado ignorado na conferencia."])
    _style_tabular_sheet(ws_errors, header_fill, header_font, thin)
    _set_widths(ws_errors, [38, 90])
    _configure_print_sheet(ws_errors)

    ws_complete = wb.create_sheet("Base completa")
    ws_complete.append(RESULT_COLUMNS)
    for row in rows:
        ws_complete.append([row.get(column, "") for column in RESULT_COLUMNS])
    _style_tabular_sheet(ws_complete, header_fill, header_font, thin, category_column=2)
    _set_widths(ws_complete, [24, 24, 16, 24, 30, 30, 18, 18, 20, 20, 18, 18, 28, 55])
    _configure_print_sheet(ws_complete)

    wb.active = wb.sheetnames.index("Para imprimir") if "Para imprimir" in wb.sheetnames else 0
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
