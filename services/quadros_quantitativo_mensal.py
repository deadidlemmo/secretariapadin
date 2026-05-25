import copy
import os
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from services.quadros_transferencias import normalize_tipo_te, serie_key_from_value
from utils.dates import detect_te_date_from_obs_flexible
from utils.excel import set_merged_cell_value
from utils.text import find_df_col, is_missing_value, safe_str


MESES_BR = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}

DEBUG_HEADERS = [
    "LINHA_ARQUIVO",
    "RM",
    "NOME",
    "SERIE",
    "OBS_ORIGINAL",
    "TE_DATA_EXTRAIDA",
    "ANO_INFERIDO",
    "TRECHO_MATCH",
    "STATUS",
    "MOTIVO",
    "TIPO_TE_RAW",
    "TIPO_TE_NORMALIZADO",
]

QUANTITATIVO_TRANSFERENCIAS_MAPPING = {
    "2º": {
        "Dentro da Rede": "K14",
        "Rede Estadual": "K15",
        "Litoral": "K16",
        "Mudança de Municipio": "K16",
        "São Paulo": "K17",
        "ABCD": "K18",
        "Interior": "K19",
        "Outros Estados": "K20",
        "Particular": "K21",
        "País": "K22",
        "Sem Informação": "K23",
    },
    "3º": {
        "Dentro da Rede": "L14",
        "Rede Estadual": "L15",
        "Litoral": "L16",
        "Mudança de Municipio": "L16",
        "São Paulo": "L17",
        "ABCD": "L18",
        "Interior": "L19",
        "Outros Estados": "L20",
        "Particular": "L21",
        "País": "L22",
        "Sem Informação": "L23",
    },
    "4º": {
        "Dentro da Rede": "M14",
        "Rede Estadual": "M15",
        "Litoral": "M16",
        "Mudança de Municipio": "M16",
        "São Paulo": "M17",
        "ABCD": "M18",
        "Interior": "M19",
        "Outros Estados": "M20",
        "Particular": "M21",
        "País": "M22",
        "Sem Informação": "M23",
    },
    "5º": {
        "Dentro da Rede": "N14",
        "Rede Estadual": "N15",
        "Litoral": "N16",
        "Mudança de Municipio": "N16",
        "São Paulo": "N17",
        "ABCD": "N18",
        "Interior": "N19",
        "Outros Estados": "N20",
        "Particular": "N21",
        "País": "N22",
        "Sem Informação": "N23",
    },
}


class QuantitativoMensalError(ValueError):
    pass


@dataclass
class QuantitativoCountResult:
    counts: dict
    debug_rows: list
    counted: int
    discarded: int


@dataclass
class QuantitativoMensalBuildResult:
    output: BytesIO
    filename: str
    counted: int
    discarded: int
    default_year: int


def get_default_mes_ano(now=None) -> str:
    now = now or datetime.now()
    return f"{MESES_BR[now.month]}/{now.year}"


@contextmanager
def temp_unprotect_sheet(ws):
    original = copy.copy(ws.protection)
    ws.protection.sheet = False
    try:
        yield
    finally:
        ws.protection = original


def recreate_debug_sheet_hidden(wb, title: str = "DEBUG_TE"):
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws_dbg = wb.create_sheet(title)
    ws_dbg.sheet_state = "hidden"
    ws_dbg.append(DEBUG_HEADERS)
    return ws_dbg


def _debug_row(
    *,
    linha_arquivo,
    rm,
    nome,
    serie_val,
    obs_val,
    te_dt,
    year_inferred,
    match_txt,
    status,
    motivo,
    tipo_raw,
    tipo_te,
):
    return [
        linha_arquivo,
        "" if rm is None else str(rm),
        "" if nome is None else str(nome),
        "" if serie_val is None else str(serie_val),
        str(obs_val).strip(),
        "" if not te_dt else te_dt.strftime("%d/%m/%Y"),
        "SIM" if year_inferred else "NAO",
        match_txt,
        status,
        motivo,
        "" if tipo_raw is None else safe_str(tipo_raw),
        "" if tipo_te is None else tipo_te,
    ]


def _add_count(counts, serie_key, tipo_te):
    counts[(serie_key, tipo_te)] = counts.get((serie_key, tipo_te), 0) + 1


def collect_quantitativo_counts(df, period_start, period_end, *, default_year):
    col_rm = find_df_col(df, ["RM"])
    col_nome = find_df_col(df, ["NOME"])
    col_serie = find_df_col(df, ["SÉRIE", "SERIE"])
    col_obs = find_df_col(df, ["OBS"])
    col_tipo_te = find_df_col(df, ["TIPO TE", "TIPO_TE", "TIPO  TE"])

    if not col_serie or not col_obs:
        raise QuantitativoMensalError(
            "Não foi possível localizar as colunas essenciais (SÉRIE e/ou OBS) na LISTA CORRIDA."
        )

    counts = {}
    debug_rows = []
    counted = 0
    discarded = 0

    for index, row in df.iterrows():
        linha_arquivo = int(index) + 2
        rm = row.get(col_rm) if col_rm else None
        nome = row.get(col_nome) if col_nome else None
        serie_val = row.get(col_serie, "")
        obs_val = row.get(col_obs, "")

        if is_missing_value(obs_val):
            continue

        te_dt, _rule, match_txt, year_inferred = detect_te_date_from_obs_flexible(
            obs_val,
            default_year=default_year,
        )

        if not match_txt:
            continue

        tipo_raw = row.get(col_tipo_te, None) if col_tipo_te else None
        tipo_te = "" if col_tipo_te is None else normalize_tipo_te(tipo_raw)

        if not te_dt:
            discarded += 1
            debug_rows.append(
                _debug_row(
                    linha_arquivo=linha_arquivo,
                    rm=rm,
                    nome=nome,
                    serie_val=serie_val,
                    obs_val=obs_val,
                    te_dt=None,
                    year_inferred=year_inferred,
                    match_txt=match_txt,
                    status="SKIPPED",
                    motivo="Data TE inválida em OBS",
                    tipo_raw=tipo_raw,
                    tipo_te=tipo_te,
                )
            )
            continue

        if not (period_start <= te_dt <= period_end):
            discarded += 1
            debug_rows.append(
                _debug_row(
                    linha_arquivo=linha_arquivo,
                    rm=rm,
                    nome=nome,
                    serie_val=serie_val,
                    obs_val=obs_val,
                    te_dt=te_dt,
                    year_inferred=year_inferred,
                    match_txt=match_txt,
                    status="SKIPPED",
                    motivo="Fora do período informado",
                    tipo_raw=tipo_raw,
                    tipo_te=tipo_te,
                )
            )
            continue

        serie_key = serie_key_from_value(serie_val)
        if not serie_key or serie_key not in QUANTITATIVO_TRANSFERENCIAS_MAPPING:
            discarded += 1
            debug_rows.append(
                _debug_row(
                    linha_arquivo=linha_arquivo,
                    rm=rm,
                    nome=nome,
                    serie_val=serie_val,
                    obs_val=obs_val,
                    te_dt=te_dt,
                    year_inferred=year_inferred,
                    match_txt=match_txt,
                    status="SKIPPED",
                    motivo="Série fora de 2º-5º ou ilegível",
                    tipo_raw=tipo_raw,
                    tipo_te=tipo_te,
                )
            )
            continue

        tipo_te = normalize_tipo_te(tipo_raw)
        if tipo_te not in QUANTITATIVO_TRANSFERENCIAS_MAPPING[serie_key]:
            tipo_te = "Sem Informação"

        _add_count(counts, serie_key, tipo_te)
        counted += 1
        debug_rows.append(
            _debug_row(
                linha_arquivo=linha_arquivo,
                rm=rm,
                nome=nome,
                serie_val=serie_val,
                obs_val=obs_val,
                te_dt=te_dt,
                year_inferred=year_inferred,
                match_txt=match_txt,
                status="COUNTED",
                motivo="",
                tipo_raw=tipo_raw,
                tipo_te=tipo_te,
            )
        )

    return QuantitativoCountResult(
        counts=counts,
        debug_rows=debug_rows,
        counted=counted,
        discarded=discarded,
    )


def fill_quantitativo_workbook(
    wb,
    count_result,
    *,
    period_start,
    period_end,
    responsavel,
    mes_ano,
    ano_letivo,
):
    ws = wb.active

    for tipos in QUANTITATIVO_TRANSFERENCIAS_MAPPING.values():
        for cell_addr in tipos.values():
            set_merged_cell_value(ws, cell_addr, 0)

    ws_dbg = recreate_debug_sheet_hidden(wb, "DEBUG_TE")
    for debug_row in count_result.debug_rows:
        ws_dbg.append(debug_row)

    for (serie_key, tipo_te), quantidade in count_result.counts.items():
        cell_addr = QUANTITATIVO_TRANSFERENCIAS_MAPPING[serie_key][tipo_te]
        current_val = ws[cell_addr].value
        current_val = current_val if isinstance(current_val, (int, float)) else 0
        set_merged_cell_value(ws, cell_addr, current_val + quantidade)

    set_merged_cell_value(ws, "B3", str(responsavel).strip())
    set_merged_cell_value(ws, "D3", f"{period_start.strftime('%d/%m/%Y')} a {period_end.strftime('%d/%m/%Y')}")

    with temp_unprotect_sheet(ws):
        set_merged_cell_value(ws, "A6", "E.M José Padin Mouta")
        set_merged_cell_value(ws, "A8", mes_ano)
        set_merged_cell_value(ws, "A10", f"QUADRO GERAL DE TRANSFERENCIAS EXPEDIDAS - {ano_letivo}")

    return wb


def build_quantitativo_mensal_file(
    *,
    fundamental_path,
    model_path,
    period_start,
    period_end,
    responsavel,
    mes_ano,
    default_year,
    ano_letivo,
):
    try:
        df = pd.read_excel(fundamental_path, sheet_name="LISTA CORRIDA")
    except Exception as exc:
        raise QuantitativoMensalError(f"Erro ao ler a Lista Piloto Fundamental: {exc}") from exc

    if not os.path.exists(model_path):
        raise QuantitativoMensalError("Modelo de Quadro Quantitativo Mensal não encontrado.")

    try:
        with open(model_path, "rb") as model_file:
            wb = load_workbook(model_file, data_only=False)
    except Exception as exc:
        raise QuantitativoMensalError(f"Erro ao ler o modelo: {exc}") from exc

    count_result = collect_quantitativo_counts(
        df,
        period_start,
        period_end,
        default_year=default_year,
    )

    fill_quantitativo_workbook(
        wb,
        count_result,
        period_start=period_start,
        period_end=period_end,
        responsavel=responsavel,
        mes_ano=mes_ano,
        ano_letivo=ano_letivo,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Quadro_Quantitativo_Fundamental_{period_start.strftime('%d%m')}_{period_end.strftime('%d%m')}.xlsx"
    return QuantitativoMensalBuildResult(
        output=output,
        filename=filename,
        counted=count_result.counted,
        discarded=count_result.discarded,
        default_year=default_year,
    )
