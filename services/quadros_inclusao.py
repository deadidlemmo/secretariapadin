import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MESES_UPPER_BR = {
    1: "JANEIRO",
    2: "FEVEREIRO",
    3: "MARÇO",
    4: "ABRIL",
    5: "MAIO",
    6: "JUNHO",
    7: "JULHO",
    8: "AGOSTO",
    9: "SETEMBRO",
    10: "OUTUBRO",
    11: "NOVEMBRO",
    12: "DEZEMBRO",
}


class QuantInclusaoError(ValueError):
    pass


@dataclass
class QuantInclusaoBuildResult:
    output: BytesIO
    filename: str
    alerts: list
    plan_without_inclusion_alerts: list


def collapse_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def normalize_rm(value) -> str:
    """
    Normaliza RM para deduplicacao.

    RMs vazios, zero e floats nao inteiros sao tratados como invalidos.
    """
    if value is None:
        return ""

    if isinstance(value, int):
        return "" if value == 0 else str(value).strip()

    if isinstance(value, float):
        if not value.is_integer():
            return ""
        int_value = int(value)
        return "" if int_value == 0 else str(int_value).strip()

    text = collapse_spaces(str(value))
    if not text or text.casefold() in {"nan", "none"}:
        return ""

    digits = re.sub(r"\D+", "", text).lstrip("0")
    return digits or ""


def normalize_turma(value) -> str:
    """
    Normaliza turma para o padrao usado no modelo, como '2oA'.
    """
    if value is None:
        return ""

    text = collapse_spaces(str(value))
    if not text:
        return ""

    text = text.replace("\u00aa", "").replace("\u00ba", "").replace("\u00b0", "")
    text = text.replace("-", "").replace("/", "").replace("\\", "")
    text = text.replace(" ", "").upper()

    match = re.match(r"^(\d{1,2})([A-Z])$", text)
    if not match:
        return ""

    number = int(match.group(1))
    letter = match.group(2)
    return f"{number}\u00ba{letter}"


def is_sim(value) -> bool:
    text = collapse_spaces("" if value is None else str(value))
    return text.casefold() == "sim"


def has_ma(value) -> bool:
    if value is None:
        return False

    text = collapse_spaces(str(value)).upper()
    if not text:
        return False

    tokens = re.findall(r"[A-Z0-9]+", text)
    return "MA" in tokens


def is_valid_prof(value) -> bool:
    if value is None:
        return False

    text = collapse_spaces(str(value))
    if not text or text in {"0", "-"}:
        return False

    return text.casefold() not in {"nan", "none"}


def prof_key(value: str) -> str:
    return collapse_spaces(value).casefold()


def build_template_map(ws_model):
    """
    Mapeia automaticamente as celulas de Inclusao, Plano e Profissionais por turma.
    """
    label_to_qty = {2: 4, 6: 8, 10: 12}  # B->D, F->H, J->L
    turma_cells = {}

    for label_col, qty_col in label_to_qty.items():
        for row_number in range(1, ws_model.max_row + 1):
            value = ws_model.cell(row_number, label_col).value
            if not isinstance(value, str):
                continue

            raw = collapse_spaces(value)
            if not re.match(r"^\d{1,2}\s*-\s*[A-Za-z]$", raw):
                continue

            turma = normalize_turma(raw)
            if not turma:
                continue

            turma_cells[turma] = {
                "inc_qtd": ws_model.cell(row_number, qty_col).coordinate,
                "plano_qtd": ws_model.cell(row_number + 1, qty_col).coordinate,
                "prof_qtd": ws_model.cell(row_number + 2, qty_col).coordinate,
            }

    return turma_cells


def collect_counts_from_lista_corrida(ws_lista, valid_turmas):
    """
    Le LISTA CORRIDA e calcula contadores de inclusao/plano/profissionais por turma.

    Colunas 0-index usadas:
      A turma = 0
      C RM = 2
      D nome = 3
      H situacao = 7, deve conter MA
      N inclusao = 13, deve ser Sim
      P profissional = 15
    """
    inc_rms = defaultdict(set)
    plano_rms = defaultdict(set)
    profs_by_turma = defaultdict(lambda: defaultdict(lambda: {"display": "", "alunos": []}))
    plan_without_inclusion_by_turma = defaultdict(list)
    plan_without_inclusion_seen = defaultdict(set)

    for row in ws_lista.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        turma = normalize_turma(row[0] if len(row) > 0 else None)
        if not turma or (valid_turmas and turma not in valid_turmas):
            continue

        rm = normalize_rm(row[2] if len(row) > 2 else None)
        if not rm:
            continue

        if not has_ma(row[7] if len(row) > 7 else None):
            continue

        nome = collapse_spaces(str(row[3])) if len(row) > 3 and row[3] is not None else ""
        is_inclusao = is_sim(row[13] if len(row) > 13 else None)

        if is_inclusao:
            inc_rms[turma].add(rm)

        prof_value = row[15] if len(row) > 15 else None
        has_prof = is_valid_prof(prof_value)

        if has_prof and not is_inclusao and rm not in plan_without_inclusion_seen[turma]:
            plan_without_inclusion_seen[turma].add(rm)
            plan_without_inclusion_by_turma[turma].append(
                {
                    "rm": rm,
                    "nome": nome,
                    "profissional": collapse_spaces(str(prof_value)),
                }
            )

        if is_inclusao and has_prof:
            plano_rms[turma].add(rm)

            display = collapse_spaces(str(prof_value))
            bucket = profs_by_turma[turma][prof_key(display)]
            if not bucket["display"]:
                bucket["display"] = display
            bucket["alunos"].append((rm, nome))

    inc_counts = {turma: len(rms) for turma, rms in inc_rms.items()}
    plano_counts = {turma: len(rms) for turma, rms in plano_rms.items()}

    return inc_counts, plano_counts, profs_by_turma, plan_without_inclusion_by_turma


def turma_sort_key(turma: str):
    try:
        number, letter = turma.split("\u00ba", 1)
        return (int(number), letter)
    except Exception:
        return (999, turma)


def build_multi_prof_alerts(profs_by_turma, valid_turmas):
    alerts = []
    for turma in sorted(valid_turmas, key=turma_sort_key):
        prof_dict = profs_by_turma.get(turma, {})
        if len(prof_dict) < 2:
            continue

        prof_names = sorted(
            [prof_dict[key]["display"] for key in prof_dict.keys()],
            key=lambda text: text.casefold(),
        )

        audit = []
        for key in sorted(prof_dict.keys(), key=lambda item: prof_dict[item]["display"].casefold()):
            students = prof_dict[key]["alunos"][:10]
            audit.append(
                {
                    "profissional": prof_dict[key]["display"],
                    "amostra_alunos": [{"rm": rm, "nome": nome} for rm, nome in students],
                }
            )

        alerts.append(
            {
                "turma": turma,
                "qtd_profissionais": len(prof_dict),
                "profissionais": prof_names,
                "auditoria": audit,
            }
        )

    return alerts


def build_plan_without_inclusion_alerts(plan_without_inclusion_by_turma, valid_turmas, max_students=20):
    alerts = []
    for turma in sorted(valid_turmas, key=turma_sort_key):
        cases = plan_without_inclusion_by_turma.get(turma, [])
        if not cases:
            continue

        alerts.append(
            {
                "turma": turma,
                "qtd_casos": len(cases),
                "alunos": cases[:max_students],
            }
        )

    return alerts


def replace_workbook_sheet(wb, title: str):
    if title in wb.sheetnames:
        wb.remove(wb[title])
    return wb.create_sheet(title)


def add_quant_inclusao_alerts_sheet(wb, multi_prof_alerts, plan_without_inclusion_alerts):
    if not multi_prof_alerts and not plan_without_inclusion_alerts:
        return

    ws_alert = replace_workbook_sheet(wb, "ALERTAS")
    ws_alert.append(["Categoria", "Turma", "Detalhe", "RM", "Nome"])

    for alert in multi_prof_alerts:
        turma = alert.get("turma", "-")
        profissionais = ", ".join(alert.get("profissionais", []))
        ws_alert.append(
            [
                "M\u00faltiplos profissionais",
                turma,
                f"{alert.get('qtd_profissionais', 0)} profissionais: {profissionais}",
                "",
                "",
            ]
        )
        for block in alert.get("auditoria", []):
            profissional = block.get("profissional", "-")
            for aluno in block.get("amostra_alunos", []):
                ws_alert.append(
                    [
                        "Amostra por profissional",
                        turma,
                        profissional,
                        aluno.get("rm", ""),
                        aluno.get("nome", ""),
                    ]
                )

    for alert in plan_without_inclusion_alerts:
        turma = alert.get("turma", "-")
        ws_alert.append(
            [
                "Plano sem inclus\u00e3o",
                turma,
                f"{alert.get('qtd_casos', 0)} caso(s)",
                "",
                "",
            ]
        )
        for aluno in alert.get("alunos", []):
            ws_alert.append(
                [
                    "Plano sem inclus\u00e3o",
                    turma,
                    aluno.get("profissional", "-"),
                    aluno.get("rm", ""),
                    aluno.get("nome", ""),
                ]
            )

    widths = [26, 14, 42, 14, 36]
    for index, width in enumerate(widths, start=1):
        ws_alert.column_dimensions[get_column_letter(index)].width = width


def get_quant_inclusao_mes_ano(now=None):
    now = now or datetime.now()
    return f"{MESES_UPPER_BR[now.month]}/{now.year}"


def apply_quant_inclusao_header(ws_model, responsavel, now=None):
    now = now or datetime.now()
    mes_ano = get_quant_inclusao_mes_ano(now)

    try:
        b4 = ws_model["B4"].value or ""
        b4s = str(b4)
        if re.search(r"MÊS\s*/\s*\d{4}", b4s, flags=re.IGNORECASE):
            ws_model["B4"] = re.sub(r"MÊS\s*/\s*\d{4}", mes_ano, b4s, flags=re.IGNORECASE)
        else:
            ws_model["B4"] = b4s if mes_ano in b4s else f"{b4s} - {mes_ano}".strip(" -")
    except Exception:
        pass

    ws_model["C8"] = responsavel.strip()
    ws_model["K8"] = now.strftime("%d/%m/%Y")


def fill_quant_inclusao_workbook(wb_model, ws_lista_reg, responsavel, now=None):
    ws_model = wb_model.active
    template_map = build_template_map(ws_model)
    valid_turmas = set(template_map.keys())

    (
        inc_counts,
        plano_counts,
        profs_by_turma,
        plan_without_inclusion_by_turma,
    ) = collect_counts_from_lista_corrida(ws_lista_reg, valid_turmas)

    for turma, cells in template_map.items():
        inc = inc_counts.get(turma, 0)
        plano = plano_counts.get(turma, 0)
        profs = len(profs_by_turma.get(turma, {}))

        ws_model[cells["inc_qtd"]] = inc
        ws_model[cells["plano_qtd"]] = plano
        ws_model[cells["prof_qtd"]] = profs

    now = now or datetime.now()
    apply_quant_inclusao_header(ws_model, responsavel, now)

    alerts = build_multi_prof_alerts(profs_by_turma, valid_turmas)
    plan_without_inclusion_alerts = build_plan_without_inclusion_alerts(
        plan_without_inclusion_by_turma,
        valid_turmas,
    )

    add_quant_inclusao_alerts_sheet(wb_model, alerts, plan_without_inclusion_alerts)

    return alerts, plan_without_inclusion_alerts


def build_quant_inclusao_file(lista_path, model_path, responsavel, now=None):
    wb_reg = None
    try:
        wb_reg = load_workbook(lista_path, data_only=True, read_only=True)
        ws_lista_reg = wb_reg["LISTA CORRIDA"]
    except Exception as exc:
        raise QuantInclusaoError(f"Erro ao ler o arquivo: {exc}") from exc

    try:
        wb_model = load_workbook(model_path, data_only=False)
    except Exception as exc:
        raise QuantInclusaoError(f"Erro ao abrir o modelo de inclusão: {exc}") from exc

    now = now or datetime.now()
    try:
        alerts, plan_without_inclusion_alerts = fill_quant_inclusao_workbook(
            wb_model,
            ws_lista_reg,
            responsavel,
            now,
        )
    finally:
        if wb_reg is not None:
            wb_reg.close()

    output = BytesIO()
    wb_model.save(output)
    output.seek(0)

    filename = f"Quadro_Quantitativo_de_Inclusao_{now.strftime('%d%m%Y')}.xlsx"
    return QuantInclusaoBuildResult(
        output=output,
        filename=filename,
        alerts=alerts,
        plan_without_inclusion_alerts=plan_without_inclusion_alerts,
    )
