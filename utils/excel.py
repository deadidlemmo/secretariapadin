from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter


def set_merged_cell_value(ws, cell_coord, value):
    """
    Atualiza o valor de uma celula mesclada em uma planilha openpyxl
    preservando a mesclagem.
    """
    cell = ws[cell_coord]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                range_str = str(merged_range)
                ws.unmerge_cells(range_str)
                min_col, min_row, _, _ = merged_range.bounds
                top_left_coord = f"{get_column_letter(min_col)}{min_row}"
                ws[top_left_coord] = value
                ws.merge_cells(range_str)
                return
    ws[cell_coord] = value
