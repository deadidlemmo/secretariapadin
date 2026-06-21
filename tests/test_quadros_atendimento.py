import unittest
from datetime import datetime
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from services.quadros_atendimento import (
    build_atendimento_mensal_file,
    extract_by_cols,
    extract_by_fallback_block,
    fill_eja_block,
    normalize_mes_ref,
    safe_int,
    write_block,
    write_header,
    write_turno_totals,
    zero_eja_block,
)


class QuadrosAtendimentoServiceTests(unittest.TestCase):
    def test_normalize_mes_ref_accepts_current_formats(self):
        fallback_now = datetime(2026, 5, 25)

        self.assertEqual(normalize_mes_ref("2026-03", now=fallback_now), "03/2026")
        self.assertEqual(normalize_mes_ref("04/2026", now=fallback_now), "04/2026")
        self.assertEqual(normalize_mes_ref("", now=fallback_now), "05/2026")
        self.assertEqual(normalize_mes_ref("invalido", now=fallback_now), "05/2026")

    def test_safe_int_keeps_existing_conversion_rules(self):
        self.assertEqual(safe_int("1.234"), 1234)
        self.assertEqual(safe_int("12,5"), 12)
        self.assertEqual(safe_int(True), 1)
        self.assertEqual(safe_int(None, default=7), 7)

    def test_extract_by_cols_sums_duplicate_turmas(self):
        wb = Workbook()
        ws = wb.active
        ws["G1"] = "MATRICULAS"
        ws.cell(row=2, column=3).value = "2\u00ba"
        ws.cell(row=2, column=4).value = "A"
        ws.cell(row=2, column=7).value = 10
        ws.cell(row=2, column=8).value = 11
        ws.cell(row=3, column=3).value = "2\u00ba"
        ws.cell(row=3, column=4).value = "A"
        ws.cell(row=3, column=7).value = 1
        ws.cell(row=3, column=8).value = 2
        ws.cell(row=4, column=3).value = "2o"
        ws.cell(row=4, column=4).value = "B"
        ws.cell(row=4, column=7).value = "3"
        ws.cell(row=4, column=8).value = "4"

        debug_log = []
        data = extract_by_cols(ws, "2\u00ba", debug_log)

        self.assertEqual(data["A"], (11, 13))
        self.assertEqual(data["B"], (3, 4))
        self.assertTrue(any("duplicadas" in item for item in debug_log))

    def test_extract_by_fallback_block_stops_when_series_changes(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=6, column=3).value = "2\u00ba"
        ws.cell(row=6, column=4).value = "A"
        ws.cell(row=6, column=7).value = 5
        ws.cell(row=6, column=8).value = 6
        ws.cell(row=7, column=3).value = "2\u00ba"
        ws.cell(row=7, column=4).value = "B"
        ws.cell(row=7, column=7).value = 7
        ws.cell(row=7, column=8).value = 8
        ws.cell(row=8, column=3).value = "3\u00ba"
        ws.cell(row=8, column=4).value = "C"

        debug_log = []
        data = extract_by_fallback_block(ws, "2\u00ba", debug_log)

        self.assertEqual(data, {"A": (5, 6), "B": (7, 8)})
        self.assertTrue(any("fallback usado" in item for item in debug_log))

    def test_write_header_block_and_turno_totals(self):
        wb_model = Workbook()
        ws_model = wb_model.active
        wb_total = Workbook()
        ws_total = wb_total.active
        ws_total.cell(row=38, column=9).value = 120
        ws_total.cell(row=40, column=9).value = 130

        debug_log = []
        write_header(ws_model, "Responsavel", "123", "05/2026", debug_log)
        write_block(ws_model, "2\u00ba", {"A": (10, 11), "C": (12, 13)}, debug_log)
        write_turno_totals(ws_model, ws_total, debug_log)

        self.assertEqual(ws_model["B5"].value, "E.M Jos\u00e9 Padin Mouta")
        self.assertEqual(ws_model["C6"].value, "Responsavel")
        self.assertEqual(ws_model["A13"].value, "05/2026")
        self.assertEqual(ws_model["B37"].value, 10)
        self.assertEqual(ws_model["C37"].value, 11)
        self.assertEqual(ws_model["D37"].value, "=B37+C37")
        self.assertEqual(ws_model["B39"].value, 12)
        self.assertEqual(ws_model["R20"].value, 120)
        self.assertEqual(ws_model["R28"].value, 130)

    def test_eja_helpers_zero_and_fill_expected_cells(self):
        wb_model = Workbook()
        ws_model = wb_model.active

        zero_eja_block(ws_model)
        self.assertEqual(ws_model["L19"].value, 0)
        self.assertEqual(ws_model["R24"].value, "-")

        wb_eja = Workbook()
        ws_eja = wb_eja.active
        ws_eja.cell(row=6, column=5).value = 21
        ws_eja.cell(row=6, column=6).value = 22
        ws_eja.cell(row=20, column=7).value = 99

        fill_eja_block(ws_model, ws_eja)

        self.assertEqual(ws_model["L19"].value, 21)
        self.assertEqual(ws_model["M19"].value, 22)
        self.assertEqual(ws_model["R32"].value, 99)
        self.assertEqual(ws_model["R24"].value, "-")

    def test_build_atendimento_mensal_file_generates_workbook(self):
        with TemporaryDirectory() as tmpdir:
            model_path = f"{tmpdir}/modelo.xlsx"
            lista_path = f"{tmpdir}/lista.xlsx"

            wb_model = Workbook()
            wb_model.save(model_path)

            wb_lista = Workbook()
            ws_total = wb_lista.active
            ws_total.title = "Total de Alunos"
            ws_total["G1"] = "MATRICULAS"
            ws_total.cell(row=2, column=3).value = "2\u00ba"
            ws_total.cell(row=2, column=4).value = "A"
            ws_total.cell(row=2, column=7).value = 10
            ws_total.cell(row=2, column=8).value = 11
            ws_total.cell(row=38, column=9).value = 120
            ws_total.cell(row=40, column=9).value = 130
            wb_lista.save(lista_path)

            result = build_atendimento_mensal_file(
                fundamental_path=lista_path,
                model_path=model_path,
                responsavel="Responsavel",
                rf="123",
                mes_ref="05/2026",
                enable_eja=False,
                now=datetime(2026, 5, 25),
            )

            self.assertEqual(result.filename, "Quadro de Atendimento Mensal - 2505.xlsx")
            self.assertEqual(result.warnings, [])
            self.assertTrue(any("[EJA] desativada" in item for item in result.debug_log))

            generated = load_workbook(result.output, data_only=False)
            ws = generated.active
            self.assertEqual(ws["B5"].value, "E.M Jos\u00e9 Padin Mouta")
            self.assertEqual(ws["C6"].value, "Responsavel")
            self.assertEqual(ws["B7"].value, "123")
            self.assertEqual(ws["A13"].value, "05/2026")
            self.assertEqual(ws["B37"].value, 10)
            self.assertEqual(ws["C37"].value, 11)
            self.assertEqual(ws["R20"].value, 120)
            self.assertEqual(ws["R28"].value, 130)
            self.assertEqual(ws["L19"].value, 0)

    def test_build_atendimento_mensal_file_warns_when_eja_sheet_missing(self):
        with TemporaryDirectory() as tmpdir:
            model_path = f"{tmpdir}/modelo.xlsx"
            lista_path = f"{tmpdir}/lista.xlsx"
            eja_path = f"{tmpdir}/eja.xlsx"

            Workbook().save(model_path)

            wb_lista = Workbook()
            ws_total = wb_lista.active
            ws_total.title = "Total de Alunos"
            ws_total.cell(row=38, column=9).value = 120
            ws_total.cell(row=40, column=9).value = 130
            wb_lista.save(lista_path)

            wb_eja = Workbook()
            wb_eja.active.title = "Outra Aba"
            wb_eja.save(eja_path)

            result = build_atendimento_mensal_file(
                fundamental_path=lista_path,
                model_path=model_path,
                responsavel="Responsavel",
                rf="123",
                mes_ref="05/2026",
                enable_eja=True,
                eja_path=eja_path,
                now=datetime(2026, 5, 25),
            )

            self.assertEqual(len(result.warnings), 1)
            self.assertIn("Total de Alunos", result.warnings[0])
            generated = load_workbook(result.output, data_only=False)
            self.assertEqual(generated.active["L19"].value, 0)


if __name__ == "__main__":
    unittest.main()
