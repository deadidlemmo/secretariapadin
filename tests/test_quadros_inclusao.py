import unittest
from datetime import datetime
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from services.quadros_inclusao import (
    add_quant_inclusao_alerts_sheet,
    build_multi_prof_alerts,
    build_quant_inclusao_file,
    build_plan_without_inclusion_alerts,
    build_template_map,
    collect_counts_from_lista_corrida,
    fill_quant_inclusao_workbook,
    get_quant_inclusao_mes_ano,
    has_ma,
    normalize_rm,
    normalize_turma,
)


def make_lista_row(turma="", rm="", nome="", situacao="", inclusao="", profissional=""):
    row = [""] * 16
    row[0] = turma
    row[2] = rm
    row[3] = nome
    row[7] = situacao
    row[13] = inclusao
    row[15] = profissional
    return row


class QuadrosInclusaoServiceTests(unittest.TestCase):
    def test_normalizers_match_current_rules(self):
        self.assertEqual(normalize_rm("000123-4"), "1234")
        self.assertEqual(normalize_rm(123.0), "123")
        self.assertEqual(normalize_rm(12.5), "")
        self.assertEqual(normalize_rm("0"), "")

        self.assertEqual(normalize_turma("2-A"), "2\u00baA")
        self.assertEqual(normalize_turma("5 D"), "5\u00baD")
        self.assertEqual(normalize_turma("sem turma"), "")

        self.assertTrue(has_ma("MA/Transferido"))
        self.assertTrue(has_ma("Ativo - MA"))
        self.assertFalse(has_ma("MARIA"))

    def test_build_template_map_detects_model_cells(self):
        wb = Workbook()
        ws = wb.active
        ws["B13"] = "2-A"
        ws["F20"] = "3-C"

        template_map = build_template_map(ws)

        self.assertEqual(template_map["2\u00baA"]["inc_qtd"], "D13")
        self.assertEqual(template_map["2\u00baA"]["plano_qtd"], "D14")
        self.assertEqual(template_map["2\u00baA"]["prof_qtd"], "D15")
        self.assertEqual(template_map["3\u00baC"]["inc_qtd"], "H20")

    def test_collect_counts_deduplicates_and_builds_alert_sources(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["header"] * 16)
        ws.append(make_lista_row("2-A", "001", "Aluno Um", "MA", "Sim", "Prof A"))
        ws.append(make_lista_row("2-A", "001", "Aluno Um Duplicado", "MA", "Sim", "Prof A"))
        ws.append(make_lista_row("2-A", "002", "Aluno Dois", "MA", "Sim", "Prof B"))
        ws.append(make_lista_row("2-A", "003", "Aluno Tres", "MA", "", "Prof C"))
        ws.append(make_lista_row("2-A", "004", "Aluno Quatro", "MARIA", "Sim", "Prof D"))
        ws.append(make_lista_row("2-A", "0", "Aluno Zero", "MA", "Sim", "Prof E"))

        inc_counts, plano_counts, profs_by_turma, plan_without = collect_counts_from_lista_corrida(
            ws,
            {"2\u00baA"},
        )

        self.assertEqual(inc_counts["2\u00baA"], 2)
        self.assertEqual(plano_counts["2\u00baA"], 2)
        self.assertEqual(len(profs_by_turma["2\u00baA"]), 2)
        self.assertEqual(len(plan_without["2\u00baA"]), 1)
        self.assertEqual(plan_without["2\u00baA"][0]["rm"], "3")

    def test_alert_builders_and_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["header"] * 16)
        ws.append(make_lista_row("2-A", "001", "Aluno Um", "MA", "Sim", "Prof A"))
        ws.append(make_lista_row("2-A", "002", "Aluno Dois", "MA", "Sim", "Prof B"))
        ws.append(make_lista_row("2-A", "003", "Aluno Tres", "MA", "", "Prof C"))

        _, _, profs_by_turma, plan_without = collect_counts_from_lista_corrida(ws, {"2\u00baA"})
        multi_prof_alerts = build_multi_prof_alerts(profs_by_turma, {"2\u00baA"})
        plan_alerts = build_plan_without_inclusion_alerts(plan_without, {"2\u00baA"})

        self.assertEqual(multi_prof_alerts[0]["qtd_profissionais"], 2)
        self.assertEqual(plan_alerts[0]["qtd_casos"], 1)

        out = Workbook()
        add_quant_inclusao_alerts_sheet(out, multi_prof_alerts, plan_alerts)

        self.assertIn("ALERTAS", out.sheetnames)
        self.assertEqual(out["ALERTAS"]["A1"].value, "Categoria")
        self.assertEqual(out["ALERTAS"]["A2"].value, "M\u00faltiplos profissionais")

    def test_fill_quant_inclusao_workbook_writes_counts_header_and_alerts(self):
        wb_lista = Workbook()
        ws_lista = wb_lista.active
        ws_lista.title = "LISTA CORRIDA"
        ws_lista.append(["header"] * 16)
        ws_lista.append(make_lista_row("2-A", "001", "Aluno Um", "MA", "Sim", "Prof A"))
        ws_lista.append(make_lista_row("2-A", "002", "Aluno Dois", "MA", "Sim", "Prof B"))
        ws_lista.append(make_lista_row("2-A", "003", "Aluno Tres", "MA", "", "Prof C"))

        wb_model = Workbook()
        ws_model = wb_model.active
        ws_model["B4"] = "M\u00caS/2025"
        ws_model["B13"] = "2-A"

        alerts, plan_alerts = fill_quant_inclusao_workbook(
            wb_model,
            ws_lista,
            "Secretaria",
            datetime(2026, 3, 4),
        )

        self.assertEqual(ws_model["D13"].value, 2)
        self.assertEqual(ws_model["D14"].value, 2)
        self.assertEqual(ws_model["D15"].value, 2)
        self.assertEqual(ws_model["B4"].value, "MAR\u00c7O/2026")
        self.assertEqual(ws_model["C8"].value, "Secretaria")
        self.assertEqual(ws_model["K8"].value, "04/03/2026")
        self.assertEqual(len(alerts), 1)
        self.assertEqual(len(plan_alerts), 1)
        self.assertIn("ALERTAS", wb_model.sheetnames)

    def test_build_quant_inclusao_file_returns_excel_payload(self):
        with TemporaryDirectory() as tmpdir:
            lista_path = f"{tmpdir}/lista.xlsx"
            model_path = f"{tmpdir}/modelo.xlsx"

            wb_lista = Workbook()
            ws_lista = wb_lista.active
            ws_lista.title = "LISTA CORRIDA"
            ws_lista.append(["header"] * 16)
            ws_lista.append(make_lista_row("2-A", "001", "Aluno Um", "MA", "Sim", "Prof A"))
            wb_lista.save(lista_path)

            wb_model = Workbook()
            ws_model = wb_model.active
            ws_model["B4"] = "REFERENCIA"
            ws_model["B13"] = "2-A"
            wb_model.save(model_path)

            result = build_quant_inclusao_file(
                lista_path,
                model_path,
                "Secretaria",
                datetime(2026, 3, 4),
            )

            self.assertEqual(result.filename, "Quadro_Quantitativo_de_Inclusao_04032026.xlsx")
            loaded = load_workbook(result.output, data_only=False)
            self.assertEqual(loaded.active["D13"].value, 1)
            self.assertEqual(loaded.active["C8"].value, "Secretaria")
            self.assertEqual(result.alerts, [])

    def test_get_quant_inclusao_mes_ano_uppercase(self):
        self.assertEqual(get_quant_inclusao_mes_ano(datetime(2026, 3, 1)), "MAR\u00c7O/2026")


if __name__ == "__main__":
    unittest.main()
