import unittest
from datetime import datetime
from tempfile import TemporaryDirectory

from openpyxl import Workbook

from services.quadros_transferencias import (
    add_transfer_alerts_sheet,
    build_transferencias_workbook,
    collect_transfer_records,
    label_set,
    normalize_tipo_te,
    push_missing_info_alert,
    serie_key_from_value,
)


class QuadrosTransferenciasServiceTests(unittest.TestCase):
    def _write_lista_corrida(self, path, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "LISTA CORRIDA"
        ws.append(["SÉRIE", "NOME", "DATA NASC.", "RA", "OBS", "LOCAL TE"])
        for row in rows:
            ws.append(row)
        wb.save(path)

    def test_push_missing_info_alert_deduplicates_and_fills_blanks(self):
        alerts = []
        seen = set()

        payload = {
            "turma": "5A",
            "nome": "Aluno Ficticio",
            "ra": "",
            "tipo": "Sem Informacao",
            "data_str": "",
            "campo": "Tipo",
            "detalhe": "Campo vazio",
        }
        push_missing_info_alert(alerts, seen, **payload)
        push_missing_info_alert(alerts, seen, **payload)

        self.assertEqual(len(alerts), 1)
        self.assertEqual(alerts[0]["turma"], "5A")
        self.assertEqual(alerts[0]["ra"], "-")
        self.assertEqual(alerts[0]["data"], "-")

    def test_add_transfer_alerts_sheet_creates_alert_tab(self):
        wb = Workbook()

        add_transfer_alerts_sheet(
            wb,
            [
                {
                    "turma": "4A",
                    "nome": "Aluno Ficticio",
                    "ra": "123",
                    "tipo": "Rede Estadual",
                    "data": "01/02/2026",
                    "campo": "TIPO TE",
                    "detalhe": "Validado",
                }
            ],
        )

        self.assertIn("ALERTAS", wb.sheetnames)
        ws = wb["ALERTAS"]
        self.assertEqual(ws["A1"].value, "Turma")
        self.assertEqual(ws["B2"].value, "Aluno Ficticio")

    def test_label_set_preserves_template_prefix(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Responsavel: antigo"

        label_set(ws, "A1", "Responsavel", "Novo")

        self.assertEqual(ws["A1"].value, "Responsavel: Novo")

    def test_serie_key_from_value_accepts_expected_series(self):
        self.assertEqual(serie_key_from_value("4\u00baF"), "4\u00ba")
        self.assertEqual(serie_key_from_value("5o D"), "5\u00ba")
        self.assertIsNone(serie_key_from_value("1\u00baA"))
        self.assertIsNone(serie_key_from_value("2019"))
        self.assertIsNone(serie_key_from_value(""))

    def test_normalize_tipo_te_keeps_current_labels(self):
        self.assertEqual(normalize_tipo_te("rede estadual"), "Rede Estadual")
        self.assertEqual(normalize_tipo_te("Sao Paulo"), "S\u00e3o Paulo")
        self.assertEqual(normalize_tipo_te("pais"), "Pa\u00eds")
        self.assertEqual(normalize_tipo_te(""), "Sem Informa\u00e7\u00e3o")

    def test_collect_transfer_records_reads_fundamental_te_rows(self):
        with TemporaryDirectory() as tmpdir:
            lista_path = f"{tmpdir}/lista.xlsx"
            self._write_lista_corrida(
                lista_path,
                [
                    ["5\u00baA", "Aluno Transferido", datetime(2015, 2, 3), "123", "TE - 15/01/2026", "Rede Estadual"],
                    ["5\u00baA", "Aluno Fora", datetime(2015, 2, 3), "456", "TE - 20/03/2026", "Rede Estadual"],
                ],
            )

            records, alerts, debug = collect_transfer_records(
                lista_path,
                datetime(2026, 1, 1),
                datetime(2026, 1, 31),
            )

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["nome"], "Aluno Transferido")
        self.assertEqual(records[0]["dn"], "03/02/2015")
        self.assertEqual(records[0]["observacao"], "Rede Estadual")
        self.assertEqual(alerts, [])
        self.assertTrue(any("TE v\u00e1lidos" in item for item in debug))

    def test_collect_transfer_records_alerts_missing_local_te(self):
        with TemporaryDirectory() as tmpdir:
            lista_path = f"{tmpdir}/lista.xlsx"
            self._write_lista_corrida(
                lista_path,
                [["4\u00baB", "Aluno Sem Local", datetime(2016, 5, 4), "789", "TE - 10/01/2026", ""]],
            )

            records, alerts, _debug = collect_transfer_records(
                lista_path,
                datetime(2026, 1, 1),
                datetime(2026, 1, 31),
            )

        self.assertEqual(records[0]["observacao"], "-")
        self.assertEqual(len(alerts), 1)
        self.assertEqual(alerts[0]["campo"], "LOCAL TE")

    def test_build_transferencias_workbook_fills_template_and_alerts(self):
        with TemporaryDirectory() as tmpdir:
            model_path = f"{tmpdir}/modelo.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A7"] = "Unidade Escolar: antigo"
            ws["A8"] = "Diretor(a): antigo"
            wb.save(model_path)

            result_wb = build_transferencias_workbook(
                model_path=model_path,
                transfer_records=[
                    {
                        "nome": "Aluno Ficticio",
                        "dn": "03/02/2015",
                        "ra": "123",
                        "situacao": "Parcial",
                        "breda": "N\u00e3o",
                        "nivel_classe": "5\u00baA",
                        "tipo": "TE",
                        "observacao": "Rede Estadual",
                        "remanejamento": "-",
                        "data": "15/01/2026",
                    }
                ],
                missing_info_alerts=[
                    {
                        "turma": "5\u00baA",
                        "nome": "Aluno Ficticio",
                        "ra": "123",
                        "tipo": "TE",
                        "data": "15/01/2026",
                        "campo": "LOCAL TE",
                        "detalhe": "Teste",
                    }
                ],
                responsavel="Secretaria",
                diretor_nome="Diretora",
                data_quadro_dt=datetime(2026, 1, 31),
            )

        ws = result_wb.active
        self.assertEqual(ws["A7"].value, "Unidade Escolar: E.M Jos\u00e9 Padin Mouta")
        self.assertEqual(ws["A8"].value, "Diretor(a): Diretora")
        self.assertEqual(ws["B9"].value, "Secretaria")
        self.assertEqual(ws["J9"].value, "31/01/2026")
        self.assertEqual(ws["A12"].value, "Aluno Ficticio")
        self.assertEqual(ws["H12"].value, "Rede Estadual")
        self.assertEqual(ws["J12"].value, "15/01/2026")
        self.assertIn("ALERTAS", result_wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
