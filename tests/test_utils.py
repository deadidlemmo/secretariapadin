import unittest
import tempfile
from pathlib import Path

from openpyxl import Workbook

from services.upload_sessions import save_excel_upload_to_session
from utils.excel import set_merged_cell_value
from utils.uploads import allowed_excel_file, allowed_image_file, save_excel_upload, unique_secure_filename


class DummyUpload:
    def __init__(self, filename, content=b"conteudo"):
        self.filename = filename
        self.content = content

    def save(self, path):
        Path(path).write_bytes(self.content)


class UploadUtilsTests(unittest.TestCase):
    def test_allowed_upload_extensions(self):
        self.assertTrue(allowed_excel_file("lista.xlsx"))
        self.assertTrue(allowed_excel_file("lista.XLSM"))
        self.assertFalse(allowed_excel_file("lista.exe"))
        self.assertTrue(allowed_image_file("foto.PNG"))
        self.assertFalse(allowed_image_file("foto.pdf"))

    def test_unique_secure_filename_keeps_extension_and_prefix(self):
        filename = unique_secure_filename("../../Lista Piloto.xlsx", prefix="confere")
        self.assertTrue(filename.startswith("confere_"))
        self.assertTrue(filename.endswith("Lista_Piloto.xlsx"))
        self.assertNotIn("..", filename)

    def test_save_excel_upload_uses_safe_unique_name(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = save_excel_upload(DummyUpload("../../Lista Piloto.xlsx"), tmpdir, prefix="regular")

            self.assertTrue(Path(path).exists())
            self.assertTrue(Path(path).name.startswith("regular_"))
            self.assertTrue(Path(path).name.endswith("Lista_Piloto.xlsx"))

    def test_save_excel_upload_to_session_stores_path(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            session_data = {}

            path = save_excel_upload_to_session(
                DummyUpload("lista.xlsx"),
                session_data,
                "lista_fundamental",
                tmpdir,
                prefix="fundamental",
            )

            self.assertEqual(session_data["lista_fundamental"], path)
            self.assertTrue(Path(path).exists())


class ExcelUtilsTests(unittest.TestCase):
    def test_set_merged_cell_value_preserves_merge(self):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:B1")

        set_merged_cell_value(ws, "B1", "valor")

        self.assertEqual(ws["A1"].value, "valor")
        self.assertIn("A1:B1", [str(rng) for rng in ws.merged_cells.ranges])


if __name__ == "__main__":
    unittest.main()
