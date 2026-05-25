import unittest
from datetime import datetime

import pandas as pd
from openpyxl import Workbook

from services.quadros_quantitativo_mensal import (
    QuantitativoCountResult,
    collect_quantitativo_counts,
    fill_quantitativo_workbook,
    get_default_mes_ano,
)


class QuadrosQuantitativoMensalServiceTests(unittest.TestCase):
    def test_get_default_mes_ano_uses_portuguese_month_name(self):
        self.assertEqual(get_default_mes_ano(datetime(2026, 3, 1)), "Março/2026")

    def test_collect_quantitativo_counts_counts_and_discards_rows(self):
        df = pd.DataFrame(
            [
                ["100", "Aluno Cinco", "5\u00baA", "TE - 15/01/2026", "Rede Estadual"],
                ["101", "Aluno Quatro", "4o B", "TE - 20/01", "Sao Paulo"],
                ["102", "Aluno Fora", "5\u00baA", "TE - 20/03/2026", "Rede Estadual"],
                ["103", "Aluno Serie Invalida", "1\u00baA", "TE - 16/01/2026", "Rede Estadual"],
                ["104", "Aluno Sem Obs", "5\u00baA", "", "Rede Estadual"],
            ],
            columns=["RM", "NOME", "S\u00c9RIE", "OBS", "TIPO TE"],
        )

        result = collect_quantitativo_counts(
            df,
            datetime(2026, 1, 1),
            datetime(2026, 1, 31),
            default_year=2026,
        )

        self.assertEqual(result.counted, 2)
        self.assertEqual(result.discarded, 2)
        self.assertEqual(result.counts[("5\u00ba", "Rede Estadual")], 1)
        self.assertEqual(result.counts[("4\u00ba", "S\u00e3o Paulo")], 1)
        self.assertEqual(len(result.debug_rows), 4)
        self.assertEqual(result.debug_rows[0][8], "COUNTED")
        self.assertEqual(result.debug_rows[2][8], "SKIPPED")

    def test_fill_quantitativo_workbook_writes_cells_and_debug_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws["N15"] = 99
        ws["M23"] = 99

        count_result = QuantitativoCountResult(
            counts={
                ("5\u00ba", "Rede Estadual"): 2,
                ("4\u00ba", "Sem Informa\u00e7\u00e3o"): 1,
            },
            debug_rows=[
                [
                    2,
                    "100",
                    "Aluno Cinco",
                    "5\u00baA",
                    "TE - 15/01/2026",
                    "15/01/2026",
                    "NAO",
                    "TE - 15/01/2026",
                    "COUNTED",
                    "",
                    "Rede Estadual",
                    "Rede Estadual",
                ]
            ],
            counted=3,
            discarded=0,
        )

        fill_quantitativo_workbook(
            wb,
            count_result,
            period_start=datetime(2026, 1, 1),
            period_end=datetime(2026, 1, 31),
            responsavel="Secretaria",
            mes_ano="Janeiro/2026",
            ano_letivo=2026,
        )

        self.assertEqual(ws["N15"].value, 2)
        self.assertEqual(ws["M23"].value, 1)
        self.assertEqual(ws["B3"].value, "Secretaria")
        self.assertEqual(ws["D3"].value, "01/01/2026 a 31/01/2026")
        self.assertEqual(ws["A6"].value, "E.M Jos\u00e9 Padin Mouta")
        self.assertEqual(ws["A8"].value, "Janeiro/2026")
        self.assertEqual(ws["A10"].value, "QUADRO GERAL DE TRANSFERENCIAS EXPEDIDAS - 2026")
        self.assertIn("DEBUG_TE", wb.sheetnames)
        self.assertEqual(wb["DEBUG_TE"].sheet_state, "hidden")
        self.assertEqual(wb["DEBUG_TE"]["I2"].value, "COUNTED")


if __name__ == "__main__":
    unittest.main()
