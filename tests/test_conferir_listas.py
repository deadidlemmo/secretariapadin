import io
import unittest

import pandas as pd
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

from services.conferir_listas import (
    build_excel_report,
    compare_lista_piloto_sed,
    extract_sed_pdf_records,
    normalize_name_for_match,
    normalizar_status_sed,
    normalize_text,
    normalize_turma_lista,
    normalize_turma_sed,
    preview_sed_pdf_scope,
    ra_keys,
    read_lista_piloto,
    status_observacao,
    status_compativel,
)
from services.confere_escolas import ConfereSchoolConfig, get_confere_school_config


def make_lista_record(nome, ra, turma, nascimento, situacao):
    return {
        "source": "lista",
        "turma": turma,
        "turma_key": normalize_turma_lista(turma),
        "nome": nome,
        "nome_norm": normalize_text(nome),
        "nome_match_norm": normalize_name_for_match(nome),
        "ra": ra,
        "ra_keys": sorted(ra_keys(ra)),
        "data_nascimento": nascimento,
        "data_nascimento_norm": nascimento,
        "situacao": situacao,
        "observacoes": "",
    }


def make_sed_record(nome, ra, turma, nascimento, situacao, origem="sed.pdf"):
    return {
        "source": "sed",
        "turma": turma,
        "turma_key": normalize_turma_sed(turma),
        "nome": nome,
        "nome_norm": normalize_text(nome),
        "nome_match_norm": normalize_name_for_match(nome),
        "ra": ra,
        "ra_keys": sorted(ra_keys(ra)),
        "data_nascimento": nascimento,
        "data_nascimento_norm": nascimento,
        "situacao": situacao,
        "pdf_origem": origem,
    }


class ConferirListasServiceTests(unittest.TestCase):
    def test_normalizes_status_and_turma_mappings(self):
        self.assertTrue(status_compativel("MA", "ATIVO"))
        self.assertTrue(status_compativel("TE", "BXTR"))
        self.assertTrue(status_compativel("TE", "TRANSF"))
        self.assertTrue(status_compativel("REM", "REMA"))
        self.assertFalse(status_compativel("MA", "TRANSF"))
        self.assertFalse(status_compativel("REM", "ATIVO"))
        self.assertEqual(normalize_turma_lista("2\u00baA"), "2A")
        self.assertEqual(normalize_turma_sed("2\u00b0 ANO AI TARDE ANUAL"), "2A")
        self.assertEqual(normalize_turma_sed("2\u00b0 ANO G TARDE ANUAL"), "2G")
        self.assertEqual(normalize_name_for_match("SANT'ANNA"), normalize_name_for_match("SANT ANNA"))
        self.assertEqual(normalize_name_for_match("SANT'ANNA"), normalize_name_for_match("SANTANNA"))
        self.assertIn("124851170", ra_keys("000124851170"))
        self.assertIn("121298529", ra_keys("121298529-1"))
        self.assertIn("121371103", ra_keys("121371103-34"))
        self.assertEqual(
            status_observacao("MA", "REMA"),
            "Aluno ativo na Lista Piloto, mas no SED aparece como remanejado.",
        )
        self.assertEqual(
            status_observacao("MA", "TRANSF"),
            "Aluno ativo na Lista Piloto, mas no SED aparece como transferido.",
        )
        self.assertEqual(normalizar_status_sed("NCOM"), "NCOM")
        self.assertEqual(normalizar_status_sed("Nao Comparecimento"), "NCOM")
        self.assertEqual(normalizar_status_sed("Nao Compareceu"), "NCOM")

    def test_reads_lista_piloto_by_real_headers_not_fixed_rm_column(self):
        output = io.BytesIO()
        data = [
            {
                "S\u00c9RIE": "2\u00baA",
                "N\u00ba": "1",
                "RM": "19017",
                "NOME": "ANTHONY BONELLO SANTOS",
                "SEXO": "M",
                "DATA NASC.": "2018-04-09 00:00:00",
                "RA": "121074859",
                "COD.": "MA",
                "OBS": "0",
            }
        ]
        df = pd.DataFrame(data)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="LISTA CORRIDA", index=False)
        output.seek(0)

        records = read_lista_piloto(output)

        self.assertEqual(records[0]["turma"], "2\u00baA")
        self.assertEqual(records[0]["nome"], "ANTHONY BONELLO SANTOS")
        self.assertEqual(records[0]["ra"], "121.074.859")
        self.assertEqual(records[0]["data_nascimento"], "09/04/2018")
        self.assertEqual(records[0]["situacao"], "MA")

    def test_reads_lista_piloto_with_school_letter_config(self):
        output = io.BytesIO()
        rows = [[""] * 11 for _ in range(7)]
        rows.append(
            [
                "ALUNA CONFIG",
                "",
                "",
                "4\u00baB",
                "",
                "121371103-34",
                "2016-07-03 00:00:00",
                "",
                "",
                "Matricula Ativa",
                "observacao",
            ]
        )
        df = pd.DataFrame(rows)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Planilha1", index=False, header=False)
        output.seek(0)

        config = ConfereSchoolConfig(
            id="escola_teste",
            nome="Escola Teste",
            sheet_name="Planilha1",
            column_mode="letters",
            data_start_row=8,
            columns={
                "nome": "A",
                "turma": "D",
                "ra": "F",
                "data_nascimento": "G",
                "situacao": "J",
                "observacoes": "K",
            },
        )

        records = read_lista_piloto(output, school_config=config)

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["school_id"], "escola_teste")
        self.assertEqual(records[0]["row_number"], 8)
        self.assertEqual(records[0]["turma"], "4\u00baB")
        self.assertEqual(records[0]["nome"], "ALUNA CONFIG")
        self.assertEqual(records[0]["ra"], "121.371.103-34")
        self.assertEqual(records[0]["data_nascimento"], "03/07/2016")
        self.assertEqual(records[0]["situacao"], "MA")

    def test_reads_mahatma_gandhi_verificacao_sed_status_codes(self):
        output = io.BytesIO()
        df = pd.DataFrame(
            [
                {
                    "N\u00ba": "1",
                    "RM": "1001",
                    "NOME DO ALUNO": "ALUNA ATIVA",
                    "SEXO": "F",
                    "Idade": "8",
                    "NASC.": "2018-02-01",
                    "R.A.": "111",
                    "C\u00d3D.": "",
                    "OBSERVA\u00c7\u00c3O": "",
                    "TURMA": "2\u00ba A",
                },
                {
                    "N\u00ba": "N\u00ba",
                    "RM": "RM",
                    "NOME DO ALUNO": "NOME DO ALUNO",
                    "SEXO": "SEXO",
                    "Idade": "Idade",
                    "NASC.": "NASC.",
                    "R.A.": "R.A.",
                    "C\u00d3D.": "C\u00d3D.",
                    "OBSERVA\u00c7\u00c3O": "OBSERVA\u00c7\u00c3O",
                    "TURMA": "TURMA",
                },
                {
                    "N\u00ba": "2",
                    "RM": "1002",
                    "NOME DO ALUNO": "ALUNA PNEE",
                    "SEXO": "F",
                    "Idade": "8",
                    "NASC.": "2018-03-01",
                    "R.A.": "222",
                    "C\u00d3D.": "PNEE",
                    "OBSERVA\u00c7\u00c3O": "",
                    "TURMA": "2\u00ba A",
                },
                {
                    "N\u00ba": "3",
                    "RM": "1003",
                    "NOME DO ALUNO": "ALUNO TR",
                    "SEXO": "M",
                    "Idade": "8",
                    "NASC.": "2018-04-01",
                    "R.A.": "333",
                    "C\u00d3D.": "T.R.",
                    "OBSERVA\u00c7\u00c3O": "",
                    "TURMA": "2\u00ba A",
                },
                {
                    "N\u00ba": "4",
                    "RM": "1004",
                    "NOME DO ALUNO": "ALUNO TE",
                    "SEXO": "M",
                    "Idade": "8",
                    "NASC.": "2018-05-01",
                    "R.A.": "444",
                    "C\u00d3D.": "T.E.",
                    "OBSERVA\u00c7\u00c3O": "",
                    "TURMA": "2\u00ba A",
                },
                {
                    "N\u00ba": "5",
                    "RM": "1005",
                    "NOME DO ALUNO": "ALUNO REM",
                    "SEXO": "M",
                    "Idade": "8",
                    "NASC.": "2018-06-01",
                    "R.A.": "555",
                    "C\u00d3D.": "REM",
                    "OBSERVA\u00c7\u00c3O": "",
                    "TURMA": "2\u00ba A",
                },
            ]
        )
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Verifica\u00e7\u00e3o SED", index=False)
        output.seek(0)

        records = read_lista_piloto(
            output,
            school_config=get_confere_school_config("mahatma_gandhi"),
        )

        self.assertEqual([record["situacao"] for record in records], ["MA", "MA", "MA", "TE", "REM"])
        self.assertEqual([record["nome"] for record in records], ["ALUNA ATIVA", "ALUNA PNEE", "ALUNO TR", "ALUNO TE", "ALUNO REM"])
        self.assertTrue(all(record["turma_key"] == "2A" for record in records))

    def test_reads_mahatma_gandhi_ra_with_punctuation_and_uf_suffix(self):
        output = io.BytesIO()
        df = pd.DataFrame(
            [
                {
                    "N\u00ba": "1",
                    "RM": "1001",
                    "NOME DO ALUNO": "ALUNA RA",
                    "SEXO": "F",
                    "Idade": "8",
                    "NASC.": "2018-02-01",
                    "R.A.": "121.519.315-4 / SP",
                    "C\u00d3D.": "",
                    "OBSERVA\u00c7\u00c3O": "",
                    "TURMA": "2\u00ba A",
                },
            ]
        )
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Verifica\u00e7\u00e3o SED", index=False)
        output.seek(0)

        lista = read_lista_piloto(
            output,
            school_config=get_confere_school_config("mahatma_gandhi"),
        )
        sed = [
            make_sed_record("ALUNA RA", "000121519315", "2\u00b0 ANO A TARDE ANUAL", "01/02/2018", "ATIVO"),
        ]
        sed[0]["ra_keys"] = sorted(ra_keys("000121519315", "4"))

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["sed.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(lista[0]["ra"], "121.519.315-4")
        self.assertIn("1215193154", lista[0]["ra_keys"])
        self.assertEqual(result["summary"]["total_ok"], 1)
        self.assertEqual(result["summary"]["total_divergencias_cadastrais"], 0)

    def test_compares_records_by_ra_and_separates_categories(self):
        lista = [
            make_lista_record("Jo\u00e3o da Silva", "0001", "2\u00baA", "01/02/2017", "MA"),
            make_lista_record("Maria Souza", "0002", "2\u00baA", "03/04/2017", "TE"),
            make_lista_record("Aluno Sem SED", "0003", "2\u00baA", "05/06/2017", "MA"),
        ]
        sed = [
            make_sed_record("JOAO DA SILVA", "1", "2\u00b0 ANO AI TARDE ANUAL", "01/02/2017", "ATIVO"),
            make_sed_record("Maria Souza", "2", "2\u00b0 ANO AI TARDE ANUAL", "03/04/2017", "ATIVO"),
            make_sed_record("Aluno Extra", "999", "2\u00b0 ANO AI TARDE ANUAL", "07/08/2017", "ATIVO"),
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["sed.pdf"], "errors": [], "duplicate_files": []},
        )
        counts = result["summary"]

        self.assertEqual(counts["total_ok"], 1)
        self.assertEqual(counts["total_inconsistencias_situacao"], 1)
        self.assertEqual(counts["total_nao_encontrados_sed"], 1)
        self.assertEqual(counts["total_nao_encontrados_lista"], 1)
        status_row = next(row for row in result["rows"] if row["categoria"] == "inconsistencia_situacao")
        self.assertEqual(status_row["campos_divergentes"], ["situacao"])

    def test_limits_lista_piloto_scope_to_uploaded_sed_turmas(self):
        lista = [
            make_lista_record("Alice da Turma A", "0001", "2\u00baA", "01/02/2017", "MA"),
            make_lista_record("Aluno da Turma G", "0007", "2\u00baG", "07/02/2017", "MA"),
        ]
        sed = [
            make_sed_record("Aluno da Turma G", "7", "2\u00b0 ANO G TARDE ANUAL", "07/02/2017", "ATIVO"),
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["2g.pdf"], "errors": [], "duplicate_files": []},
        )
        counts = result["summary"]

        self.assertEqual(counts["total_lista_piloto"], 1)
        self.assertEqual(counts["total_lista_piloto_geral"], 2)
        self.assertEqual(counts["turmas_conferidas"], "2G")
        self.assertEqual(counts["total_ok"], 1)
        self.assertEqual(counts["total_nao_encontrados_sed"], 0)
        self.assertFalse(any(row["turma_lista"] == "2\u00baA" for row in result["rows"]))

    def test_treats_apostrophe_spacing_name_variations_as_same_name(self):
        lista = [
            make_lista_record(
                "ARETA SANTANNA QUADROS DIAS DE MELO",
                "121298529-1",
                "2\u00baG",
                "21/08/2018",
                "MA",
            )
        ]
        sed = [
            make_sed_record(
                "ARETA SANT ANNA QUADROS DIAS DE MELO",
                "000121298529",
                "2\u00b0 ANO G TARDE ANUAL",
                "21/08/2018",
                "ATIVO",
            )
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["2g.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(result["summary"]["total_ok"], 1)
        self.assertEqual(result["summary"]["total_divergencias_cadastrais"], 0)
        self.assertEqual(result["rows"][0]["categoria"], "ok")

    def test_matches_lista_ra_with_multi_digit_suffix_to_sed_base_ra(self):
        lista = [
            make_lista_record(
                "ALUNO FICTICIO TESTE",
                "121371103-34",
                "4\u00baB",
                "03/07/2016",
                "MA",
            )
        ]
        sed = [
            make_sed_record(
                "ALUNO FICTICIO TESTE",
                "000121371103",
                "4\u00b0 ANO BI MANHA ANUAL",
                "03/06/2016",
                "ATIVO",
            )
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["4b.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(result["summary"]["total_nao_encontrados_sed"], 0)
        self.assertEqual(result["summary"]["total_nao_encontrados_lista"], 0)
        self.assertEqual(result["summary"]["total_divergencias_cadastrais"], 1)
        row = result["rows"][0]
        self.assertEqual(row["categoria"], "divergencia_cadastral")
        self.assertEqual(row["observacao"], "Data de nascimento divergente entre as bases.")
        self.assertEqual(row["campos_divergentes"], ["data_nascimento"])

    def test_reports_probable_same_student_with_different_ra_as_single_divergence(self):
        lista = [
            make_lista_record(
                "DAVI FICTICIO DA SILVA",
                "1234801644",
                "3\u00baB",
                "29/06/2017",
                "MA",
            )
        ]
        sed = [
            make_sed_record(
                "DAVI FICTICIO SILVA",
                "000126528685",
                "3\u00b0 ANO BI TARDE ANUAL",
                "29/06/2017",
                "ATIVO",
            )
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["3b.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(result["summary"]["total_nao_encontrados_sed"], 0)
        self.assertEqual(result["summary"]["total_nao_encontrados_lista"], 0)
        self.assertEqual(result["summary"]["total_divergencias_cadastrais"], 1)
        self.assertEqual(len(result["rows"]), 1)
        row = result["rows"][0]
        self.assertEqual(row["categoria"], "divergencia_cadastral")
        self.assertEqual(row["observacao"], "RA divergente entre as bases.")
        self.assertEqual(row["campos_divergentes"], ["ra"])

    def test_ignores_sed_transfer_student_missing_from_lista_piloto(self):
        lista = [
            make_lista_record("Aluno Ativo Lista", "0001", "2\u00baG", "01/02/2018", "MA"),
        ]
        sed = [
            make_sed_record("Aluno Ativo Lista", "1", "2\u00b0 ANO G TARDE ANUAL", "01/02/2018", "ATIVO"),
            make_sed_record("Alice Muraca Said", "000122217587", "2\u00b0 ANO G TARDE ANUAL", "26/02/2019", "TRAN"),
            make_sed_record("Aluno NCOM", "000122217588", "2\u00b0 ANO G TARDE ANUAL", "27/02/2019", "NCOM"),
            make_sed_record("Aluno Nao Comparecimento", "000122217589", "2\u00b0 ANO G TARDE ANUAL", "28/02/2019", "Nao Comparecimento"),
            make_sed_record("Aluno Nao Compareceu", "000122217590", "2\u00b0 ANO G TARDE ANUAL", "01/03/2019", "Nao Compareceu"),
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["2g.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(result["summary"]["total_ok"], 1)
        self.assertEqual(result["summary"]["total_nao_encontrados_lista"], 0)
        self.assertEqual(result["summary"]["total_sed_inativos_sem_lista_ignorados"], 4)
        self.assertFalse(any(row["nome_sed"] == "Alice Muraca Said" for row in result["rows"]))
        self.assertFalse(any("NCOM" in row["nome_sed"] for row in result["rows"]))

    def test_prefers_active_same_turma_when_sed_has_remanejamento_history(self):
        lista = [
            make_lista_record("MIGUEL CARNAHIBA DE SOUZA", "120617416X", "3\u00baC", "15/01/2018", "MA"),
        ]
        sed = [
            make_sed_record("MIGUEL CARNAHIBA DE SOUZA", "000120617416", "3\u00b0 ANO DI TARDE ANUAL", "15/01/2018", "REMA"),
            make_sed_record("MIGUEL CARNAHIBA DE SOUZA", "000120617416", "3\u00b0 ANO CI TARDE ANUAL", "15/01/2018", "ATIVO"),
        ]
        for record in sed:
            record["ra_keys"] = sorted(ra_keys(record["ra"], "X"))

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 2, "successful_files": ["3c.pdf", "3d.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(result["summary"]["total_ok"], 1)
        self.assertEqual(result["summary"]["total_nao_encontrados_lista"], 0)
        self.assertEqual(result["summary"]["total_inconsistencias_situacao"], 0)
        self.assertEqual(result["summary"]["total_duplicidades_sed"], 0)
        self.assertEqual(sum(1 for row in result["rows"] if "mais de uma turma/PDF" in row["observacao"]), 0)
        ok_row = next(row for row in result["rows"] if row["categoria"] == "ok")
        self.assertEqual(ok_row["turma_sed"], "3\u00b0 ANO CI TARDE ANUAL")

    def test_accepts_active_lista_with_sed_rema_in_previous_turma(self):
        lista = [
            make_lista_record("ATHENA LUISA REIS DA CRUZ", "121.613.794-8", "3\u00baB", "12/09/2017", "MA"),
        ]
        sed = [
            make_sed_record("ATHENA LUISA REIS DA CRUZ", "000121613794", "3\u00b0 ANO DI TARDE ANUAL", "12/09/2017", "REMA"),
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["3d.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(result["summary"]["total_ok"], 1)
        self.assertEqual(result["summary"]["total_inconsistencias_situacao"], 0)
        self.assertEqual(result["summary"]["total_divergencias_cadastrais"], 0)
        row = result["rows"][0]
        self.assertEqual(row["categoria"], "ok")
        self.assertEqual(
            row["observacao"],
            "Aluno ativo na Lista Piloto e com registro de remanejamento no SED. Sem divergencia.",
        )

    def test_reports_active_lista_with_sed_rema_in_same_turma(self):
        lista = [
            make_lista_record("ATHENA LUISA REIS DA CRUZ", "121.613.794-8", "3\u00baD", "12/09/2017", "MA"),
        ]
        sed = [
            make_sed_record("ATHENA LUISA REIS DA CRUZ", "000121613794", "3\u00b0 ANO DI TARDE ANUAL", "12/09/2017", "REMA"),
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["3d.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(result["summary"]["total_ok"], 0)
        self.assertEqual(result["summary"]["total_inconsistencias_situacao"], 1)
        row = result["rows"][0]
        self.assertEqual(row["categoria"], "inconsistencia_situacao")
        self.assertEqual(row["campos_divergentes"], ["situacao"])

    def test_reports_sed_student_found_in_lista_piloto_with_different_turma(self):
        lista = [
            make_lista_record("LORENZO CRISTIANO TEODORO DOS SANTOS", "123387476", "2\u00baC", "04/07/2018", "MA"),
        ]
        sed = [
            make_sed_record(
                "LORENZO CRISTIANO TEODORO DOS SANTOS",
                "000123387476",
                "2\u00b0 ANO FI TARDE ANUAL",
                "04/07/2018",
                "ATIVO",
            ),
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["2f.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(result["summary"]["total_nao_encontrados_lista"], 0)
        self.assertEqual(result["summary"]["total_divergencias_cadastrais"], 1)
        row = result["rows"][0]
        self.assertEqual(row["categoria"], "divergencia_cadastral")
        self.assertEqual(row["turma_lista"], "2\u00baC")
        self.assertEqual(row["turma_sed"], "2\u00b0 ANO FI TARDE ANUAL")
        self.assertEqual(
            row["observacao"],
            "Aluno em turma diferente: Lista 2\u00baC; SED 2\u00baF. Confira a turma correta.",
        )

    def test_reports_same_ra_with_conflicting_identity_data(self):
        lista = [
            make_lista_record("ALUNO FICTICIO DA COSTA", "121167757-6", "4\u00baC", "26/07/2016", "MA"),
        ]
        sed = [
            make_sed_record(
                "ALUNO FICTICIO DE SOUZA",
                "000121167757",
                "2\u00b0 ANO DI TARDE ANUAL",
                "29/06/2016",
                "ATIVO",
            ),
        ]

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["2d.pdf"], "errors": [], "duplicate_files": []},
        )

        row = result["rows"][0]
        self.assertEqual(row["categoria"], "divergencia_cadastral")
        self.assertEqual(
            row["observacao"],
            "Mesmo RA em alunos diferentes: Lista 4\u00baC; SED 2\u00baD. Confira qual RA esta correto.",
        )
        self.assertEqual(row["campos_divergentes"], ["data_nascimento", "nome", "ra", "turma"])

    def test_reports_duplicate_lista_piloto_turma_instead_of_missing_sed(self):
        lista = [
            make_lista_record("MELISSA MIRANDA ROCHA", "122292805-X", "2\u00baD", "31/05/2018", "MA"),
            make_lista_record("MELISSA MIRANDA ROCHA", "122292805-X", "2\u00baA", "31/05/2018", "MA"),
        ]
        sed = [
            make_sed_record("MELISSA MIRANDA ROCHA", "000122292805", "2\u00b0 ANO AI TARDE ANUAL", "31/05/2018", "ATIVO"),
        ]
        sed[0]["ra_keys"] = sorted(ra_keys(sed[0]["ra"], "X"))

        result = compare_lista_piloto_sed(
            lista,
            sed,
            {"total_files": 1, "successful_files": ["2a.pdf"], "errors": [], "duplicate_files": []},
        )

        self.assertEqual(result["summary"]["total_ok"], 1)
        self.assertEqual(result["summary"]["total_nao_encontrados_sed"], 0)
        self.assertEqual(result["summary"]["total_divergencias_cadastrais"], 1)
        duplicate_row = next(row for row in result["rows"] if row["categoria"] == "divergencia_cadastral")
        self.assertEqual(duplicate_row["turma_lista"], "2\u00baD")
        self.assertEqual(duplicate_row["turma_sed"], "2\u00b0 ANO AI TARDE ANUAL")
        self.assertEqual(
            duplicate_row["observacao"],
            "Aluno duplicado em turmas diferentes: Lista 2\u00baD; SED 2\u00baA. Verifique duplicidade na Lista Piloto.",
        )

    def test_extracts_turma_and_student_from_sed_pdf_table(self):
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=18, rightMargin=18)
        styles = getSampleStyleSheet()
        table = Table(
            [
                [
                    "Série",
                    "N\u00ba",
                    "Nome do Aluno",
                    "RA",
                    "Dígito do RA",
                    "UF RA",
                    "Data de Nascimento",
                    "Data Movimentação",
                    "Situação",
                    "Condições educacionais especiais",
                    "Transtornos",
                ],
                [
                    "2",
                    "1",
                    "ALUNO FICTICIO TESTE",
                    "000123456",
                    "X",
                    "SP",
                    "10/03/2018",
                    "",
                    "ATIVO",
                    "",
                    "",
                ],
                [
                    "S\u00e9rie",
                    "N\u00ba",
                    "Nome do Aluno",
                    "RA",
                    "D\u00edgito do RA",
                    "UF RA",
                    "Data de Nascimento",
                    "Data Movimenta\u00e7\u00e3o",
                    "Situa\u00e7\u00e3o",
                    "Condi\u00e7\u00f5es educacionais especiais",
                    "Transtornos",
                ],
            ],
            colWidths=[34, 24, 150, 58, 44, 36, 72, 72, 52, 110, 90],
        )
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        doc.build([Paragraph("Turma: 2\u00b0 ANO AI TARDE ANUAL", styles["Normal"]), Spacer(1, 8), table])

        records = extract_sed_pdf_records(output.getvalue(), "sed_teste.pdf")

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["turma_key"], "2A")
        self.assertEqual(records[0]["nome"], "ALUNO FICTICIO TESTE")
        self.assertEqual(records[0]["ra"], "123.456-X")
        self.assertIn("123456X", records[0]["ra_keys"])
        self.assertEqual(records[0]["situacao"], "ATIVO")

        preview = preview_sed_pdf_scope([{"filename": "sed_teste.pdf", "content": output.getvalue()}])
        self.assertEqual(preview["turmas"], ["2ºA"])
        self.assertEqual(preview["pdfs_validos"], 1)
        self.assertEqual(preview["pdfs_ignorados"], 0)
        self.assertEqual(
            preview["file_scopes"],
            [
                {
                    "arquivo": "sed_teste.pdf",
                    "turmas": ["2ºA"],
                    "turma_keys": ["2A"],
                    "alunos": 1,
                }
            ],
        )

    def test_extracts_multiple_turmas_from_single_sed_pdf(self):
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=18, rightMargin=18)
        styles = getSampleStyleSheet()
        header = [
            "S\u00e9rie",
            "N\u00ba",
            "Nome do Aluno",
            "RA",
            "D\u00edgito do RA",
            "UF RA",
            "Data de Nascimento",
            "Data Movimenta\u00e7\u00e3o",
            "Situa\u00e7\u00e3o",
            "Condi\u00e7\u00f5es educacionais especiais",
            "Transtornos",
        ]

        def build_table(row):
            table = Table([header, row], colWidths=[34, 24, 150, 58, 44, 36, 72, 72, 52, 110, 90])
            table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("FONTSIZE", (0, 0), (-1, -1), 6),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            return table

        doc.build(
            [
                Paragraph("Turma: 2\u00b0 ANO AI TARDE ANUAL Ativos: 1", styles["Normal"]),
                Spacer(1, 6),
                build_table(["2", "1", "ALUNO TURMA A", "000111111111", "", "SP", "10/03/2018", "", "ATIVO", "", ""]),
                Spacer(1, 18),
                Paragraph("Turma: 3\u00b0 ANO BI TARDE ANUAL Ativos: 1", styles["Normal"]),
                Spacer(1, 6),
                build_table(["3", "1", "ALUNO TURMA B", "000222222222", "", "SP", "11/04/2017", "", "ATIVO", "", ""]),
            ]
        )

        records = extract_sed_pdf_records(output.getvalue(), "sed_multiturmas.pdf")

        self.assertEqual(len(records), 2)
        self.assertEqual([(record["nome"], record["turma_key"]) for record in records], [("ALUNO TURMA A", "2A"), ("ALUNO TURMA B", "3B")])

        preview = preview_sed_pdf_scope([{"filename": "sed_multiturmas.pdf", "content": output.getvalue()}])
        self.assertEqual(preview["turmas"], ["2\u00baA", "3\u00baB"])
        self.assertEqual(preview["file_scopes"][0]["turma_keys"], ["2A", "3B"])
        self.assertEqual(preview["file_scopes"][0]["alunos"], 2)

    def test_reuses_last_sed_turma_on_continuation_pages(self):
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=18, rightMargin=18)
        styles = getSampleStyleSheet()

        def build_table(nome, ra):
            table = Table(
                [
                    [
                        "S\u00e9rie",
                        "N\u00ba",
                        "Nome do Aluno",
                        "RA",
                        "D\u00edgito do RA",
                        "UF RA",
                        "Data de Nascimento",
                        "Data Movimenta\u00e7\u00e3o",
                        "Situa\u00e7\u00e3o",
                        "Condi\u00e7\u00f5es educacionais especiais",
                        "Transtornos",
                    ],
                    [
                        "5",
                        "1",
                        nome,
                        ra,
                        "",
                        "SP",
                        "18/12/2016",
                        "",
                        "ATIVO",
                        "",
                        "",
                    ],
                ],
                colWidths=[34, 24, 150, 58, 44, 36, 72, 72, 52, 110, 90],
            )
            table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("FONTSIZE", (0, 0), (-1, -1), 6),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            return table

        doc.build(
            [
                Paragraph("Turma: 5\u00b0 ANO BI MANHA ANUAL", styles["Normal"]),
                Spacer(1, 8),
                build_table("ALUNO PRIMEIRA PAGINA", "000121111111"),
                PageBreak(),
                build_table("ALUNO CONTINUACAO PAGINA", "000121222222"),
            ]
        )

        records = extract_sed_pdf_records(output.getvalue(), "sed_duas_paginas.pdf")

        self.assertEqual(len(records), 2)
        self.assertTrue(all(record["turma"] == "5\u00b0 ANO BI MANHA ANUAL" for record in records))
        self.assertTrue(all(record["turma_key"] == "5B" for record in records))

    def test_merges_sed_student_name_continuation_across_page_break(self):
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=18, rightMargin=18)
        styles = getSampleStyleSheet()

        header = [
            "S\u00e9rie",
            "N\u00ba",
            "Nome do Aluno",
            "RA",
            "D\u00edgito do RA",
            "UF RA",
            "Data de Nascimento",
            "Data Movimenta\u00e7\u00e3o",
            "Situa\u00e7\u00e3o",
            "Condi\u00e7\u00f5es educacionais especiais",
            "Transtornos",
        ]

        def build_table(rows):
            table = Table([header, *rows], colWidths=[34, 24, 150, 58, 44, 36, 72, 72, 52, 110, 90])
            table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("FONTSIZE", (0, 0), (-1, -1), 6),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            return table

        doc.build(
            [
                Paragraph("Turma: 3\u00b0 ANO BI TARDE ANUAL", styles["Normal"]),
                Spacer(1, 8),
                build_table(
                    [
                        [
                            "3",
                            "28",
                            "ALUNA FICTICIA",
                            "000123450025",
                            "5",
                            "SP",
                            "22/05/2017",
                            "18/12/",
                            "ATIVO",
                            "",
                            "",
                        ],
                    ]
                ),
                PageBreak(),
                build_table(
                    [
                        [
                            "",
                            "",
                            "GONCALVES\nCORDEIRO",
                            "",
                            "",
                            "",
                            "",
                            "2026",
                            "",
                            "",
                            "",
                        ],
                    ]
                ),
            ]
        )

        records = extract_sed_pdf_records(output.getvalue(), "sed_nome_quebrado.pdf")

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["nome"], "ALUNA FICTICIA GONCALVES CORDEIRO")
        self.assertEqual(records[0]["turma_key"], "3B")
        self.assertIn("1234500255", records[0]["ra_keys"])
        self.assertEqual(records[0]["data_nascimento"], "22/05/2017")
        self.assertEqual(records[0]["situacao"], "ATIVO")

    def test_build_excel_report_has_printable_review_tabs(self):
        result = compare_lista_piloto_sed(
            [
                make_lista_record("ALUNO OK", "1", "2\u00baA", "01/01/2018", "MA"),
                make_lista_record("ALUNO SEM SED", "2", "2\u00baA", "02/01/2018", "MA"),
                make_lista_record("ALUNO STATUS", "3", "2\u00baA", "03/01/2018", "MA"),
            ],
            [
                make_sed_record("ALUNO OK", "1", "2\u00b0 ANO AI TARDE ANUAL", "01/01/2018", "ATIVO"),
                make_sed_record("ALUNO STATUS", "3", "2\u00b0 ANO AI TARDE ANUAL", "03/01/2018", "REMA"),
                make_sed_record("ALUNO EXTRA", "9", "2\u00b0 ANO AI TARDE ANUAL", "09/01/2018", "ATIVO"),
            ],
            {"total_files": 1, "successful_files": ["sed.pdf"], "errors": [], "duplicate_files": []},
        )

        output = build_excel_report(result)
        wb = load_workbook(output)

        self.assertEqual(wb.sheetnames[0], "Resumo")
        for sheet_name in [
            "Para imprimir",
            "Inconsistencias",
            "Sem SED",
            "Sem Lista",
            "Cadastro",
            "OK",
            "Erros PDFs",
            "Base completa",
        ]:
            self.assertIn(sheet_name, wb.sheetnames)

        ws_print = wb["Para imprimir"]
        headers = [cell.value for cell in ws_print[1]]
        self.assertIn("Correcao / observacao manual", headers)
        self.assertNotIn("PDF SED", headers)
        self.assertEqual(ws_print.max_row, 4)
        self.assertEqual(ws_print.page_setup.orientation, "landscape")

        printed_statuses = [ws_print.cell(row=row_number, column=2).value for row_number in range(2, ws_print.max_row + 1)]
        self.assertNotIn("OK", printed_statuses)
        self.assertIn("Inconsistencia", printed_statuses)
        self.assertIn("Nao encontrado no SED", printed_statuses)
        self.assertIn("Nao encontrado na Lista", printed_statuses)

        sit_lista_col = headers.index("Sit. Lista") + 1
        sit_sed_col = headers.index("Sit. SED") + 1
        lista_statuses = [ws_print.cell(row=row_number, column=sit_lista_col).value for row_number in range(2, ws_print.max_row + 1)]
        sed_statuses = [ws_print.cell(row=row_number, column=sit_sed_col).value for row_number in range(2, ws_print.max_row + 1)]
        self.assertIn("Matricula Ativa", lista_statuses)
        self.assertIn("Remanejado", sed_statuses)
        self.assertIn("Ativo", sed_statuses)


if __name__ == "__main__":
    unittest.main()
