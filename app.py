from flask import (
    Flask,
    request,
    redirect,
    url_for,
    render_template,
    jsonify,
    session,
    flash,
    current_app,
    send_file,
)
import copy
import os
import re
import locale
from contextlib import contextmanager
from datetime import datetime
from functools import wraps
from io import BytesIO

import pandas as pd
import xlrd  # Para ler arquivos .xls, se necessário
from werkzeug.exceptions import RequestEntityTooLarge
from openpyxl import load_workbook, Workbook  # Usado para trabalhar com XLSX

from config import configure_app
from services.carteirinhas_log import (
    get_printed_set as _get_printed_set,
    mark_printed_rms as _mark_printed_rms,
)
from services.declaracoes import (
    DECLARACAO_PRINT_CSS,
    build_declaracao_escolar_context,
    build_declaracao_personalizada_context,
    build_notas_tabela_html,
    data_extenso_praia_grande,
    format_date_br,
    format_eja_rm,
    format_rm,
    format_serie_ano,
)
from services.fotos import get_student_photo_url, save_student_photo, student_has_photo
from services.prazos import build_deadline_alerts as build_deadline_alerts_service
from services.quadros_atendimento import (
    ATENDIMENTO_CONFIG,
    extract_by_cols as _extract_atendimento_by_cols,
    extract_by_fallback_block as _extract_atendimento_by_fallback_block,
    fill_eja_block as _fill_atendimento_eja_block,
    normalize_mes_ref as _normalize_atendimento_mes_ref,
    write_block as _write_atendimento_block,
    write_header as _write_atendimento_header,
    write_turno_totals as _write_atendimento_turno_totals,
    zero_eja_block as _zero_atendimento_eja_block,
)
from services.quadros_inclusao import (
    add_quant_inclusao_alerts_sheet as _add_quant_inclusao_alerts_sheet,
    build_multi_prof_alerts as _build_quant_multi_prof_alerts,
    build_plan_without_inclusion_alerts as _build_quant_plan_without_inclusion_alerts,
    build_template_map as _build_quant_template_map,
    collect_counts_from_lista_corrida as _collect_quant_counts_from_lista_corrida,
)
from services.quadros_transferencias import (
    RX_EJA_TRANSFER as _RX_EJA,
    add_transfer_alerts_sheet as _add_transfer_alerts_sheet,
    label_set as _label_set,
    normalize_tipo_te as _normalize_tipo_te,
    push_missing_info_alert as _push_missing_info_alert,
    serie_key_from_value as _serie_key_from_value,
)
from services.upload_sessions import save_excel_upload_to_session
from utils.dates import (
    detect_te_date_from_obs_flexible,
    extract_te_date_from_text as _extract_te_date_from_text,
    parse_date_flexible,
    parse_period_date as _parse_period_date,
    parse_user_date as _parse_user_date,
)
from utils.excel import set_merged_cell_value
from utils.text import (
    build_colmap as _build_colmap,
    find_df_col as _find_df_col,
    is_missing_text as _is_missing_text,
    is_missing_value as _is_missing_value,
    norm_header_compact as _norm_header_compact,
    pick_col as _pick_col,
    safe_str as _safe_str,
)
from utils.uploads import (
    save_excel_upload,
    validate_excel_upload,
)


# ==========================================================
#  CONFIGURAÇÃO BÁSICA DA APLICAÇÃO
# ==========================================================

# Tenta definir a localidade para formatação de datas em português
try:
    locale.setlocale(locale.LC_TIME, "pt_BR.UTF-8")
except locale.Error:
    pass

app = Flask(__name__)
ACCESS_TOKEN = configure_app(app)


# Cria os diretórios necessários, se não existirem
os.makedirs("static/fotos", exist_ok=True)
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

# Caminho do arquivo CSV relativo ao diretório do script
CSV_PATH = os.path.join(os.path.dirname(__file__), "uploads", "escolas.csv")

# Variável global para armazenar os dados do CSV
escolas_df = None


# ==========================================================
#  FUNÇÕES AUXILIARES – ESCOLAS (escolas.csv)
# ==========================================================

def carregar_escolas():
    """Carrega o CSV de escolas em um DataFrame global."""
    global escolas_df
    if os.path.exists(CSV_PATH):
        try:
            escolas_df = pd.read_csv(CSV_PATH, encoding="latin1", sep=";")
            print(f"[INFO] Arquivo {CSV_PATH} carregado com sucesso.")
        except Exception as e:
            escolas_df = None
            print(f"[ERRO] Falha ao carregar {CSV_PATH}: {e}")
    else:
        escolas_df = None
        print(f"[ERRO] Arquivo {CSV_PATH} não encontrado.")


def get_escolas_df():
    """Garante que o DataFrame de escolas está carregado."""
    global escolas_df
    if escolas_df is None or escolas_df.empty:
        print("[INFO] Recarregando arquivo escolas.csv...")
        carregar_escolas()
    return escolas_df


@app.before_request
def inicializar_escolas():
    """Garante que escolas.csv está carregado antes de cada requisição."""
    global escolas_df
    if escolas_df is None or (isinstance(escolas_df, pd.DataFrame) and escolas_df.empty):
        carregar_escolas()


@app.route("/escolas/search")
def escolas_search():
    """Endpoint para o Select2 buscar escolas no CSV."""
    df = get_escolas_df()
    query = request.args.get("q", "").lower().strip()
    results = []

    if df is not None and not df.empty and query:
        # Filtra usando pandas (assumindo coluna 3 = nome da escola)
        df_filtered = df[df.iloc[:, 3].str.lower().str.contains(query, na=False)]

        # Limita a 50 resultados para não sobrecarregar
        df_filtered = df_filtered.head(50)

        for _, row in df_filtered.iterrows():
            nome = str(row[3]).strip()
            municipio = str(row[2]).strip()
            uf = str(row[1]).strip()
            text = f"{nome} - {municipio}/{uf}"
            results.append({"id": nome, "text": text})

    return jsonify(results)


# Carrega o CSV na inicialização do sistema
carregar_escolas()


@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(_error):
    limit_mb = app.config.get("MAX_CONTENT_LENGTH", 0) // (1024 * 1024)
    message = f"Arquivo muito grande. O limite atual é de {limit_mb} MB."
    wants_json = request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html
    if wants_json:
        return jsonify({"error": message}), 413
    flash(message, "error")
    return redirect(request.referrer or url_for("dashboard"))


# ==========================================================
#  BLUEPRINTS
# ==========================================================

from confere import confere_bp

app.register_blueprint(confere_bp, url_prefix="/confere")


# ==========================================================
#  HELPERS GERAIS
# ==========================================================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login_route", next=request.url))
        return f(*args, **kwargs)

    return decorated_function


def convert_xls_to_xlsx(file_like):
    """
    Converte um arquivo XLS (file-like) para um Workbook do openpyxl.
    """
    book_xlrd = xlrd.open_workbook(file_contents=file_like.read())
    wb = Workbook()
    # Remove a planilha padrão criada pelo openpyxl, se houver
    if "Sheet" in wb.sheetnames and len(book_xlrd.sheet_names()) > 0:
        std = wb.active
        wb.remove(std)

    for sheet_name in book_xlrd.sheet_names():
        sheet_xlrd = book_xlrd.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        for row in range(sheet_xlrd.nrows):
            for col in range(sheet_xlrd.ncols):
                ws.cell(row=row + 1, column=col + 1, value=sheet_xlrd.cell_value(row, col))

    return wb


def load_workbook_model(file):
    """
    Abre o arquivo do modelo XLSX (ou XLS convertendo-o para XLSX)
    preservando a formatação.
    """
    ext = os.path.splitext(file.filename)[1].lower()
    file.seek(0)
    if ext == ".xlsx":
        return load_workbook(file, data_only=False)
    elif ext == ".xls":
        content = file.read()
        return convert_xls_to_xlsx(BytesIO(content))
    else:
        raise ValueError("Formato de arquivo não suportado para o quadro modelo.")


def is_valid_plano(val):
    """Avalia se o valor do Plano de Ação é considerado válido."""
    if val is None:
        return False
    s = str(val).strip()
    return s not in ["", "-", "0", "#REF"]


# ==========================================================
#  CARTEIRINHAS – GERAÇÃO DE HTML
# ==========================================================

def gerar_html_carteirinhas(arquivo_excel, somente_com_foto=False, somente_nao_impressas=False, ano=None):
    ano = int(ano or datetime.now().year)
    printed_set = _get_printed_set(ano)

    # Lê a planilha do Fundamental
    planilha = pd.read_excel(arquivo_excel, sheet_name="LISTA CORRIDA")

    dados = planilha[["RM", "NOME", "DATA NASC.", "RA", "SAI SOZINHO?", "SÉRIE", "HORÁRIO"]].copy()
    dados["RM"] = dados["RM"].fillna(0).astype(int)

    # Filtra apenas alunos com RM válido
    registros_validos = []
    for _, row in dados.iterrows():
        rm_str = str(row["RM"]).strip()
        if not rm_str or rm_str == "0":
            continue
        registros_validos.append(row)

    alunos_sem_fotos_list = []
    alunos = []

    for row in registros_validos:
        rm_int = int(row["RM"])
        nome = row["NOME"]
        data_nasc = row["DATA NASC."]
        serie = row["SÉRIE"]
        horario = row["HORÁRIO"]

        # Data de nascimento
        if pd.notna(data_nasc):
            try:
                data_nasc = pd.to_datetime(data_nasc, errors="coerce")
                if pd.notna(data_nasc):
                    data_nasc_str = data_nasc.strftime("%d/%m/%Y")
                else:
                    data_nasc_str = "Desconhecida"
            except Exception:
                data_nasc_str = "Desconhecida"
        else:
            data_nasc_str = "Desconhecida"

        ra = row["RA"]

        # Sai sozinho?
        sai_sozinho_raw = str(row["SAI SOZINHO?"]).strip().upper()
        if sai_sozinho_raw in ("SIM", "S", "YES", "Y"):
            classe_cor = "verde"
            status_texto = "Sai sozinho"
            status_icon = "&#10003;"
        else:
            classe_cor = "vermelho"
            status_texto = "Não sai sozinho"
            status_icon = "&#9888;"

        # Foto
        foto_url = get_student_photo_url(rm_int)

        if not foto_url:
            alunos_sem_fotos_list.append(
                {
                    "rm": rm_int,
                    "nome": nome,
                    "serie": serie,
                }
            )

        # flag de impressão (lógica nova)
        impresso = (rm_int in printed_set)

        alunos.append(
            {
                "rm": rm_int,
                "nome": nome,
                "data_nasc": data_nasc_str,
                "ra": ra,
                "serie": serie,
                "horario": horario,
                "classe_cor": classe_cor,
                "status_texto": status_texto,
                "status_icon": status_icon,
                "foto_url": foto_url,
                "impresso": impresso,  # <<< NOVO
            }
        )

    # --- FILTROS ---
    alunos_para_exibir = alunos

    # mantém seu filtro atual
    if somente_com_foto:
        alunos_para_exibir = [a for a in alunos_para_exibir if a.get("foto_url")]

    # novo filtro: não impressas
    if somente_nao_impressas:
        alunos_para_exibir = [a for a in alunos_para_exibir if not a.get("impresso")]

    # Paginação: 6 carteirinhas por página (mantém)
    pages = []
    for i in range(0, len(alunos_para_exibir), 6):
        pages.append(alunos_para_exibir[i: i + 6])

    total_sem_foto = len(alunos_sem_fotos_list)

    return render_template(
        "gerar_carteirinhas.html",
        pages=pages,
        alunos_sem_foto=alunos_sem_fotos_list,
        total_sem_foto=total_sem_foto,
        somente_com_foto=somente_com_foto,
        somente_nao_impressas=somente_nao_impressas,  # <<< NOVO (se quiser mostrar no template)
        ano=ano,                                      # <<< NOVO (útil p/ log por ano)
    )


# ==========================================================
#  (NECESSÁRIO) CARTEIRINHAS – MARCAR COMO IMPRESSAS
#  -> seu JS chama isso depois de imprimir (ou ao clicar em "Imprimir")
# ==========================================================

# OBS: se você já tem login_required, mantenha aqui também
# @login_required
def _normalize_rms(rms):
    out = []
    for x in (rms or []):
        try:
            v = int(str(x).strip())
            if v > 0:
                out.append(v)
        except Exception:
            pass
    # unique preservando ordem
    return list(dict.fromkeys(out))


# Se você usa CSRFProtect (Flask-WTF), descomente a linha do csrf.exempt
# from flask_wtf.csrf import CSRFProtect
# csrf = CSRFProtect(app)

@app.route("/carteirinhas/marcar_impressas", methods=["POST"])
@login_required
# @csrf.exempt  # use isto se CSRF estiver bloqueando POST JSON
def marcar_carteirinhas_impressas():
    payload = request.get_json(silent=True) or {}
    rms = _normalize_rms(payload.get("rms", []))
    ano = int(payload.get("ano") or datetime.now().year)

    # REGRA-CHAVE: só marca como "impressa" se tiver foto de verdade
    rms = [rm for rm in rms if student_has_photo(rm)]

    added, total_printed = _mark_printed_rms(ano, rms)

    return jsonify({
        "ok": True,
        "ano": ano,
        "received": len(_normalize_rms(payload.get("rms", []))),  # recebido bruto
        "considered_with_photo": len(rms),                        # após filtro foto
        "added": added,
        "total_printed": total_printed,
    })



# ==========================================================
#  DECLARAÇÕES – GERAÇÃO HTML (SINGULAR)
# ==========================================================

def gerar_declaracao_escolar(
    file_path,
    rm,
    tipo,
    file_path2=None,
    deve_historico=False,
    unidade_anterior=None,
    dados_frequencia=None,
):
    """
    Gera o HTML de uma declaração escolar (Escolaridade, Transferência, Conclusão
    ou Frequência) tanto para Fundamental quanto EJA, de acordo com
    session['declaracao_tipo'].

    file_path  -> caminho/arquivo padrão da lista piloto (salvo em sessão/ao entrar no sistema)
    file_path2 -> caminho/arquivo opcional, usado quando o usuário reenviar a lista
                  (por exemplo, após o servidor free acordar). SE informado, TERÁ PRIORIDADE.
    dados_frequencia -> dicionário opcional com os dados de frequência por mês,
                        utilizado apenas quando tipo == "Frequencia".
    """
    global escolas_df

    # Se um segundo caminho foi informado (lista reenviada), ele tem prioridade.
    effective_path = file_path2 if file_path2 is not None else file_path
    if file_path2 is not None:
        print("[DEBUG] gerar_declaracao_escolar: usando file_path2 =", effective_path)
    else:
        print("[DEBUG] gerar_declaracao_escolar: usando file_path  =", effective_path)

    # HTML da tabela de notas (usado apenas na transferência do Fundamental)
    notas_tabela_html = ""

    # ------------------------------------------------------
    # 1) CARREGAMENTO DOS DADOS DO ALUNO (FUNDAMENTAL x EJA)
    # ------------------------------------------------------
    if session.get("declaracao_tipo") != "EJA":
        # ---------- FUNDAMENTAL ----------
        planilha = pd.read_excel(effective_path, sheet_name="LISTA CORRIDA")
        planilha.columns = [c.strip().upper() for c in planilha.columns]

        planilha["RM_str"] = planilha["RM"].apply(format_rm)

        rm_num = format_rm(rm)

        aluno = planilha[planilha["RM_str"] == rm_num]
        if aluno.empty:
            return None

        row = aluno.iloc[0]

        semestre_texto = ""  # Fundamental não usa semestre
        nome = row["NOME"]
        serie = row["SÉRIE"]

        if isinstance(serie, str):
            # transforma "5ºA" em "5º ano A"
            serie = format_serie_ano(serie)

        data_nasc = row["DATA NASC."]
        ra = row["RA"]
        horario = row.get("HORÁRIO", "Desconhecido")

        if pd.isna(horario) or not str(horario).strip():
            horario = "Desconhecido"
        else:
            horario = str(horario).strip()

        ra_label = "RA"

        data_nasc = format_date_br(data_nasc)

        # --------------------------------------------------
        # NOVO: busca as notas na aba NOTAS (apenas Transferência/Fundamental)
        # --------------------------------------------------
        if tipo == "Transferencia":
            notas_tabela_html = build_notas_tabela_html(effective_path, rm_num)
    else:
        # ---------- EJA ----------
        df = pd.read_excel(effective_path, sheet_name=0, header=None, skiprows=1)
        df.columns = [str(c).strip().upper() for c in df.columns]

        # RM (coluna 2)
        df["RM_str"] = df.iloc[:, 2].apply(format_eja_rm)
        # Nome (coluna 3)
        df["NOME"] = df.iloc[:, 3]
        # Nascimento (coluna 6)
        df["NASC."] = df.iloc[:, 6]

        def get_ra(row_local):
            try:
                val = row_local.iloc[7]
                if pd.isna(val) or float(val) == 0:
                    return row_local.iloc[8]
                else:
                    return val
            except Exception:
                return row_local.iloc[7]

        df["RA"] = df.apply(get_ra, axis=1)
        # Série (coluna 0)
        df["SÉRIE"] = df.iloc[:, 0]

        rm_num = format_rm(rm)

        aluno = df[df["RM_str"] == rm_num]
        if aluno.empty:
            return None

        row = aluno.iloc[0]

        # Semestre (quando existir na planilha)
        if len(row) > 29:
            semestre = row.iloc[29]
            semestre_texto = str(semestre).strip() if pd.notna(semestre) else ""
        else:
            semestre_texto = ""

        nome = row["NOME"]
        serie = row["SÉRIE"]
        if isinstance(serie, str):
            serie = format_serie_ano(serie)

        data_nasc = row["NASC."]
        ra = row["RA"]
        original_ra = row.iloc[7]

        # Se RA for vazio / 0, trata como RG
        if pd.isna(original_ra) or (
            isinstance(original_ra, (int, float)) and float(original_ra) == 0
        ):
            ra_label = "RG"
        else:
            ra_label = "RA"

        data_nasc = format_date_br(data_nasc)


    context = build_declaracao_escolar_context(
        tipo=tipo,
        segmento=session.get("declaracao_tipo"),
        nome=nome,
        ra=ra,
        ra_label=ra_label,
        data_nasc=data_nasc,
        serie=serie,
        horario=horario if session.get("declaracao_tipo") != "EJA" else "Desconhecido",
        semestre_texto=semestre_texto,
        row=row,
        notas_tabela_html=notas_tabela_html,
        deve_historico=deve_historico,
        unidade_anterior=unidade_anterior,
        escolas_df=escolas_df,
        dados_frequencia=dados_frequencia,
    )
    if context is None:
        return None

    data_extenso_str = data_extenso_praia_grande()
    additional_css = DECLARACAO_PRINT_CSS

    return render_template(
        "declaracao_print.html",
        titulo=context["titulo"],
        data_extenso=data_extenso_str,
        declaracao_text=context["declaracao_text"],
        additional_css=additional_css,
        body_classes=context["body_classes"],
        print_body_padding="0.5cm 0.5cm",
    )


# ==========================================================
#  NOVA FUNÇÃO – LOTE DE ESCOLARIDADE 5º ANO (FUNDAMENTAL)
# ==========================================================

def gerar_lote_escolaridade_5ano(file_path, file_path2=None):
    """
    Gera os dados para DECLARAÇÕES DE ESCOLARIDADE de todos os alunos de 5º ano
    (Fundamental) em lote.
    """
    effective_path = file_path2 if file_path2 is not None else file_path
    if not effective_path:
        raise ValueError(
            "Caminho do arquivo Excel não informado para o lote de escolaridade 5º ano."
        )

    planilha = pd.read_excel(effective_path, sheet_name="LISTA CORRIDA")
    planilha.columns = [c.strip().upper() for c in planilha.columns]

    planilha["RM_str"] = planilha["RM"].apply(format_rm)

    registros = []

    for _, row in planilha.iterrows():
        rm_str = str(row.get("RM_str", "")).strip()
        if rm_str in ("", "0"):
            continue

        serie_raw = str(row.get("SÉRIE", "")).strip()
        if not serie_raw:
            continue

        if "5º" not in serie_raw and "5°" not in serie_raw:
            continue

        nome = str(row.get("NOME", "")).strip()
        ra = str(row.get("RA", "")).strip()

        data_nasc_val = row.get("DATA NASC.")
        data_nasc = format_date_br(data_nasc_val)

        horario = str(row.get("HORÁRIO", "")).strip()
        if not horario:
            horario = "Desconhecido"

        serie_fmt = serie_raw
        try:
            serie_fmt = format_serie_ano(serie_fmt)
        except Exception:
            pass

        texto = (
            f"Declaro, para os devidos fins, que o(a) aluno(a) "
            f"<strong><u>{nome}</u></strong>, portador(a) do RA "
            f"<strong><u>{ra}</u></strong>, nascido(a) em "
            f"<strong><u>{data_nasc}</u></strong>, "
            f"encontra-se regularmente matriculado(a) na "
            f"E.M José Padin Mouta, cursando atualmente o(a) "
            f"<strong><u>{serie_fmt}</u></strong> no horário de aula: "
            f"<strong><u>{horario}</u></strong>."
        )

        registros.append(
            {
                "nome": nome,
                "ra": ra,
                "data_nasc": data_nasc,
                "serie_fmt": serie_fmt,
                "horario": horario,
                "texto": texto,
            }
        )

    data_extenso_str = data_extenso_praia_grande()
    titulo = "Declaração de Escolaridade"

    return registros, data_extenso_str, titulo


# ==========================================================
#  DECLARAÇÃO PERSONALIZADA (Fundamental / EJA)
# ==========================================================

def gerar_declaracao_personalizada(dados):
    """
    Gera o HTML de declarações personalizadas (Conclusão, Matrícula cancelada
    ou Não Comparecimento - NCOM).
    """
    context = build_declaracao_personalizada_context(dados)
    if context is None:
        return None

    data_extenso_str = data_extenso_praia_grande()
    additional_css = DECLARACAO_PRINT_CSS

    return render_template(
        "declaracao_print.html",
        titulo=context["titulo"],
        data_extenso=data_extenso_str,
        declaracao_text=context["declaracao_text"],
        additional_css=additional_css,
        body_classes=[],
        print_body_padding="1.5cm 1.5cm",
    )


# ==========================================================
#  AUTENTICAÇÃO / DASHBOARD
# ==========================================================

@app.route("/login", methods=["GET", "POST"])
def login_route():
    error = None
    if request.method == "POST":
        if not ACCESS_TOKEN:
            error = "ACCESS_TOKEN não configurado no servidor. Configure a variável de ambiente antes de usar o sistema."
            return render_template("login.html", error=error), 503

        token = request.form.get("token")
        if token == ACCESS_TOKEN:
            session["logged_in"] = True

            # ==========================================================
            # AJUSTE: EJA NÃO É MAIS OBRIGATÓRIA
            # - Só exige lista_fundamental para seguir para o dashboard
            # - lista_eja permanece suportada, mas opcional
            # ==========================================================
            if "lista_fundamental" not in session:
                return redirect(url_for("upload_listas"))

            return redirect(url_for("dashboard"))
        else:
            error = "Token inválido. Tente novamente."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout_route():
    session.clear()
    return redirect(url_for("login_route"))


@app.route("/upload_listas", methods=["GET", "POST"])
@login_required
def upload_listas():
    if request.method == "POST":
        fundamental_file = request.files.get("lista_fundamental")
        eja_file = request.files.get("lista_eja")  # agora é opcional

        if not fundamental_file or fundamental_file.filename == "":
            flash("Selecione a Lista Piloto - REGULAR - 2025", "error")
            return redirect(url_for("upload_listas"))

        valid, message = validate_excel_upload(fundamental_file)
        if not valid:
            flash(message, "error")
            return redirect(url_for("upload_listas"))

        valid, message = validate_excel_upload(eja_file, required=False)
        if not valid:
            flash(f"Lista EJA inválida: {message}", "error")
            return redirect(url_for("upload_listas"))

        # Salva Fundamental (obrigatório)
        save_excel_upload_to_session(
            fundamental_file,
            session,
            "lista_fundamental",
            app.config["UPLOAD_FOLDER"],
            prefix="fundamental",
        )

        # Salva EJA (opcional)
        eja_salva = False
        if eja_file and eja_file.filename:
            save_excel_upload_to_session(
                eja_file,
                session,
                "lista_eja",
                app.config["UPLOAD_FOLDER"],
                prefix="eja",
            )
            eja_salva = True

        if eja_salva:
            flash("Listas carregadas com sucesso (Fundamental e EJA).", "success")
        else:
            flash(
                "Lista do Fundamental carregada com sucesso. A lista de EJA é opcional e não foi enviada.",
                "success",
            )

        return redirect(url_for("dashboard"))

    return render_template("upload_listas.html")


@app.route("/", methods=["GET"])
@login_required
def dashboard():
    return render_template("dashboard.html")


# ==========================================================
#  CARTEIRINHAS – ROTA PRINCIPAL (AJUSTADA)
#  - Mantém tudo que já existe
#  - Adiciona suporte ao filtro "somente_nao_impressas"
#  - Persiste ambos filtros em session p/ GET
# ==========================================================

@app.route("/carteirinhas", methods=["GET", "POST"])
@login_required
def carteirinhas():
    if request.method == "POST":
        file_path = None

        # Checkbox no form (ex.: <input type="checkbox" name="somente_com_foto">)
        somente_com_foto = request.form.get("somente_com_foto") in ("1", "on", "true", "True", "SIM", "sim")
        session["carteirinhas_somente_com_foto"] = somente_com_foto

        # NOVO: checkbox "somente_nao_impressas"
        # (ex.: <input type="checkbox" name="somente_nao_impressas">)
        somente_nao_impressas = request.form.get("somente_nao_impressas") in ("1", "on", "true", "True", "SIM", "sim")
        session["carteirinhas_somente_nao_impressas"] = somente_nao_impressas

        if "excel_file" in request.files and request.files["excel_file"].filename != "":
            file = request.files["excel_file"]
            valid, message = validate_excel_upload(file)
            if not valid:
                flash(message, "error")
                return redirect(url_for("carteirinhas"))

            file_path = save_excel_upload_to_session(
                file,
                session,
                "lista_fundamental",
                app.config["UPLOAD_FOLDER"],
                prefix="carteirinhas",
            )
        else:
            file_path = session.get("lista_fundamental")

        if not file_path or not os.path.exists(file_path):
            flash("Nenhum arquivo selecionado. Envie a lista piloto do Fundamental.", "info")
            return redirect(url_for("carteirinhas"))

        flash("Gerando carteirinhas. Aguarde...", "info")
        html_result = gerar_html_carteirinhas(
            file_path,
            somente_com_foto=somente_com_foto,
            somente_nao_impressas=somente_nao_impressas,  # NOVO
            ano=datetime.now().year,                      # NOVO (para log por ano)
        )
        return html_result

    # GET: passa o estado atual dos filtros para o template marcar os checkbox
    somente_com_foto = session.get("carteirinhas_somente_com_foto", False)
    somente_nao_impressas = session.get("carteirinhas_somente_nao_impressas", False)

    return render_template(
        "carteirinhas.html",
        somente_com_foto=somente_com_foto,
        somente_nao_impressas=somente_nao_impressas,  # NOVO
    )


#  DECLARAÇÕES – CONCLUSÃO 5º ANO (LOTE)
# ==========================================================

@app.route("/declaracao/conclusao_5ano")
@login_required
def declaracao_conclusao_5ano():
    if session.get("declaracao_tipo") != "Fundamental":
        flash(
            "As declarações em lote de 5º ano estão disponíveis apenas para o Fundamental.",
            "error",
        )
        return redirect(url_for("declaracao_tipo"))

    file_path = session.get("declaracao_excel") or session.get("lista_fundamental")

    if not file_path or not os.path.exists(file_path):
        flash(
            "Arquivo Excel do Fundamental não encontrado. "
            "Anexe a lista piloto novamente pela tela de declarações.",
            "error",
        )
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    planilha = pd.read_excel(file_path, sheet_name="LISTA CORRIDA")
    planilha.columns = [c.strip().upper() for c in planilha.columns]

    planilha["RM_str"] = planilha["RM"].apply(format_rm)

    registros = []

    for _, row in planilha.iterrows():
        rm_str = str(row.get("RM_str", "")).strip()
        if rm_str in ("", "0"):
            continue

        serie_raw = str(row.get("SÉRIE", "")).strip()
        if not serie_raw:
            continue

        if "5º" not in serie_raw and "5°" not in serie_raw:
            continue

        nome = str(row.get("NOME", "")).strip()
        ra = str(row.get("RA", "")).strip()

        data_nasc_val = row.get("DATA NASC.")
        data_nasc = format_date_br(data_nasc_val)

        horario = str(row.get("HORÁRIO", "")).strip()
        if not horario:
            horario = "Desconhecido"

        serie_fmt = serie_raw
        try:
            serie_fmt = format_serie_ano(serie_fmt)
        except Exception:
            pass

        series_text = "a série subsequente"
        m = re.search(r"(\d+)º", serie_fmt)
        if m:
            try:
                next_year = int(m.group(1)) + 1
                series_text = f"{next_year}º ano"
            except Exception:
                pass

        valor_bolsa = str(row.get("BOLSA FAMILIA", "")).strip().upper()

        declaracao_text = (
            f"Declaro, para os devidos fins, que o(a) aluno(a) "
            f"<strong><u>{nome}</u></strong>, "
            f"portador(a) do RA <strong><u>{ra}</u></strong>, nascido(a) em "
            f"<strong><u>{data_nasc}</u></strong>, concluiu com êxito o "
            f"<strong><u>{serie_fmt}</u></strong>, estando apto(a) a ingressar no "
            f"<strong><u>{series_text}</u></strong>."
        )

        if valor_bolsa == "SIM":
            declaracao_text += "<br><br><strong>Observações:</strong><br>"
            declaracao_text += (
                '<label class="checkbox-label" '
                'style="display: block; text-align: justify; font-size:14px;">'
            )
            declaracao_text += (
                f'<img src="{url_for("static", filename="logos/bolsa_familia.jpg")}" '
                'alt="Bolsa Família" '
                'style="width:28px; vertical-align:middle; margin-right:5px;">'
                "O aluno é beneficiário do Programa Bolsa Família."
            )
            declaracao_text += "</label>"

        registros.append(
            {
                "nome": nome,
                "ra": ra,
                "data_nasc": data_nasc,
                "serie_fmt": serie_fmt,
                "series_text": series_text,
                "texto": declaracao_text,
            }
        )

    if not registros:
        flash("Nenhum aluno de 5º ano encontrado na lista piloto.", "error")
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    data_extenso_str = current_app.config.get(
        "CONCLUSAO_5ANO_DATE_TEXT",
        f"Praia Grande, 22 de dezembro de {datetime.now().year}",
    )
    titulo = "Declaração de Conclusão"

    registros_duas_vias = []
    for reg in registros:
        reg1 = reg.copy()
        reg1["via"] = 1
        registros_duas_vias.append(reg1)

        reg2 = reg.copy()
        reg2["via"] = 2
        registros_duas_vias.append(reg2)

    return render_template(
        "declaracao_conclusao_5ano.html",
        registros=registros_duas_vias,
        data_extenso=data_extenso_str,
        titulo=titulo,
        total=len(registros_duas_vias),
    )


# ==========================================================
#  DECLARAÇÕES – TELA ÚNICA (Fundamental / EJA / Personalizada)
# ==========================================================

@app.route("/declaracao/tipo", methods=["GET", "POST"])
@login_required
def declaracao_tipo():
    if request.method == "POST":
        modo_declaracao = request.form.get("modo_declaracao")

        # -------------------------------
        # FLUXO: DECLARAÇÃO PERSONALIZADA
        # -------------------------------
        if modo_declaracao == "personalizada":
            segmento_pers = request.form.get("segmento_personalizado")
            if segmento_pers not in ("Fundamental", "EJA"):
                flash(
                    "Selecione o segmento (Ensino Fundamental ou EJA) na declaração personalizada.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

            nome_aluno = (request.form.get("nome_aluno") or "").strip()
            data_nascimento = request.form.get("data_nascimento")
            ra = (request.form.get("ra") or "").strip()
            tipo_pers = request.form.get("tipo_declaracao_personalizada")

            if (
                not nome_aluno
                or not data_nascimento
                or not ra
                or tipo_pers not in ("Conclusao", "MatriculaCancelada", "NCOM")
            ):
                flash(
                    "Preencha todos os dados obrigatórios da declaração personalizada.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

            dados_personalizados = {
                "segmento": segmento_pers,
                "nome_aluno": nome_aluno,
                "data_nascimento": data_nascimento,
                "ra": ra,
                "tipo_declaracao": tipo_pers,
            }

            if tipo_pers == "Conclusao":
                ano_serie_concluida = (request.form.get("ano_serie_concluida") or "").strip()
                ano_conclusao = (request.form.get("ano_conclusao") or "").strip()
                deve_hist_unidade = request.form.get("deve_historico_unidade")
                semestre_conclusao = (request.form.get("semestre_conclusao") or "").strip()

                campos_invalidos = (
                    not ano_serie_concluida
                    or not ano_conclusao
                    or deve_hist_unidade not in ("Sim", "Não")
                )

                if segmento_pers == "EJA" and not semestre_conclusao:
                    campos_invalidos = True

                if campos_invalidos:
                    flash(
                        "Preencha todos os campos da declaração personalizada de conclusão "
                        "(para EJA é obrigatório informar o semestre).",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

                dados_personalizados.update(
                    {
                        "ano_serie_concluida": ano_serie_concluida,
                        "ano_conclusao": ano_conclusao,
                        "deve_historico_unidade": (deve_hist_unidade == "Sim"),
                        "semestre_conclusao": semestre_conclusao,
                    }
                )

            elif tipo_pers == "MatriculaCancelada":
                ano_serie_matricula = (request.form.get("ano_serie_matricula") or "").strip()
                ano_matricula = (request.form.get("ano_matricula") or "").strip()
                semestre_matricula = (request.form.get("semestre_matricula") or "").strip()

                if not ano_serie_matricula or not ano_matricula or not semestre_matricula:
                    flash(
                        "Preencha todos os campos da declaração de matrícula cancelada.",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

                dados_personalizados.update(
                    {
                        "ano_serie_matricula": ano_serie_matricula,
                        "ano_matricula": ano_matricula,
                        "semestre_matricula": semestre_matricula,
                    }
                )

            elif tipo_pers == "NCOM":
                ano_serie_vaga = (request.form.get("ano_serie_vaga") or "").strip()
                ano_referencia_ncom = (request.form.get("ano_referencia_ncom") or "").strip()
                semestre_referencia_ncom = (
                    request.form.get("semestre_referencia_ncom") or ""
                ).strip()

                if not ano_serie_vaga or not ano_referencia_ncom:
                    flash(
                        "Preencha todos os campos obrigatórios da declaração de Não Comparecimento (NCOM).",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

                dados_personalizados.update(
                    {
                        "ano_serie_vaga": ano_serie_vaga,
                        "ano_referencia_ncom": ano_referencia_ncom,
                        "semestre_referencia_ncom": semestre_referencia_ncom,
                    }
                )

            declaracao_html = gerar_declaracao_personalizada(dados_personalizados)

            if declaracao_html is None:
                flash(
                    "Não foi possível gerar a declaração personalizada. Verifique os dados informados.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

            return declaracao_html

        # -------------------------------
        # FLUXO NORMAL: FUNDAMENTAL / EJA
        # -------------------------------
        segmento = request.form.get("segmento_escolhido")
        if segmento not in ("Fundamental", "EJA"):
            flash("Selecione se a declaração é do Fundamental ou EJA antes de gerar.", "error")
            return redirect(url_for("declaracao_tipo"))

        rm = (request.form.get("rm") or "").strip()
        tipo = (request.form.get("tipo") or "").strip()

        tipo_lower = tipo.lower()
        if tipo_lower in ("transferencia", "transferência"):
            tipo = "Transferencia"
        elif tipo_lower in ("conclusao", "conclusão"):
            tipo = "Conclusão"
        elif tipo_lower in ("frequencia", "frequência"):
            tipo = "Frequencia"

        deve_historico_str = request.form.get("deve_historico")

        unidade_select = (request.form.get("unidade_anterior_select") or "").strip()
        unidade_manual = (request.form.get("unidade_anterior_manual") or "").strip()
        unidade_anterior = unidade_select or unidade_manual

        file_path = None
        excel_file = request.files.get("excel_file")
        novo_upload = excel_file is not None and excel_file.filename

        if novo_upload:
            valid, message = validate_excel_upload(excel_file)
            if not valid:
                flash(message, "error")
                return redirect(url_for("declaracao_tipo", segmento=segmento))

            session_key = "lista_fundamental" if segmento == "Fundamental" else "lista_eja"
            file_path = save_excel_upload_to_session(
                excel_file,
                session,
                session_key,
                app.config["UPLOAD_FOLDER"],
                prefix="declaracao",
            )
        else:
            if segmento == "Fundamental":
                file_path = session.get("lista_fundamental")
            else:
                file_path = session.get("lista_eja")

        if not file_path or not os.path.exists(file_path):
            flash(
                "Nenhuma lista piloto encontrada para este segmento. Anexe o arquivo em Excel.",
                "error",
            )
            return redirect(url_for("declaracao_tipo", segmento=segmento))

        session["declaracao_tipo"] = segmento
        session["declaracao_excel"] = file_path

        if novo_upload and (not rm or not tipo):
            flash(
                "Lista piloto carregada com sucesso. Agora selecione o aluno e o tipo de declaração.",
                "success",
            )
            return redirect(url_for("declaracao_tipo", segmento=segmento))

        if not rm or not tipo:
            flash("Escolha o aluno e o tipo de declaração.", "error")
            return redirect(url_for("declaracao_tipo", segmento=segmento))

        if tipo in ("Transferencia", "Conclusão"):
            if deve_historico_str not in ("sim", "nao"):
                flash("Por favor, responda se o aluno deve o histórico escolar.", "error")
                return redirect(url_for("declaracao_tipo", segmento=segmento))

            if deve_historico_str == "sim" and not unidade_anterior:
                flash(
                    "Informe a unidade escolar anterior para a qual o aluno deve o histórico.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento=segmento))

            deve_historico = deve_historico_str == "sim"
        else:
            deve_historico = False
            unidade_anterior = ""

        dados_frequencia = None

        if tipo == "Frequencia":
            meses = [
                ("jan", "Janeiro"),
                ("fev", "Fevereiro"),
                ("mar", "Março"),
                ("abr", "Abril"),
                ("mai", "Maio"),
                ("jun", "Junho"),
                ("jul", "Julho"),
                ("ago", "Agosto"),
                ("set", "Setembro"),
                ("out", "Outubro"),
                ("nov", "Novembro"),
                ("dez", "Dezembro"),
            ]

            dados_frequencia = {"meses": []}
            algum_valido = False

            for mes_id, mes_nome in meses:
                dias_raw = (request.form.get(f"freq_{mes_id}_dias") or "").strip()
                faltas_raw = (request.form.get(f"freq_{mes_id}_faltas") or "").strip()

                if not dias_raw and not faltas_raw:
                    dados_frequencia["meses"].append(
                        {
                            "id": mes_id,
                            "nome": mes_nome,
                            "dias_letivos": None,
                            "faltas": None,
                            "frequencia": None,
                            "preenchido": False,
                        }
                    )
                    continue

                try:
                    dias = float(dias_raw.replace(",", ".")) if dias_raw else None
                    faltas = float(faltas_raw.replace(",", ".")) if faltas_raw else None
                except ValueError:
                    flash(
                        "Verifique os valores de dias letivos e faltas informados na frequência.",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento=segmento))

                if dias is None or faltas is None:
                    flash(
                        "Para cada mês de frequência preenchido, informe tanto os dias letivos quanto as faltas.",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento=segmento))

                if dias <= 0 or faltas < 0 or faltas > dias:
                    flash(
                        "Os valores de dias letivos e faltas são inválidos em um ou mais meses. "
                        "Verifique e tente novamente.",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento=segmento))

                freq_percent = ((dias - faltas) / dias) * 100.0
                algum_valido = True

                dados_frequencia["meses"].append(
                    {
                        "id": mes_id,
                        "nome": mes_nome,
                        "dias_letivos": dias,
                        "faltas": faltas,
                        "frequencia": round(freq_percent, 1),
                        "preenchido": True,
                    }
                )

            if not algum_valido:
                flash(
                    "Informe ao menos um mês de frequência com dias letivos e faltas válidos.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento=segmento))

        declaracao_html = gerar_declaracao_escolar(
            file_path=file_path,
            rm=rm,
            tipo=tipo,
            deve_historico=deve_historico,
            unidade_anterior=unidade_anterior,
            dados_frequencia=dados_frequencia,
        )

        if declaracao_html is None:
            flash("Aluno não encontrado na lista piloto.", "error")
            return redirect(url_for("declaracao_tipo", segmento=segmento))

        return declaracao_html

    # --------------------------------------
    # GET: EXIBE A TELA
    # --------------------------------------
    segmento = request.args.get("segmento")
    if segmento not in ("Fundamental", "EJA", "Personalizado"):
        segmento = None

    alunos = []
    tem_lista = False

    if segmento == "Fundamental":
        file_path = session.get("lista_fundamental")
        if file_path and os.path.exists(file_path):
            tem_lista = True
            session["declaracao_tipo"] = "Fundamental"
            session["declaracao_excel"] = file_path

            planilha = pd.read_excel(file_path, sheet_name="LISTA CORRIDA")

            planilha["RM_str"] = planilha["RM"].apply(format_rm)
            alunos_df = (
                planilha[planilha["RM_str"] != "0"][["RM_str", "NOME"]].drop_duplicates()
            )

            alunos = [
                {"rm": row["RM_str"], "nome": row["NOME"]}
                for _, row in alunos_df.iterrows()
            ]

    elif segmento == "EJA":
        # EJA permanece suportada, porém agora é opcional no sistema como um todo.
        file_path = session.get("lista_eja")
        if file_path and os.path.exists(file_path):
            tem_lista = True
            session["declaracao_tipo"] = "EJA"
            session["declaracao_excel"] = file_path

            df = pd.read_excel(file_path, sheet_name=0, header=None, skiprows=1)
            df["RM_str"] = df.iloc[:, 2].apply(format_eja_rm)
            df["NOME"] = df.iloc[:, 3]
            alunos_df = df[df["RM_str"] != ""][["RM_str", "NOME"]].drop_duplicates()

            alunos = [
                {"rm": row["RM_str"], "nome": row["NOME"]}
                for _, row in alunos_df.iterrows()
            ]
        else:
            tem_lista = False
            alunos = []

    dashboard_url = url_for("dashboard")
    conclusao_5ano_url = url_for("declaracao_conclusao_5ano")
    escolaridade_5ano_url = url_for("declaracao_escolaridade_5ano")

    return render_template(
        "declaracao_tipo.html",
        segmento=segmento,
        tem_lista=tem_lista,
        alunos=alunos,
        dashboard_url=dashboard_url,
        conclusao_5ano_url=conclusao_5ano_url,
        escolaridade_5ano_url=escolaridade_5ano_url,
    )


@app.route("/declaracao/escolaridade_5ano")
@login_required
def declaracao_escolaridade_5ano(file_path_arg=None):
    if file_path_arg:
        file_path = file_path_arg
    else:
        file_path = session.get("declaracao_excel")

    if session.get("declaracao_tipo") != "Fundamental":
        flash(
            "As declarações de escolaridade de 5º ano só podem ser geradas "
            "com a lista piloto do Ensino Fundamental.",
            "error",
        )
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    if not file_path or not os.path.exists(file_path):
        flash(
            "Nenhuma lista piloto do Ensino Fundamental está carregada. "
            "Anexe a lista piloto novamente para gerar as declarações.",
            "error",
        )
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    registros, data_extenso_str, titulo = gerar_lote_escolaridade_5ano(file_path)

    if not registros:
        flash(
            "Nenhum aluno de 5º ano foi encontrado na lista piloto para "
            "gerar as declarações de escolaridade.",
            "error",
        )
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    return render_template(
        "declaracao_escolaridade_5ano.html",
        registros=registros,
        data_extenso=data_extenso_str,
        titulo=titulo,
    )


# ==========================================================
#  UPLOAD DE FOTOS (CARTEIRINHAS)
# ==========================================================

@app.route("/upload_foto", methods=["POST"])
@login_required
def upload_foto():
    rm = (request.form.get("rm") or "").strip()
    if not rm:
        flash("RM não fornecido.", "error")
        return redirect(url_for("carteirinhas"))

    file = request.files.get("foto_file")
    try:
        save_student_photo(file, rm)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("carteirinhas"))

    flash("Foto anexada com sucesso.", "success")
    return redirect(url_for("carteirinhas"))


@app.route("/upload_multiplas_fotos", methods=["POST"])
@login_required
def upload_multiplas_fotos():
    rms = request.form.getlist("rm[]")
    files = request.files.getlist("foto_file[]")

    if not files:
        flash("Nenhuma foto enviada.", "error")
        return redirect(url_for("carteirinhas"))

    total_salvas = 0
    total_ignoradas = 0

    for rm, file in zip(rms, files):
        rm = (rm or "").strip()
        try:
            save_student_photo(file, rm)
        except ValueError:
            total_ignoradas += 1
            continue

        total_salvas += 1

    if total_salvas:
        msg = f"Foto(s) anexada(s) com sucesso: {total_salvas} arquivo(s)."
        if total_ignoradas:
            msg += f" {total_ignoradas} arquivo(s) foram ignorados por RM ou formato inválido."
        flash(msg, "success")
    else:
        flash("Nenhuma foto válida foi enviada.", "error")

    return redirect(url_for("carteirinhas"))


@app.route("/upload_inline_foto", methods=["POST"])
@login_required
def upload_inline_foto():
    file = request.files.get("foto_file")
    rm = (request.form.get("rm") or "").strip()

    if not rm:
        return jsonify({"error": "RM não fornecido"}), 400

    try:
        saved = save_student_photo(file, rm)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify({"url": saved["url"], "message": "Foto anexada com sucesso"}), 200


# ==========================================================
#  QUADROS – HELPERS COMUNS + EJA (LIGA/DESLIGA)
#  (FIX RÁPIDO E SEGURO: sem colisão de nomes, sem sobrescrita de helpers)
# ==========================================================

# ----------------------------------------------------------
# EJA: liga/desliga (FIX)
# ----------------------------------------------------------
def _is_eja_enabled() -> bool:
    """
    Desativada por padrão:
      - Ative via variável de ambiente: ENABLE_EJA=1
      - Ou via config: app.config["ENABLE_EJA"] = True

    FIX: remove duplicidade/override incorreto.
    """
    def _to_bool(v) -> bool:
        if v is None:
            return False
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        return s in ("1", "true", "t", "yes", "y", "on")

    # 1) prioridade para config do Flask
    cfg = None
    try:
        cfg = current_app.config.get("ENABLE_EJA", None)
    except Exception:
        cfg = None

    if cfg is not None:
        return _to_bool(cfg)

    # 2) fallback: env var
    return _to_bool(os.getenv("ENABLE_EJA", "0"))


def _save_upload_to_session(file_storage, session_key: str, prefix: str) -> str:
    """
    Salva arquivo enviado em UPLOAD_FOLDER e grava em session[session_key].
    Retorna o path salvo.
    """
    upload_folder = None
    try:
        upload_folder = current_app.config.get("UPLOAD_FOLDER")
    except Exception:
        upload_folder = None

    if not upload_folder:
        # mantém compatibilidade com seu app.py original (que define app.config["UPLOAD_FOLDER"])
        upload_folder = "uploads"

    return save_excel_upload_to_session(
        file_storage,
        session,
        session_key,
        upload_folder,
        prefix=prefix,
    )


def _find_sheet_case_insensitive(wb, target_name: str):
    """
    Busca uma aba ignorando maiúsculas/minúsculas e espaços.
    Retorna o nome real encontrado ou None.
    """
    target = (target_name or "").strip().lower()
    for name in wb.sheetnames:
        if (name or "").strip().lower() == target:
            return name
    return None


# ----------------------------------------------------------
# Helpers comuns: normalização de cabeçalho e strings
# ----------------------------------------------------------
@contextmanager
def _temp_unprotect_sheet(ws):
    """Desabilita proteção da planilha temporariamente para escrita e restaura depois."""
    original = copy.copy(ws.protection)
    try:
        ws.protection.sheet = False
        yield
    finally:
        ws.protection = original


# ==========================================================
#  QUADROS – MENU PRINCIPAL
# ==========================================================
@app.route("/quadros")
@login_required
def quadros():
    return render_template("quadros.html")


# ==========================================================
#  QUADRO – INCLUSÃO (DESATIVADO / ARQUIVADO)
# ==========================================================
@app.route("/quadros/inclusao", methods=["GET", "POST"])
@login_required
def quadros_inclusao():
    flash("Quadro de Inclusão foi migrado para Drive e está desativado no sistema no momento.", "info")
    return redirect(url_for("quadros"))


# ==========================================================
#  QUADRO – ATENDIMENTO MENSAL (CORRIGIDO / FUTURE-PROOF)
# ==========================================================

@app.route('/quadros/atendimento_mensal', methods=['GET', 'POST'])
@login_required
def quadro_atendimento_mensal():
    if request.method == 'POST':
        debug_log = []
        enable_eja = _is_eja_enabled()

        responsavel = (request.form.get("responsavel") or "").strip()
        rf = (request.form.get("rf") or "").strip()

        mes_ref = _normalize_atendimento_mes_ref(request.form.get("mes_ref"))

        fundamental_file = request.files.get('lista_fundamental')
        eja_file = request.files.get('lista_eja')

        try:
            if fundamental_file and fundamental_file.filename != '':
                _save_upload_to_session(fundamental_file, 'lista_fundamental', prefix='atendimento')

            if enable_eja and eja_file and eja_file.filename != '':
                _save_upload_to_session(eja_file, 'lista_eja', prefix='atendimento_eja')
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        file_path = session.get('lista_fundamental')
        if not file_path or not os.path.exists(file_path):
            flash("Nenhum arquivo da Lista Piloto FUNDAMENTAL disponível.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        model_path = os.path.join("modelos", "Quadro de Atendimento Mensal - Modelo.xlsx")
        if not os.path.exists(model_path):
            flash("Modelo Atendimento Mensal não encontrado.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        try:
            with open(model_path, "rb") as f:
                wb_modelo = load_workbook(f, data_only=False)
        except Exception as e:
            flash(f"Erro ao ler o modelo de atendimento mensal: {str(e)}", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        ws_modelo = wb_modelo.worksheets[1] if len(wb_modelo.worksheets) > 1 else wb_modelo.active

        _write_atendimento_header(ws_modelo, responsavel, rf, mes_ref, debug_log)

        try:
            wb_lista = load_workbook(file_path, data_only=True, read_only=True)
        except Exception:
            flash("Erro ao ler o arquivo da Lista Piloto FUNDAMENTAL.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        sheet_name = _find_sheet_case_insensitive(wb_lista, "Total de Alunos")
        if not sheet_name:
            flash("A aba 'Total de Alunos' não foi encontrada na Lista Piloto FUNDAMENTAL.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        ws_total = wb_lista[sheet_name]

        # 1º ano: SEMPRE ZERADO (Masc + Fem + Total)
        _write_atendimento_block(ws_modelo, "1º", {}, debug_log)

        # séries 2º..5º
        for serie_label in ["2º", "3º", "4º", "5º"]:
            data = _extract_atendimento_by_cols(ws_total, serie_label, debug_log)
            if not data:
                debug_log.append(f"[{serie_label}] prioridade 1 (C/D) não encontrou turmas; tentando fallback...")
                data = _extract_atendimento_by_fallback_block(ws_total, serie_label, debug_log)
            _write_atendimento_block(ws_modelo, serie_label, data, debug_log)

        # totais manhã / tarde
        _write_atendimento_turno_totals(ws_modelo, ws_total, debug_log)

        # EJA: mantém comportamento atual
        if not enable_eja:
            _zero_atendimento_eja_block(ws_modelo)
            debug_log.append("[EJA] desativada: bloco EJA zerado.")
        else:
            eja_path = session.get('lista_eja')
            if not eja_path or not os.path.exists(eja_path):
                _zero_atendimento_eja_block(ws_modelo)
                debug_log.append("[EJA] habilitada, mas sem arquivo: bloco EJA zerado.")
            else:
                try:
                    wb_eja = load_workbook(eja_path, data_only=True, read_only=True)
                except Exception as e:
                    flash(f"Erro ao ler a Lista Piloto EJA: {str(e)}. Gerando sem EJA.", "warning")
                    _zero_atendimento_eja_block(ws_modelo)
                    debug_log.append(f"[EJA] erro ao ler: {e}. Bloco EJA zerado.")
                else:
                    sheet_name_eja = _find_sheet_case_insensitive(wb_eja, "Total de Alunos")
                    if not sheet_name_eja:
                        flash("A aba 'Total de Alunos' não foi encontrada na Lista Piloto EJA. Gerando sem EJA.", "warning")
                        _zero_atendimento_eja_block(ws_modelo)
                        debug_log.append("[EJA] aba Total de Alunos não encontrada. Bloco EJA zerado.")
                    else:
                        ws_total_eja = wb_eja[sheet_name_eja]
                        _fill_atendimento_eja_block(ws_modelo, ws_total_eja)
                        debug_log.append("[EJA] preenchida com sucesso.")

        if ATENDIMENTO_CONFIG["ENABLE_DEBUG_LOG"]:
            try:
                current_app.logger.info("=== DEBUG ATENDIMENTO MENSAL ===\n%s", "\n".join(debug_log))
            except Exception:
                print("=== DEBUG ATENDIMENTO MENSAL ===")
                print("\n".join(debug_log))

        output = BytesIO()
        wb_modelo.save(output)
        output.seek(0)

        filename = f"Quadro de Atendimento Mensal - {datetime.now().strftime('%d%m')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return render_template('quadro_atendimento_mensal.html', enable_eja=_is_eja_enabled())


# ==========================================================
#  QUADRO – TRANSFERÊNCIAS (FIX H/I/J + ALERTAS DE FALTA DE INFORMAÇÃO)
# ==========================================================

@app.route("/quadros/transferencias", methods=["GET", "POST"])
@login_required
def quadro_transferencias():
    if request.method == "POST":
        enable_eja = _is_eja_enabled()

        period_start_str = request.form.get("period_start")
        period_end_str = request.form.get("period_end")

        responsavel = request.form.get("responsavel")
        diretor_nome = request.form.get("diretor_nome") or "Luciana Rocha Augustinho"
        data_quadro_in = request.form.get("data_quadro")  # opcional

        fundamental_file = request.files.get("lista_fundamental")
        eja_file = request.files.get("lista_eja")

        if not period_start_str or not period_end_str or not responsavel:
            flash("Por favor, preencha todos os campos.", "error")
            return redirect(url_for("quadro_transferencias"))

        # FUNDAMENTAL
        if fundamental_file and fundamental_file.filename != "":
            valid, message = validate_excel_upload(fundamental_file)
            if not valid:
                flash(message, "error")
                return redirect(url_for("quadro_transferencias"))

            upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
            fundamental_path = save_excel_upload_to_session(
                fundamental_file,
                session,
                "lista_fundamental",
                upload_folder,
                prefix="fundamental",
            )
        else:
            fundamental_path = session.get("lista_fundamental")
            if not fundamental_path or not os.path.exists(fundamental_path):
                flash("Lista Piloto Fundamental não encontrada.", "error")
                return redirect(url_for("quadro_transferencias"))

        # EJA opcional
        eja_path = None
        if enable_eja and eja_file and eja_file.filename != "":
            valid, message = validate_excel_upload(eja_file)
            if not valid:
                flash(f"Lista EJA inválida: {message}", "error")
                return redirect(url_for("quadro_transferencias"))

            upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
            eja_path = save_excel_upload_to_session(
                eja_file,
                session,
                "lista_eja",
                upload_folder,
                prefix="eja",
            )
        elif enable_eja:
            eja_path = session.get("lista_eja")

        try:
            period_start = _parse_period_date(period_start_str, "a data inicial")
            period_end = _parse_period_date(period_end_str, "a data final")
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for("quadro_transferencias"))

        if period_end < period_start:
            flash("A data final não pode ser menor que a data inicial.", "error")
            return redirect(url_for("quadro_transferencias"))

        data_quadro_dt = _parse_user_date(data_quadro_in) or datetime.now()

        # ALERTAS de inconsistência que NÃO impedem gerar o arquivo
        missing_info_alerts = []
        missing_info_seen = set()

        # Lê LISTA CORRIDA (Fundamental)
        try:
            df_fundamental = pd.read_excel(fundamental_path, sheet_name="LISTA CORRIDA")
        except Exception as e:
            flash(f"Erro ao ler a Lista Piloto Fundamental: {str(e)}", "error")
            return redirect(url_for("quadro_transferencias"))

        colmap = _build_colmap(df_fundamental)

        col_serie = _pick_col(colmap, "SÉRIE", "SERIE")
        col_nome = _pick_col(colmap, "NOME")
        col_dn = _pick_col(colmap, "DATA NASC.", "DATA NASC", "DATANASC")
        col_ra = _pick_col(colmap, "RA")
        col_obs = _pick_col(colmap, "OBS", "OBSERVACAO", "OBSERVAÇÃO")
        col_local_te = _pick_col(colmap, "LOCAL TE", "LOCALTE")

        if not col_nome or not col_dn or not col_ra or not col_serie or not col_obs:
            flash("A aba 'LISTA CORRIDA' não contém cabeçalhos essenciais (SÉRIE, NOME, DATA NASC., RA, OBS).", "error")
            return redirect(url_for("quadro_transferencias"))

        debug = []
        debug.append("[quadro_transferencias] Aba lida: LISTA CORRIDA (Fundamental)")
        debug.append(
            f"[quadro_transferencias] Colunas detectadas: "
            f"SÉRIE='{col_serie}', NOME='{col_nome}', DATA NASC.='{col_dn}', RA='{col_ra}', "
            f"OBS='{col_obs}', LOCAL TE='{col_local_te}'"
        )

        if not col_local_te:
            _push_missing_info_alert(
                missing_info_alerts,
                missing_info_seen,
                turma="(Estrutura do arquivo)",
                nome="-",
                ra="-",
                tipo="TE",
                data_str="-",
                campo="LOCAL TE",
                detalhe="A coluna 'LOCAL TE' não foi encontrada na Lista Piloto Fundamental."
            )

        transfer_records = []
        invalid_te_dates = 0

        use_cols = [col_serie, col_nome, col_dn, col_ra, col_obs]
        if col_local_te and col_local_te not in use_cols:
            use_cols.append(col_local_te)

        df_sub = df_fundamental[use_cols].copy()

        for row in df_sub.itertuples(index=False, name=None):
            row_dict = dict(zip(df_sub.columns, row))

            te_dt, te_match_txt, _ = _extract_te_date_from_text(
                row_dict.get(col_obs), period_start, period_end
            )

            if te_match_txt and not te_dt:
                invalid_te_dates += 1
                continue

            if not te_dt:
                continue

            if not (period_start <= te_dt <= period_end):
                continue

            nome = _safe_str(row_dict.get(col_nome))

            dn_val = row_dict.get(col_dn)
            dn_str = ""
            if pd.notna(dn_val):
                try:
                    dn_dt = pd.to_datetime(dn_val, errors="coerce")
                    dn_str = dn_dt.strftime("%d/%m/%Y") if pd.notna(dn_dt) else ""
                except Exception:
                    dn_str = ""

            ra = _safe_str(row_dict.get(col_ra))
            nivel_classe = _safe_str(row_dict.get(col_serie))

            # FIX H/I/J
            local_te_raw = _safe_str(row_dict.get(col_local_te)) if col_local_te else ""
            local_te = "-" if _is_missing_text(local_te_raw) else local_te_raw

            # ALERTA: registro encontrado, mas sem LOCAL TE
            if _is_missing_text(local_te_raw):
                _push_missing_info_alert(
                    missing_info_alerts,
                    missing_info_seen,
                    turma=nivel_classe,
                    nome=nome,
                    ra=ra,
                    tipo="TE",
                    data_str=te_dt.strftime("%d/%m/%Y"),
                    campo="LOCAL TE",
                    detalhe="Registro encontrado no período, mas o campo LOCAL TE está vazio ou inválido."
                )

            record = {
                "nome": nome,
                "dn": dn_str,
                "ra": ra,
                "situacao": "Parcial",
                "breda": "Não",
                "nivel_classe": nivel_classe,
                "tipo": "TE",
                "observacao": local_te,              # H
                "remanejamento": "-",                # I
                "data": te_dt.strftime("%d/%m/%Y"),  # J
            }
            transfer_records.append(record)

        debug.append(f"[quadro_transferencias] TE válidos no período: {len(transfer_records)}")
        debug.append(f"[quadro_transferencias] Datas TE inválidas descartadas: {invalid_te_dates}")

        # EJA opcional (mantido)
        if enable_eja and eja_path and os.path.exists(eja_path):
            try:
                df_eja = pd.read_excel(eja_path, sheet_name="LISTA CORRIDA")
            except Exception as e:
                flash(f"Erro ao ler a Lista Piloto EJA: {str(e)}", "error")
                return redirect(url_for("quadro_transferencias"))

            colmap_eja = _build_colmap(df_eja)
            eja_col_nome = _pick_col(colmap_eja, "NOME")
            eja_col_dn = _pick_col(colmap_eja, "DATA NASC.", "DATA NASC")
            eja_col_ra = _pick_col(colmap_eja, "RA")
            eja_col_serie = _pick_col(colmap_eja, "SÉRIE", "SERIE")
            eja_col_obs = _pick_col(colmap_eja, "OBS")
            eja_col_local_te = _pick_col(colmap_eja, "LOCAL TE", "LOCALTE")

            if not eja_col_local_te:
                _push_missing_info_alert(
                    missing_info_alerts,
                    missing_info_seen,
                    turma="(Estrutura do arquivo EJA)",
                    nome="-",
                    ra="-",
                    tipo="TE/MC/MCC",
                    data_str="-",
                    campo="LOCAL TE",
                    detalhe="A coluna 'LOCAL TE' não foi encontrada na Lista Piloto EJA."
                )

            if eja_col_nome and eja_col_ra and eja_col_serie and eja_col_obs:
                df_eja_sub = df_eja[
                    [c for c in [eja_col_serie, eja_col_nome, eja_col_dn, eja_col_ra, eja_col_obs, eja_col_local_te] if c]
                ].copy()

                for row in df_eja_sub.itertuples(index=False, name=None):
                    row_dict = dict(zip(df_eja_sub.columns, row))
                    txt = _safe_str(row_dict.get(eja_col_obs))

                    m = _RX_EJA.search(txt)
                    if not m:
                        continue

                    tipo_str = m.group(1).upper()
                    day = int(m.group(2))
                    month = int(m.group(3))
                    year_raw = m.group(4)

                    if year_raw:
                        y = int(year_raw)
                        if y < 100:
                            y += 2000
                    else:
                        y = period_start.year

                    try:
                        dt = datetime(y, month, day)
                    except Exception:
                        continue

                    if not (period_start <= dt <= period_end):
                        continue

                    nome = _safe_str(row_dict.get(eja_col_nome))
                    ra = _safe_str(row_dict.get(eja_col_ra))
                    nivel_classe = _safe_str(row_dict.get(eja_col_serie))

                    dn_str = ""
                    if eja_col_dn:
                        dn_val = row_dict.get(eja_col_dn)
                        if pd.notna(dn_val):
                            try:
                                dn_dt = pd.to_datetime(dn_val, errors="coerce")
                                dn_str = dn_dt.strftime("%d/%m/%Y") if pd.notna(dn_dt) else ""
                            except Exception:
                                dn_str = ""

                    local_te_raw = _safe_str(row_dict.get(eja_col_local_te)) if eja_col_local_te else ""
                    local_te = "-" if _is_missing_text(local_te_raw) else local_te_raw

                    # ALERTA: registro encontrado, mas sem LOCAL TE
                    if _is_missing_text(local_te_raw):
                        _push_missing_info_alert(
                            missing_info_alerts,
                            missing_info_seen,
                            turma=nivel_classe,
                            nome=nome,
                            ra=ra,
                            tipo=tipo_str,
                            data_str=dt.strftime("%d/%m/%Y"),
                            campo="LOCAL TE",
                            detalhe="Registro encontrado no período, mas o campo LOCAL TE está vazio ou inválido."
                        )

                    transfer_records.append({
                        "nome": nome,
                        "dn": dn_str,
                        "ra": ra,
                        "situacao": "Parcial",
                        "breda": "Não",
                        "nivel_classe": nivel_classe,
                        "tipo": tipo_str,
                        "observacao": local_te,           # H
                        "remanejamento": "-",             # I
                        "data": dt.strftime("%d/%m/%Y"),  # J
                    })

        if not transfer_records:
            flash("Nenhum registro de TE/MC/MCC encontrado no período especificado.", "error")
            try:
                current_app.logger.info("\n".join(debug))
            except Exception:
                print("\n".join(debug))
            return redirect(url_for("quadro_transferencias"))

        model_path = os.path.join("modelos", "Quadro Informativo - Modelo.xlsx")
        if not os.path.exists(model_path):
            flash("Modelo de Quadro Informativo (Transferências) não encontrado.", "error")
            return redirect(url_for("quadro_transferencias"))

        try:
            with open(model_path, "rb") as f:
                wb = load_workbook(f, data_only=False)
        except Exception as e:
            flash(f"Erro ao ler o modelo: {str(e)}", "error")
            return redirect(url_for("quadro_transferencias"))

        ws = wb.active

        SCHOOL_NAME = "E.M José Padin Mouta"
        _label_set(ws, "A7", "Unidade Escolar", SCHOOL_NAME)
        _label_set(ws, "A8", "Diretor(a)", diretor_nome)

        set_merged_cell_value(ws, "B9", responsavel)
        set_merged_cell_value(ws, "J9", data_quadro_dt.strftime("%d/%m/%Y"))

        current_row = 12
        for record in transfer_records:
            set_merged_cell_value(ws, f"A{current_row}", record["nome"])
            set_merged_cell_value(ws, f"B{current_row}", record["dn"])
            set_merged_cell_value(ws, f"C{current_row}", record["ra"])
            set_merged_cell_value(ws, f"D{current_row}", record["situacao"])
            set_merged_cell_value(ws, f"E{current_row}", record["breda"])
            set_merged_cell_value(ws, f"F{current_row}", record["nivel_classe"])
            set_merged_cell_value(ws, f"G{current_row}", record["tipo"])

            # FIX H/I/J
            set_merged_cell_value(ws, f"H{current_row}", record["observacao"])
            set_merged_cell_value(ws, f"I{current_row}", "-")
            set_merged_cell_value(ws, f"J{current_row}", record["data"])

            current_row += 1

        debug.append(f"[quadro_transferencias] Linhas preenchidas no modelo: {len(transfer_records)} (início A12)")
        debug.append(f"[quadro_transferencias] Alertas de falta de informação: {len(missing_info_alerts)}")

        try:
            current_app.logger.info("\n".join(debug))
        except Exception:
            print("\n".join(debug))

        _add_transfer_alerts_sheet(wb, missing_info_alerts)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Quadro_de_Transferencias_{period_start.strftime('%d%m')}_{period_end.strftime('%d%m')}.xlsx"
        resp = send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Dados pessoais ficam na aba ALERTAS do Excel, não em headers HTTP.
        if missing_info_alerts:
            resp.headers["X-Transferencias-MissingInfo-Count"] = str(len(missing_info_alerts))

        return resp

    return render_template("quadro_transferencias.html")


# ==========================================================
#  QUADRO – QUANTITATIVO MENSAL (Fundamental)
#  (mantido: parse flexível do período + debug sheet oculta)
# ==========================================================

def _recreate_debug_sheet_hidden(wb, title: str = "DEBUG_TE"):
    """Cria/zera uma aba de debug (oculta) no workbook."""
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws_dbg = wb.create_sheet(title)
    ws_dbg.sheet_state = "hidden"
    ws_dbg.append(
        [
            "LINHA_ARQUIVO",
            "RM",
            "NOME",
            "SERIE",
            "OBS_ORIGINAL",
            "TE_DATA_EXTRAIDA",
            "ANO_INFERIDO",
            "TRECHO_MATCH",
            "STATUS",
            "MOTIVO",
            "TIPO_TE_RAW",
            "TIPO_TE_NORMALIZADO",
        ]
    )
    return ws_dbg


@app.route("/quadros/quantitativo_mensal", methods=["GET", "POST"])
@login_required
def quadro_quantitativo_mensal():
    if request.method == "POST":
        period_start_str = request.form.get("period_start")
        period_end_str = request.form.get("period_end")
        responsavel = request.form.get("responsavel")
        mes_ano = request.form.get("mes_ano")

        if not responsavel or not str(responsavel).strip():
            flash("Preencha o campo Responsável.", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        default_year = datetime.now().year

        try:
            period_start = parse_date_flexible(period_start_str, default_year=default_year, field_label="a data inicial")
            period_end = parse_date_flexible(period_end_str, default_year=default_year, field_label="a data final")
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        if period_end < period_start:
            flash("A data final não pode ser menor que a data inicial.", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        if not mes_ano or not str(mes_ano).strip():
            meses = {
                1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
                5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
                9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
            }
            mes_ano = f"{meses[datetime.now().month]}/{datetime.now().year}"
        mes_ano = str(mes_ano).strip()

        fundamental_file = request.files.get("lista_fundamental")
        if fundamental_file and fundamental_file.filename:
            valid, message = validate_excel_upload(fundamental_file)
            if not valid:
                flash(message, "error")
                return redirect(url_for("quadro_quantitativo_mensal"))

            upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
            fundamental_path = save_excel_upload_to_session(
                fundamental_file,
                session,
                "lista_fundamental",
                upload_folder,
                prefix="fundamental",
            )
        else:
            fundamental_path = session.get("lista_fundamental")
            if not fundamental_path or not os.path.exists(fundamental_path):
                flash("Arquivo da Lista Piloto Fundamental não encontrado.", "error")
                return redirect(url_for("quadro_quantitativo_mensal"))

        try:
            df = pd.read_excel(fundamental_path, sheet_name="LISTA CORRIDA")
        except Exception as e:
            flash(f"Erro ao ler a Lista Piloto Fundamental: {str(e)}", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        col_rm = _find_df_col(df, ["RM"])
        col_nome = _find_df_col(df, ["NOME"])
        col_serie = _find_df_col(df, ["SÉRIE", "SERIE"])
        col_obs = _find_df_col(df, ["OBS"])
        col_tipo_te = _find_df_col(df, ["TIPO TE", "TIPO_TE", "TIPO  TE"])

        if not col_serie or not col_obs:
            flash("Não foi possível localizar as colunas essenciais (SÉRIE e/ou OBS) na LISTA CORRIDA.", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        model_path = os.path.join("modelos", "Quadro Quantitativo Mensal - Modelo.xlsx")
        if not os.path.exists(model_path):
            flash("Modelo de Quadro Quantitativo Mensal não encontrado.", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        try:
            with open(model_path, "rb") as f:
                wb = load_workbook(f, data_only=False)
        except Exception as e:
            flash(f"Erro ao ler o modelo: {str(e)}", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        ws = wb.active

        mapping = {
            "2º": {
                "Dentro da Rede": "K14",
                "Rede Estadual": "K15",
                "Litoral": "K16",
                "Mudança de Municipio": "K16",
                "São Paulo": "K17",
                "ABCD": "K18",
                "Interior": "K19",
                "Outros Estados": "K20",
                "Particular": "K21",
                "País": "K22",
                "Sem Informação": "K23",
            },
            "3º": {
                "Dentro da Rede": "L14",
                "Rede Estadual": "L15",
                "Litoral": "L16",
                "Mudança de Municipio": "L16",
                "São Paulo": "L17",
                "ABCD": "L18",
                "Interior": "L19",
                "Outros Estados": "L20",
                "Particular": "L21",
                "País": "L22",
                "Sem Informação": "L23",
            },
            "4º": {
                "Dentro da Rede": "M14",
                "Rede Estadual": "M15",
                "Litoral": "M16",
                "Mudança de Municipio": "M16",
                "São Paulo": "M17",
                "ABCD": "M18",
                "Interior": "M19",
                "Outros Estados": "M20",
                "Particular": "M21",
                "País": "M22",
                "Sem Informação": "M23",
            },
            "5º": {
                "Dentro da Rede": "N14",
                "Rede Estadual": "N15",
                "Litoral": "N16",
                "Mudança de Municipio": "N16",
                "São Paulo": "N17",
                "ABCD": "N18",
                "Interior": "N19",
                "Outros Estados": "N20",
                "Particular": "N21",
                "País": "N22",
                "Sem Informação": "N23",
            },
        }

        # Zera determinísticamente todas as células-alvo
        for tipos in mapping.values():
            for cell_addr in tipos.values():
                set_merged_cell_value(ws, cell_addr, 0)

        ws_dbg = _recreate_debug_sheet_hidden(wb, "DEBUG_TE")

        counted = 0
        discarded = 0

        for i, row in df.iterrows():
            linha_arquivo = int(i) + 2

            rm = row.get(col_rm) if col_rm else None
            nome = row.get(col_nome) if col_nome else None
            serie_val = row.get(col_serie, "")
            obs_val = row.get(col_obs, "")

            if _is_missing_value(obs_val):
                continue

            te_dt, rule, match_txt, year_inferred = detect_te_date_from_obs_flexible(
                obs_val,
                default_year=default_year,
            )

            if not match_txt:
                continue

            if not te_dt:
                discarded += 1
                ws_dbg.append(
                    [
                        linha_arquivo,
                        "" if rm is None else str(rm),
                        "" if nome is None else str(nome),
                        "" if serie_val is None else str(serie_val),
                        str(obs_val).strip(),
                        "",
                        "SIM" if year_inferred else "NAO",
                        match_txt,
                        "SKIPPED",
                        "Data TE inválida em OBS",
                        "" if col_tipo_te is None else _safe_str(row.get(col_tipo_te)),
                        "" if col_tipo_te is None else _normalize_tipo_te(row.get(col_tipo_te)),
                    ]
                )
                continue

            if not (period_start <= te_dt <= period_end):
                discarded += 1
                ws_dbg.append(
                    [
                        linha_arquivo,
                        "" if rm is None else str(rm),
                        "" if nome is None else str(nome),
                        "" if serie_val is None else str(serie_val),
                        str(obs_val).strip(),
                        te_dt.strftime("%d/%m/%Y"),
                        "SIM" if year_inferred else "NAO",
                        match_txt,
                        "SKIPPED",
                        "Fora do período informado",
                        "" if col_tipo_te is None else _safe_str(row.get(col_tipo_te)),
                        "" if col_tipo_te is None else _normalize_tipo_te(row.get(col_tipo_te)),
                    ]
                )
                continue

            serie_key = _serie_key_from_value(serie_val)
            if not serie_key or serie_key not in mapping:
                discarded += 1
                ws_dbg.append(
                    [
                        linha_arquivo,
                        "" if rm is None else str(rm),
                        "" if nome is None else str(nome),
                        "" if serie_val is None else str(serie_val),
                        str(obs_val).strip(),
                        te_dt.strftime("%d/%m/%Y"),
                        "SIM" if year_inferred else "NAO",
                        match_txt,
                        "SKIPPED",
                        "Série fora de 2º–5º ou ilegível",
                        "" if col_tipo_te is None else _safe_str(row.get(col_tipo_te)),
                        "" if col_tipo_te is None else _normalize_tipo_te(row.get(col_tipo_te)),
                    ]
                )
                continue

            tipo_raw = row.get(col_tipo_te, None) if col_tipo_te else None
            tipo_te = _normalize_tipo_te(tipo_raw)
            if tipo_te not in mapping[serie_key]:
                tipo_te = "Sem Informação"

            cell_addr = mapping[serie_key][tipo_te]
            current_val = ws[cell_addr].value
            current_val = current_val if isinstance(current_val, (int, float)) else 0
            set_merged_cell_value(ws, cell_addr, current_val + 1)

            counted += 1
            ws_dbg.append(
                [
                    linha_arquivo,
                    "" if rm is None else str(rm),
                    "" if nome is None else str(nome),
                    "" if serie_val is None else str(serie_val),
                    str(obs_val).strip(),
                    te_dt.strftime("%d/%m/%Y"),
                    "SIM" if year_inferred else "NAO",
                    match_txt,
                    "COUNTED",
                    "",
                    "" if tipo_raw is None else str(tipo_raw),
                    tipo_te,
                ]
            )

        set_merged_cell_value(ws, "B3", str(responsavel).strip())
        set_merged_cell_value(ws, "D3", f"{period_start.strftime('%d/%m/%Y')} a {period_end.strftime('%d/%m/%Y')}")

        with _temp_unprotect_sheet(ws):
            set_merged_cell_value(ws, "A6", "E.M José Padin Mouta")
            set_merged_cell_value(ws, "A8", mes_ano)
            ano_letivo = int(current_app.config.get("SCHOOL_YEAR", datetime.now().year))
            set_merged_cell_value(ws, "A10", f"QUADRO GERAL DE TRANSFERENCIAS EXPEDIDAS - {ano_letivo}")

        current_app.logger.info(
            "[QUADRO_QUANTITATIVO] periodo=%s..%s | counted=%s | discarded=%s | ano_padrao_sem_ano=%s",
            period_start.strftime("%d/%m/%Y"),
            period_end.strftime("%d/%m/%Y"),
            counted,
            discarded,
            default_year,
        )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Quadro_Quantitativo_Fundamental_{period_start.strftime('%d%m')}_{period_end.strftime('%d%m')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    meses = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
    }
    default_mes_ano = f"{meses[datetime.now().month]}/{datetime.now().year}"
    return render_template("quadro_quantitativo_mensal.html", default_mes_ano=default_mes_ano)

# ==========================================================
#  QUADRO QUANTITATIVO DE INCLUSÃO – REGULAR (EJA DESCONSIDERADA)
#  - Contagem por turma via LISTA CORRIDA
#  - Regras:
#      * Turma: Coluna A
#      * RM (identificador único): Coluna C (NÃO contabiliza vazio/0)
#      * Matrícula Ativa: Coluna H deve conter "MA"
#      * Inclusão: Coluna N == "Sim" (case/trim)
#      * Plano de Ação: Coluna P com profissional válido,
#                       MAS SOMENTE se o aluno também for Inclusão
#      * Profissionais: únicos por turma (dedupe case-insensitive e colapsando espaços)
#  - Alertas:
#      * Turmas com 2+ profissionais distintos (Plano de Ação)
#      * Aluno com profissional na P, mas sem Inclusão na N
#  - Preenchimento por mapeamento automático do MODELO
# ==========================================================

# OBS: este código assume que você já tem:
# - app = Flask(__name__)
# - app.config["UPLOAD_FOLDER"]
# no seu arquivo principal.


@app.route("/quantinclusao", methods=["GET", "POST"])
@login_required
def quantinclusao():
    if request.method == "POST":
        reg_file = (
            request.files.get("lista_regular")
            or request.files.get("lista_fundamental")
            or request.files.get("lista")
        )
        responsavel = request.form.get("responsavel")

        if not reg_file or reg_file.filename == "":
            flash("Selecione o arquivo da Lista Piloto (Regular/Fundamental).", "error")
            return redirect(url_for("quantinclusao"))

        valid, message = validate_excel_upload(reg_file)
        if not valid:
            flash(message, "error")
            return redirect(url_for("quantinclusao"))

        if not responsavel or responsavel.strip() == "":
            flash("Informe o Responsável pelo preenchimento.", "error")
            return redirect(url_for("quantinclusao"))

        reg_path = save_excel_upload(
            reg_file,
            app.config["UPLOAD_FOLDER"],
            prefix="regular",
        )

        try:
            wb_reg = load_workbook(reg_path, data_only=True, read_only=True)
            ws_lista_reg = wb_reg["LISTA CORRIDA"]
        except Exception as e:
            flash(f"Erro ao ler o arquivo: {str(e)}", "error")
            return redirect(url_for("quantinclusao"))

        model_path = os.path.join("modelos", "Quadro Quantitativo de Inclusão - Modelo.xlsx")
        try:
            wb_model = load_workbook(model_path, data_only=False)
            ws_model = wb_model.active
        except Exception as e:
            flash(f"Erro ao abrir o modelo de inclusão: {str(e)}", "error")
            return redirect(url_for("quantinclusao"))

        template_map = _build_quant_template_map(ws_model)
        valid_turmas = set(template_map.keys())

        (
            inc_counts,
            plano_counts,
            profs_by_turma,
            plan_without_inclusion_by_turma,
        ) = _collect_quant_counts_from_lista_corrida(ws_lista_reg, valid_turmas)

        # Preenche as 3 linhas por turma: Inclusão / Plano / Profissionais
        for turma, cells in template_map.items():
            inc = inc_counts.get(turma, 0)
            plano = plano_counts.get(turma, 0)
            profs = len(profs_by_turma.get(turma, {}))

            ws_model[cells["inc_qtd"]] = inc
            ws_model[cells["plano_qtd"]] = plano
            ws_model[cells["prof_qtd"]] = profs

        # Cabeçalho
        meses = {
            1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
            5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
            9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO",
        }
        now = datetime.now()
        mes_ano = f"{meses[now.month]}/{now.year}"

        try:
            b4 = ws_model["B4"].value or ""
            b4s = str(b4)
            if re.search(r"MÊS\s*/\s*\d{4}", b4s, flags=re.IGNORECASE):
                ws_model["B4"] = re.sub(r"MÊS\s*/\s*\d{4}", mes_ano, b4s, flags=re.IGNORECASE)
            else:
                ws_model["B4"] = b4s if mes_ano in b4s else f"{b4s} - {mes_ano}".strip(" -")
        except Exception:
            pass

        ws_model["C8"] = responsavel.strip()
        ws_model["K8"] = now.strftime("%d/%m/%Y")

        alerts = _build_quant_multi_prof_alerts(profs_by_turma, valid_turmas)
        plan_without_inclusion_alerts = _build_quant_plan_without_inclusion_alerts(
            plan_without_inclusion_by_turma,
            valid_turmas,
        )

        _add_quant_inclusao_alerts_sheet(wb_model, alerts, plan_without_inclusion_alerts)

        # Gera arquivo
        output = BytesIO()
        wb_model.save(output)
        output.seek(0)

        filename = f"Quadro_Quantitativo_de_Inclusao_{now.strftime('%d%m%Y')}.xlsx"
        resp = send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Dados pessoais ficam na aba ALERTAS do Excel, não em headers HTTP.
        if alerts:
            resp.headers["X-QuantInclusao-Alerts-Count"] = str(len(alerts))

        if plan_without_inclusion_alerts:
            resp.headers["X-QuantInclusao-PlanWithoutInclusion-Count"] = str(len(plan_without_inclusion_alerts))

        return resp

    return render_template("quantinclusao.html")

def build_deadline_alerts(today=None):
    return build_deadline_alerts_service(
        today=today,
        holidays_path=app.config.get("HOLIDAYS_JSON_PATH"),
        weekly_due_weekday=app.config.get("INFORMATIVO_WEEKDAY_DUE", 4),
    )


# Injeta automaticamente para TODOS os templates
@app.context_processor
def _inject_deadline_alerts():
    return {"deadline_alerts": build_deadline_alerts()}

# ==========================================================
#  MAIN
# ==========================================================

if __name__ == "__main__":
    app.run(debug=True)
